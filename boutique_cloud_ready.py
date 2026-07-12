import streamlit as st
from pymongo import MongoClient
import pandas as pd
from datetime import date, datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import os
import time
import hmac
import hashlib
import secrets
import uuid
import re
import json
import base64
from html import escape as html_escape
from dotenv import load_dotenv

load_dotenv("credentials/.env")

BRAND_NAME = "Shree Krishna Boutique"
BRAND_SHORT_NAME = "Shree Krishna"
BILL_SYMBOL_PATH = os.path.join(os.path.dirname(__file__), "krishna_symbol.png")

# =====================================================
# PASSWORD HASH UTILITY
# Run once in a Python shell to generate your hash:
#
#   import bcrypt
#   h = bcrypt.hashpw(b"your_password_here", bcrypt.gensalt())
#   print(h.decode())
#
# Then set PASSWORD_HASH=<output> in credentials/.env
# or in your Streamlit secrets.toml as:
#   PASSWORD_HASH = "<output>"
# =====================================================

# =====================================================
# PAGE CONFIG
# =====================================================

st.set_page_config(
    page_title=BRAND_NAME,
    page_icon="◆",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =====================================================
# CSS — DARK NAVY BLUE THEME
# =====================================================

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,wght@0,300;0,400;0,500;0,600;1,400&family=DM+Serif+Display:ital@0;1&display=swap');

:root {
    /* Light mode palette */
    --bg:           #F4F7FC;
    --bg-2:         #EBF0FA;
    --surface:      #FFFFFF;
    --surface-2:    #F8FAFF;
    --blue:         #2563EB;
    --blue-soft:    #3B82F6;
    --blue-pale:    #BFDBFE;
    --blue-glow:    rgba(37,99,235,0.12);
    --text:         #0F172A;
    --text-2:       #334155;
    --muted:        #64748B;
    --dim:          #94A3B8;
    --emerald:      #059669;
    --rose:         #DC2626;
    --amber:        #D97706;
    --r:            10px;
    --r-lg:         14px;
    --r-xl:         22px;
    --border:       rgba(37,99,235,0.12);
    --border-hover: rgba(37,99,235,0.35);
    --shadow-sm:    0 1px 3px rgba(15,23,42,0.06), 0 1px 2px rgba(15,23,42,0.04);
    --shadow:       0 4px 16px rgba(15,23,42,0.08), 0 2px 6px rgba(15,23,42,0.05);
    --shadow-lg:    0 12px 40px rgba(37,99,235,0.14), 0 4px 12px rgba(15,23,42,0.06);
    /* Legacy aliases kept so existing rules don't break */
    --navy-1:       #F4F7FC;
    --navy-2:       #EBF0FA;
    --navy-3:       #FFFFFF;
    --navy-4:       #F0F5FF;
    --navy-5:       #E2EAFF;
    --cream:        #0F172A;
    --cream-dim:    #334155;
}

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif !important;
    background: var(--bg) !important;
    color: var(--text) !important;
}

/* ━━━ ALL INPUTS — DARK TEXT FOR LIGHT MODE ━━━ */
input:not([type="radio"]):not([type="checkbox"]),
textarea,
[data-baseweb="input"] input,
[data-baseweb="base-input"] input,
[data-baseweb="textarea"] textarea,
[data-baseweb="date-picker"] input,
[data-baseweb="select"] input {
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
    caret-color: var(--blue) !important;
    background-color: var(--surface) !important;
}
[data-baseweb="input"] *,
[data-baseweb="base-input"] *,
[data-baseweb="textarea"] * {
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
}
/* Selectbox displayed value */
[data-baseweb="select"] > div > div,
[data-baseweb="select"] > div > div > div,
[data-baseweb="select"] span,
[class*="ValueContainer"] > div,
[class*="singleValue"] {
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
}
/* Number stepper */
.stNumberInput div[data-baseweb="input"] input,
.stNumberInput input[type="number"] {
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
}
/* Date displayed value */
.stDateInput div[data-baseweb="input"] input,
.stDateInput input[type="text"] {
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
}
.stApp {
    background: var(--bg) !important;
    background-image:
        radial-gradient(ellipse 900px 600px at 0% 0%, rgba(37,99,235,0.04) 0%, transparent 70%),
        radial-gradient(ellipse 700px 500px at 100% 100%, rgba(37,99,235,0.03) 0%, transparent 70%);
    background-attachment: fixed;
}
::-webkit-scrollbar { width: 4px; height: 4px; }
::-webkit-scrollbar-track { background: var(--bg-2); border-radius: 99px; }
::-webkit-scrollbar-thumb { background: rgba(37,99,235,0.3); border-radius: 99px; }
::-webkit-scrollbar-thumb:hover { background: rgba(37,99,235,0.5); }

h1, h2, h3 {
    font-family: 'DM Serif Display', serif !important;
    color: var(--text) !important;
    letter-spacing: -0.01em;
    font-weight: 400 !important;
}
h4, h5, h6 {
    font-family: 'DM Sans', sans-serif !important;
    color: var(--muted) !important;
    text-transform: uppercase !important;
    letter-spacing: 0.1em !important;
    font-size: 0.7rem !important;
    font-weight: 600 !important;
}

.page-title {
    font-family: 'DM Serif Display', serif;
    font-size: 2.1rem;
    font-weight: 400;
    color: var(--text);
    letter-spacing: -0.02em;
    line-height: 1.15;
    margin-bottom: 0.2rem;
}
.page-sub {
    font-size: 0.7rem;
    text-transform: uppercase;
    letter-spacing: 0.18em;
    color: var(--dim);
    margin-bottom: 2.4rem;
    font-weight: 500;
}
.settings-top-pad { height: 0.35rem; }
.settings-meta {
    font-size: 0.78rem;
    color: var(--muted);
    line-height: 1.45;
    padding: 0.2rem 0 0.45rem;
}
.settings-status {
    border: 1px solid var(--border);
    border-radius: var(--r);
    background: var(--surface-2);
    padding: 0.55rem 0.7rem;
    margin-bottom: 0.65rem;
    font-size: 0.78rem;
    color: var(--text-2);
}
.rule { height:1px; background:linear-gradient(90deg, var(--blue) 0%, rgba(37,99,235,0.15) 60%, transparent 100%); margin:2rem 0; border:none; }
.rule-sm { height:1px; background:linear-gradient(90deg, rgba(37,99,235,0.25), transparent); margin:1.2rem 0; border:none; }

/* ━━━ SIDEBAR ━━━ */
[data-testid="stSidebar"] {
    background: var(--surface) !important;
    border-right: 1px solid var(--border) !important;
    box-shadow: 2px 0 12px rgba(15,23,42,0.05) !important;
}
[data-testid="stSidebar"] * { color: var(--text) !important; }
.sb-brand { padding: 2rem 1.5rem 0.5rem; text-align: center; }
.sb-logo {
    font-family: 'DM Serif Display', serif;
    font-size: 1.75rem;
    font-weight: 400;
    color: var(--blue);
    letter-spacing: -0.01em;
    line-height: 1;
}
.sb-mark {
    font-size: 0.6rem;
    text-transform: uppercase;
    letter-spacing: 0.22em;
    color: var(--muted);
    margin-top: 0.3rem;
    font-weight: 600;
}
[data-testid="stSidebar"] .stRadio > div {
    gap: 2px !important;
    flex-direction: column !important;
}
[data-testid="stSidebar"] .stRadio > div > label {
    background: transparent !important;
    border: none !important;
    border-radius: var(--r) !important;
    color: var(--muted) !important;
    font-size: 0.84rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.01em !important;
    padding: 0.55rem 1rem 0.55rem 0.75rem !important;
    transition: all 0.18s ease !important;
    cursor: pointer;
    display: flex;
    align-items: center;
}
[data-testid="stSidebar"] .stRadio > div > label:hover {
    background: var(--blue-glow) !important;
    color: var(--blue) !important;
}
[data-testid="stSidebar"] .stRadio > div > label > div:first-child {
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    width: 16px !important;
    height: 16px !important;
    min-width: 16px !important;
    border: 2px solid var(--dim) !important;
    border-radius: 50% !important;
    margin-right: 8px !important;
    background: transparent !important;
    transition: all 0.18s !important;
}
.sb-user {
    font-size: 0.75rem;
    color: var(--muted);
    text-align: center;
    padding: 0.6rem 0 1.5rem;
    letter-spacing: 0.04em;
    font-weight: 500;
}
.sb-sep { height: 1px; background: var(--border); margin: 0.8rem 1rem; }

/* ━━━ PUBLIC BANNER ━━━ */
.pub-banner {
    background: linear-gradient(135deg, #FFFFFF 0%, #F0F6FF 100%);
    border: 1px solid var(--border);
    border-radius: var(--r-xl);
    padding: 2.2rem 2.8rem 1.8rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
    box-shadow: var(--shadow);
}
.pub-banner::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, var(--blue), var(--blue-soft), transparent);
    border-radius: var(--r-xl) var(--r-xl) 0 0;
}
.pub-banner-title {
    font-family: 'DM Serif Display', serif;
    font-size: 2rem;
    font-weight: 400;
    color: var(--text);
    letter-spacing: -0.02em;
    line-height: 1.1;
    margin-bottom: 0.3rem;
}
.pub-banner-sub {
    font-size: 0.68rem;
    text-transform: uppercase;
    letter-spacing: 0.22em;
    color: var(--muted);
    font-weight: 600;
}

/* ━━━ ADMIN LOGIN PANEL (bottom) ━━━ */
.admin-strip {
    margin-top: 3rem;
    border-top: 1px solid var(--border);
    padding-top: 1.5rem;
}
.admin-strip-label {
    font-size: 0.62rem;
    text-transform: uppercase;
    letter-spacing: 0.22em;
    color: var(--dim);
    text-align: center;
    margin-bottom: 0.8rem;
    font-weight: 600;
}

/* ━━━ METRICS ━━━ */
[data-testid="stMetric"] {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--r-lg) !important;
    padding: 1.25rem 1.4rem !important;
    position: relative;
    overflow: hidden;
    box-shadow: var(--shadow-sm) !important;
    transition: box-shadow 0.2s ease, transform 0.2s ease !important;
}
[data-testid="stMetric"]::after {
    content: '';
    position: absolute;
    bottom: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, var(--blue), var(--blue-soft), transparent);
    opacity: 0.6;
}
[data-testid="stMetric"]:hover {
    box-shadow: var(--shadow) !important;
    transform: translateY(-2px);
}
[data-testid="stMetricLabel"] > div {
    color: var(--muted) !important;
    font-size: 0.68rem !important;
    text-transform: uppercase !important;
    letter-spacing: 0.12em !important;
    font-weight: 600 !important;
    font-family: 'DM Sans', sans-serif !important;
}
[data-testid="stMetricValue"] {
    color: var(--text) !important;
    font-family: 'DM Serif Display', serif !important;
    font-size: 1.7rem !important;
    font-weight: 400 !important;
    letter-spacing: -0.02em !important;
    line-height: 1.2 !important;
}

/* ━━━ BUTTONS ━━━ */
.stButton > button {
    background: var(--surface) !important;
    color: var(--blue) !important;
    border: 1.5px solid rgba(37,99,235,0.28) !important;
    border-radius: var(--r) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.8rem !important;
    letter-spacing: 0.04em !important;
    padding: 0.6rem 1.4rem !important;
    transition: all 0.18s ease !important;
    box-shadow: var(--shadow-sm) !important;
    width: 100% !important;
}
.stButton > button:hover {
    background: var(--blue) !important;
    border-color: var(--blue) !important;
    color: #FFFFFF !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 16px rgba(37,99,235,0.25) !important;
}
.stButton > button:active { transform: scale(0.98) !important; }

.stDownloadButton > button {
    background: var(--surface-2) !important;
    color: var(--muted) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--r) !important;
    font-size: 0.78rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.04em !important;
    transition: all 0.18s ease !important;
    width: 100% !important;
}
.stDownloadButton > button:hover {
    border-color: var(--border-hover) !important;
    color: var(--blue) !important;
    background: var(--blue-glow) !important;
}

/* ━━━ FORM SUBMIT ━━━ */
.stForm button[type="submit"] {
    background: var(--blue) !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: var(--r) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.84rem !important;
    letter-spacing: 0.04em !important;
    padding: 0.85rem 2.5rem !important;
    width: 100% !important;
    transition: all 0.2s ease !important;
    box-shadow: 0 2px 8px rgba(37,99,235,0.2) !important;
}
.stForm button[type="submit"]:hover {
    background: #1D4ED8 !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 24px rgba(37,99,235,0.35) !important;
}

/* ━━━ INPUTS ━━━ */
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stTextArea > div > div > textarea,
.stDateInput > div > div > input,
.stDateInput input,
input[type="text"], input[type="number"], input[type="date"], textarea {
    background: var(--surface) !important;
    border: 1.5px solid rgba(37,99,235,0.18) !important;
    border-radius: var(--r) !important;
    color: var(--text) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.88rem !important;
    font-weight: 400 !important;
    padding: 0.6rem 0.9rem !important;
    transition: border-color 0.18s ease, box-shadow 0.18s ease !important;
    -webkit-text-fill-color: var(--text) !important;
}
.stTextInput > div > div > input:focus,
.stNumberInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus,
.stDateInput > div > div > input:focus {
    border-color: var(--blue) !important;
    box-shadow: 0 0 0 3px rgba(37,99,235,0.1) !important;
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
}
.stTextInput > div > div > input::placeholder,
.stTextArea > div > div > textarea::placeholder,
.stDateInput > div > div > input::placeholder { color: var(--dim) !important; -webkit-text-fill-color: var(--dim) !important; }

[data-testid="InputInstructions"] {
    display: none !important;
}

/* Date picker calendar popup */
[data-baseweb="calendar"] {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    box-shadow: var(--shadow-lg) !important;
    border-radius: var(--r-lg) !important;
}
[data-baseweb="calendar"] * { color: var(--text) !important; background: transparent !important; }
[data-baseweb="calendar"] [aria-selected="true"] { background: var(--blue) !important; color: #FFFFFF !important; }
[data-baseweb="calendar"] button:hover { background: var(--bg-2) !important; }

/* Date input wrapper */
.stDateInput > div {
    background: var(--surface) !important;
    border-radius: var(--r) !important;
}
.stDateInput > div > div { background: var(--surface) !important; }
.stDateInput svg { fill: var(--muted) !important; }

/* Selectbox */
.stSelectbox > div > div,
.stSelectbox [data-baseweb="select"] > div {
    background: var(--surface) !important;
    border: 1.5px solid rgba(37,99,235,0.18) !important;
    border-radius: var(--r) !important;
    color: var(--text) !important;
    transition: border-color 0.18s ease !important;
}
.stSelectbox > div > div:hover,
.stSelectbox [data-baseweb="select"] > div:hover { border-color: var(--border-hover) !important; }
.stSelectbox [data-baseweb="select"] span,
.stSelectbox [data-baseweb="select"] div { color: var(--text) !important; }

/* Selectbox dropdown menu */
[data-baseweb="popover"] [data-baseweb="menu"],
[data-baseweb="popover"] ul {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    box-shadow: var(--shadow-lg) !important;
    border-radius: var(--r-lg) !important;
}
[data-baseweb="popover"] li,
[data-baseweb="popover"] [role="option"] {
    background: var(--surface) !important;
    color: var(--text) !important;
}
[data-baseweb="popover"] li:hover,
[data-baseweb="popover"] [role="option"]:hover { background: var(--bg-2) !important; }

/* Number input */
.stNumberInput > div > div { background: var(--surface) !important; border-radius: var(--r) !important; }
.stNumberInput input { color: var(--text) !important; -webkit-text-fill-color: var(--text) !important; }

/* ━━━ LABELS ━━━ */
.stTextInput label, .stNumberInput label, .stSelectbox label,
.stTextArea label, .stDateInput label, .stRadio label, .stCheckbox label,
.stTextInput label p, .stNumberInput label p, .stSelectbox label p,
.stTextArea label p, .stDateInput label p, .stRadio label p,
[data-testid="stWidgetLabel"], [data-testid="stWidgetLabel"] p,
[data-testid="stWidgetLabel"] span {
    color: var(--text-2) !important;
    -webkit-text-fill-color: var(--text-2) !important;
    font-size: 0.75rem !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.1em !important;
    font-family: 'DM Sans', sans-serif !important;
}

/* ━━━ DATAFRAME ━━━ */
.stDataFrame {
    border-radius: var(--r-lg) !important;
    border: 1px solid var(--border) !important;
    overflow: hidden !important;
    box-shadow: var(--shadow-sm) !important;
}
[data-testid="stDataFrame"] th {
    background: var(--bg-2) !important;
    color: var(--muted) !important;
    font-size: 0.68rem !important;
    text-transform: uppercase !important;
    letter-spacing: 0.1em !important;
    font-weight: 700 !important;
    border-bottom: 1px solid var(--border) !important;
    padding: 0.75rem 1rem !important;
}
[data-testid="stDataFrame"] td {
    background: var(--surface) !important;
    color: var(--text-2) !important;
    font-size: 0.85rem !important;
    font-weight: 400 !important;
    border-bottom: 1px solid rgba(37,99,235,0.06) !important;
    padding: 0.7rem 1rem !important;
}
[data-testid="stDataFrame"] tr:hover td { background: var(--bg-2) !important; }

/* ━━━ TABS ━━━ */
.stTabs [data-baseweb="tab-list"] {
    background: transparent !important;
    border-bottom: 2px solid var(--border) !important;
    border-radius: 0 !important;
    padding: 0 !important;
    gap: 0 !important;
}
.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    border-radius: 0 !important;
    color: var(--dim) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.8rem !important;
    letter-spacing: 0.04em !important;
    padding: 0.75rem 1.25rem !important;
    border-bottom: 2px solid transparent !important;
    transition: all 0.18s !important;
    margin-bottom: -2px !important;
}
.stTabs [aria-selected="true"] {
    color: var(--blue) !important;
    border-bottom-color: var(--blue) !important;
    font-weight: 600 !important;
}
.stTabs [data-baseweb="tab"]:hover { color: var(--text) !important; }

/* ━━━ ALERTS ━━━ */
.stSuccess {
    background: rgba(5,150,105,0.07) !important;
    border: 1px solid rgba(5,150,105,0.25) !important;
    border-radius: var(--r) !important;
}
.stSuccess * { color: #065f46 !important; }
.stInfo {
    background: rgba(37,99,235,0.06) !important;
    border: 1px solid rgba(37,99,235,0.2) !important;
    border-radius: var(--r) !important;
}
.stInfo * { color: #1e3a8a !important; }
.stWarning {
    background: rgba(217,119,6,0.07) !important;
    border: 1px solid rgba(217,119,6,0.25) !important;
    border-radius: var(--r) !important;
}
.stWarning * { color: #92400e !important; }
.stError {
    background: rgba(220,38,38,0.06) !important;
    border: 1px solid rgba(220,38,38,0.22) !important;
    border-radius: var(--r) !important;
}
.stError * { color: #991b1b !important; }

/* ━━━ RADIO (inline) ━━━ */
.stRadio > div { gap: 0.5rem !important; flex-direction: row !important; }
.stRadio > div > label {
    background: var(--surface) !important;
    border: 1.5px solid var(--border) !important;
    border-radius: var(--r) !important;
    padding: 0.5rem 1.1rem !important;
    color: var(--muted) !important;
    font-size: 0.8rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.02em !important;
    transition: all 0.18s !important;
    cursor: pointer;
    box-shadow: var(--shadow-sm) !important;
}
.stRadio > div > label:hover {
    border-color: var(--blue) !important;
    color: var(--blue) !important;
    background: var(--blue-glow) !important;
}

/* ━━━ SECTION HEADERS ━━━ */
.sec-head {
    font-family: 'DM Serif Display', serif;
    font-size: 1.05rem;
    font-weight: 400;
    font-style: italic;
    color: var(--text);
    margin: 1.8rem 0 1rem;
    padding-bottom: 0.6rem;
    border-bottom: 1px solid var(--border);
    letter-spacing: -0.01em;
}

/* ━━━ BADGES ━━━ */
.badge {
    display: inline-block;
    font-size: 0.68rem;
    font-weight: 700;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    padding: 0.22rem 0.65rem;
    border-radius: 6px;
}
.badge-gold  { background: rgba(37,99,235,0.1); color: var(--blue); }
.badge-green { background: rgba(5,150,105,0.1); color: #065f46; }
.badge-red   { background: rgba(220,38,38,0.09); color: #991b1b; }
.badge-muted { background: var(--bg-2); color: var(--muted); }

/* ━━━ PAYMENT COLLECTION ━━━ */
.pay-card {
    background: linear-gradient(135deg, rgba(255,255,255,0.98), rgba(248,250,255,0.96));
    border: 1.5px solid rgba(37,99,235,0.16);
    border-left: 4px solid var(--blue);
    border-radius: var(--r-lg);
    box-shadow: var(--shadow-sm);
    padding: 1rem 1.15rem;
    margin: 0.85rem 0 0.35rem;
}
.pay-grid {
    display: grid;
    grid-template-columns: minmax(240px, 2fr) minmax(150px, 1fr) minmax(140px, 1fr) minmax(140px, 1fr);
    gap: 1rem;
    align-items: center;
}
.pay-name {
    color: var(--text);
    font-size: 0.95rem;
    font-weight: 700;
}
.pay-meta {
    color: var(--muted);
    font-size: 0.78rem;
    margin-top: 0.2rem;
}
.pay-amount {
    color: var(--text);
    font-size: 1rem;
    font-weight: 700;
}
.pay-label {
    color: var(--dim);
    font-size: 0.66rem;
    font-weight: 700;
    letter-spacing: 0.12em;
    text-transform: uppercase;
}
.pay-form-note {
    color: var(--muted);
    font-size: 0.78rem;
    margin: 0.15rem 0 0.75rem;
}
@media (max-width: 760px) {
    .pay-grid { grid-template-columns: 1fr; gap: 0.55rem; }
}

/* ━━━ EMPTY STATE ━━━ */
.empty { text-align: center; padding: 4rem 2rem; color: var(--dim); }
.empty-glyph { font-size: 2rem; margin-bottom: 1rem; color: var(--border); }

/* Number input spinners */
button[data-testid="stNumberInputStepDown"],
button[data-testid="stNumberInputStepUp"] {
    background: var(--bg-2) !important;
    border-color: var(--border) !important;
    color: var(--muted) !important;
}

/* Expanders */
.streamlit-expanderHeader {
    background: var(--surface-2) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--r) !important;
    color: var(--text-2) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.84rem !important;
    font-weight: 500 !important;
}
.streamlit-expanderContent {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-top: none !important;
    border-radius: 0 0 var(--r) var(--r) !important;
    padding: 1rem !important;
}

/* ━━━ BACKUP & RESTORE CARDS ━━━ */
.bk-card {
    background: var(--surface);
    border: 1.5px solid var(--border);
    border-radius: var(--r-xl);
    padding: 1.8rem 2rem;
    box-shadow: var(--shadow);
    transition: box-shadow 0.2s ease, transform 0.2s ease;
    animation: fadeSlideUp 0.4s ease forwards;
    opacity: 0;
}
.bk-card:hover { box-shadow: var(--shadow-lg); transform: translateY(-2px); }
@keyframes fadeSlideUp {
    from { opacity: 0; transform: translateY(16px); }
    to   { opacity: 1; transform: translateY(0);   }
}
.bk-card-icon {
    width: 42px; height: 42px;
    border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-size: 1.25rem;
    margin-bottom: 1rem;
}
.bk-icon-blue  { background: rgba(37,99,235,0.1); }
.bk-icon-green { background: rgba(5,150,105,0.1); }
.bk-icon-amber { background: rgba(217,119,6,0.1); }
.bk-card-title {
    font-family: 'DM Sans', sans-serif;
    font-size: 0.95rem;
    font-weight: 700;
    color: var(--text);
    margin-bottom: 0.2rem;
    letter-spacing: -0.01em;
}
.bk-card-desc {
    font-size: 0.78rem;
    color: var(--muted);
    margin-bottom: 1.2rem;
    line-height: 1.5;
}
.bk-status-badge {
    display: inline-flex; align-items: center; gap: 0.4rem;
    font-size: 0.7rem; font-weight: 600;
    text-transform: uppercase; letter-spacing: 0.1em;
    padding: 0.28rem 0.7rem;
    border-radius: 999px;
}
.bk-status-ok    { background: rgba(5,150,105,0.1); color: #065f46; }
.bk-status-warn  { background: rgba(217,119,6,0.1);  color: #92400e; }
.bk-status-info  { background: rgba(37,99,235,0.1);  color: #1e3a8a; }
.bk-header {
    background: linear-gradient(135deg, #FFFFFF 0%, #EBF3FF 100%);
    border: 1.5px solid var(--border);
    border-radius: var(--r-xl);
    padding: 2rem 2.4rem;
    margin-bottom: 1.8rem;
    position: relative;
    overflow: hidden;
    box-shadow: var(--shadow);
}
.bk-header::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, var(--blue), var(--blue-soft));
    border-radius: var(--r-xl) var(--r-xl) 0 0;
}
.bk-header-icon {
    font-size: 2rem; margin-bottom: 0.6rem; display: block;
}
.bk-header-title {
    font-family: 'DM Serif Display', serif;
    font-size: 1.75rem; font-weight: 400; color: var(--text);
    letter-spacing: -0.02em; line-height: 1.1; margin-bottom: 0.3rem;
}
.bk-header-sub {
    font-size: 0.72rem; text-transform: uppercase; letter-spacing: 0.16em;
    color: var(--muted); font-weight: 600;
}
.bk-progress-bar {
    height: 6px; border-radius: 999px;
    background: linear-gradient(90deg, var(--blue), var(--blue-soft));
    animation: progressPulse 1.5s ease-in-out infinite alternate;
}
@keyframes progressPulse {
    from { opacity: 0.6; width: 30%; }
    to   { opacity: 1;   width: 85%; }
}
.bk-ts {
    font-size: 0.72rem; color: var(--muted);
    font-family: 'DM Sans', sans-serif; margin-top: 0.5rem;
}

/* File uploader */
[data-testid="stFileUploader"] > div {
    background: var(--surface-2) !important;
    border: 2px dashed rgba(37,99,235,0.25) !important;
    border-radius: var(--r-lg) !important;
    transition: border-color 0.18s !important;
}
[data-testid="stFileUploader"] > div:hover {
    border-color: var(--blue) !important;
    background: var(--blue-glow) !important;
}
[data-testid="stFileUploader"] * { color: var(--muted) !important; }

.stCaption, .stCaption * { color: var(--dim) !important; }

p, span { color: var(--text-2); }
</style>
""", unsafe_allow_html=True)

# =====================================================
# PLOTLY TEMPLATE
# =====================================================


# ── THEME TOGGLE — pure session_state, no JS needed ─────────────────────────
if "theme" not in st.session_state:
    st.session_state.theme = "light"

_DARK_CSS = """
<style>
/* ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   DARK MODE OVERRIDE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ */
:root {
    --bg:           #070C18 !important;
    --bg-2:         #0B1221 !important;
    --surface:      #0F1A2E !important;
    --surface-2:    #152238 !important;
    --blue:         #4D8AE8 !important;
    --blue-soft:    #6BA3F0 !important;
    --blue-pale:    #A8C4F0 !important;
    --blue-glow:    rgba(77,138,232,0.15) !important;
    --text:         #E8EEF8 !important;
    --text-2:       #C8D4E8 !important;
    --muted:        #6A84A8 !important;
    --dim:          #3D5478 !important;
    --border:       rgba(77,138,232,0.18) !important;
    --border-hover: rgba(77,138,232,0.42) !important;
    --shadow-sm:    0 1px 6px rgba(0,0,0,0.4) !important;
    --shadow:       0 4px 24px rgba(0,0,0,0.5) !important;
    --shadow-lg:    0 8px 48px rgba(0,0,0,0.65) !important;
}
.stApp, .stApp > div, .main, [data-testid="stAppViewContainer"], [data-testid="block-container"] {
    background: #070C18 !important;
}
[data-testid="stSidebar"], [data-testid="stSidebar"] > div { background: #0B1221 !important; }
[data-testid="stSidebar"] * { color: #E8EEF8 !important; -webkit-text-fill-color: #E8EEF8 !important; }
p, span, div, label { color: #C8D4E8 !important; }
h1, h2, h3 { color: #E8EEF8 !important; }
[data-testid="stMetric"] { background: #0F1A2E !important; }
[data-testid="stMetricValue"] { color: #E8EEF8 !important; -webkit-text-fill-color: #E8EEF8 !important; }
input, textarea, [data-baseweb="input"] input, [data-baseweb="base-input"] input {
    background: #0F1A2E !important; color: #E8EEF8 !important; -webkit-text-fill-color: #E8EEF8 !important;
}
.stSelectbox > div > div, .stSelectbox [data-baseweb="select"] > div, [data-baseweb="select"] > div { background: #0F1A2E !important; }
[data-baseweb="select"] span, [data-baseweb="select"] div, [class*="singleValue"] { color: #E8EEF8 !important; -webkit-text-fill-color: #E8EEF8 !important; }
[data-baseweb="popover"] [data-baseweb="menu"], [data-baseweb="popover"] ul { background: #0F1A2E !important; }
[data-baseweb="popover"] li, [data-baseweb="popover"] [role="option"] { background: #0F1A2E !important; color: #E8EEF8 !important; }
[data-baseweb="popover"] li:hover, [data-baseweb="popover"] [role="option"]:hover { background: #1C2D47 !important; }
[data-baseweb="calendar"] { background: #0F1A2E !important; }
[data-baseweb="calendar"] * { color: #E8EEF8 !important; }
.stDateInput > div, .stDateInput > div > div { background: #0F1A2E !important; }
[data-testid="stDataFrame"] th { background: #0F1A2E !important; color: #6A84A8 !important; }
[data-testid="stDataFrame"] td { background: #0B1221 !important; color: #C8D4E8 !important; }
[data-testid="stDataFrame"] tr:hover td { background: #0F1A2E !important; }
.stButton > button { background: transparent !important; color: #4D8AE8 !important; border-color: rgba(77,138,232,0.42) !important; }
.stButton > button:hover { background: rgba(77,138,232,0.14) !important; color: #A8C4F0 !important; }
.stDownloadButton > button { background: transparent !important; color: #6A84A8 !important; }
.stTabs [data-baseweb="tab"] { color: #3D5478 !important; }
.stTabs [aria-selected="true"] { color: #E8EEF8 !important; }
.streamlit-expanderHeader { background: #0F1A2E !important; color: #6A84A8 !important; border-color: rgba(77,138,232,0.18) !important; }
.streamlit-expanderContent { background: #0B1221 !important; border-color: rgba(77,138,232,0.18) !important; }
.stRadio > div > label { background: #0F1A2E !important; color: #6A84A8 !important; border-color: rgba(77,138,232,0.18) !important; }
.pub-banner { background: linear-gradient(135deg, #0B1221 0%, #0F1A2E 100%) !important; }
.pub-banner-title { color: #E8EEF8 !important; }
.bk-card { background: #0F1A2E !important; }
.bk-header { background: linear-gradient(135deg, #0B1221 0%, #0F1A2E 100%) !important; }
[data-testid="stWidgetLabel"], [data-testid="stWidgetLabel"] p, [data-testid="stWidgetLabel"] span { color: #C8D4E8 !important; -webkit-text-fill-color: #C8D4E8 !important; }
.stSuccess { background: rgba(61,154,108,0.12) !important; } .stSuccess * { color: #7ADFA0 !important; }
.stInfo { background: rgba(77,138,232,0.1) !important; } .stInfo * { color: #A8C4F0 !important; }
.stWarning { background: rgba(200,160,50,0.1) !important; } .stWarning * { color: #E8C840 !important; }
.stError { background: rgba(192,80,96,0.1) !important; } .stError * { color: #E08090 !important; }
[data-testid="stFileUploader"] > div { background: #0F1A2E !important; border-color: rgba(77,138,232,0.3) !important; }
</style>
"""

def inject_theme():
    if st.session_state.theme == "dark":
        st.markdown(_DARK_CSS, unsafe_allow_html=True)

PLOT_LAYOUT = dict(
    paper_bgcolor="rgba(255,255,255,0)",
    plot_bgcolor="rgba(255,255,255,0)",
    font=dict(family="DM Sans", color="#64748B", size=11),
    title=dict(font=dict(family="DM Serif Display", size=17, color="#0F172A"), pad=dict(b=12), x=0),
    xaxis=dict(gridcolor="rgba(37,99,235,0.08)", linecolor="rgba(37,99,235,0.15)", tickfont=dict(size=10, color="#64748B"), showgrid=True, zeroline=False),
    yaxis=dict(gridcolor="rgba(37,99,235,0.08)", linecolor="rgba(37,99,235,0.15)", tickfont=dict(size=10, color="#64748B"), showgrid=True, zeroline=False),
    legend=dict(bgcolor="rgba(255,255,255,0.92)", bordercolor="rgba(37,99,235,0.15)", borderwidth=1, font=dict(color="#64748B", size=10)),
    margin=dict(l=12, r=12, t=44, b=12),
    colorway=["#2563EB","#3B82F6","#059669","#8BACD8","#DC2626","#1D4ED8","#10B981","#0EA5E9"],
    hoverlabel=dict(bgcolor="rgba(255,255,255,0.97)", bordercolor="rgba(37,99,235,0.3)", font=dict(color="#0F172A", size=11, family="DM Sans"), align="left"),
    bargap=0.35,
)

_DARK_PLOT_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="DM Sans", color="#3D5478", size=11),
    title=dict(font=dict(family="DM Serif Display", size=17, color="#C8D4E8"), pad=dict(b=12), x=0),
    xaxis=dict(gridcolor="rgba(77,138,232,0.08)", linecolor="rgba(77,138,232,0.15)", tickfont=dict(size=10, color="#3D5478"), showgrid=True, zeroline=False),
    yaxis=dict(gridcolor="rgba(77,138,232,0.08)", linecolor="rgba(77,138,232,0.15)", tickfont=dict(size=10, color="#3D5478"), showgrid=True, zeroline=False),
    legend=dict(bgcolor="rgba(7,12,24,0.88)", bordercolor="rgba(77,138,232,0.2)", borderwidth=1, font=dict(color="#6A84A8", size=10)),
    margin=dict(l=12, r=12, t=44, b=12),
    colorway=["#4D8AE8","#6BA3F0","#7ADFA0","#8BACD8","#E08090","#1A3D80","#3D9A6C","#4A9AC8"],
    hoverlabel=dict(bgcolor="rgba(7,12,24,0.96)", bordercolor="rgba(77,138,232,0.3)", font=dict(color="#E8EEF8", size=11, family="DM Sans"), align="left"),
    bargap=0.35,
)

def get_plot_layout():
    """Return plotly layout dict adjusted for current theme."""
    if st.session_state.get("theme") == "dark":
        return _DARK_PLOT_LAYOUT
    return PLOT_LAYOUT

def styled_fig(fig, height=340):
    fig.update_layout(**get_plot_layout(), height=height)
    return fig

# =====================================================
# CONSTANTS
# =====================================================

CATEGORIES      = ["Sarees","Salwar Suits","Lehengas","Kurtis","Western Wear","Accessories","Kids Wear","Blouse","Fabric","Other"]
PAYMENT_METHODS = ["Cash","UPI","Card","Bank Transfer","Part Payment","Credit"]
PAYMENT_COLLECTION_METHODS = ["Cash", "UPI", "Bank Transfer", "Card"]
BILL_SCOPE_ALL = "All Transactions"
BILL_SCOPE_LAST = "Last Transactions"
BILL_SCOPE_PENDING = "Pending Transactions"
BILL_SCOPE_OPTIONS = [BILL_SCOPE_ALL, BILL_SCOPE_LAST, BILL_SCOPE_PENDING]
STATE_OPTIONS   = ["Tamil Nadu","Maharashtra","Karnataka","Delhi","Gujarat","Rajasthan","West Bengal","Uttar Pradesh","Andhra Pradesh","Telangana","Other"]
VENDOR_MANUAL_OPTION = "Add new vendor..."
ADMIN_NAV_OPTIONS = [
    "Dashboard",
    "Add Sale",
    "Review Accounts",
    "Update Transaction",
    "Customer List",
    "Vendor List",
    "Analytics",
    "Reminders & Alerts",
    "Generate Bill",
    "Passbook Reader",
    "Work Notes",
    "AI Assistant",
    "Technical",
    "Security & Devices",
    "Backup & Restore",
    "Logout",
]
MEMBER_NAV_OPTIONS = [
    "Add Sale",
    "Review Accounts",
    "Customer List",
    "Vendor List",
    "Generate Bill",
    "Logout",
]
APP_SETTING_DEFAULTS = {
    "theme": "light",
    "default_page": "Dashboard",
    "default_payment_method": "UPI",
    "default_bill_scope": BILL_SCOPE_ALL,
    "default_review_days": 90,
    "default_inventory_category": "All",
    "ai_provider": "auto",
    "gemini_model": "",
    "openai_model": "",
}
MANAGED_SECRET_KEYS = [
    "GEMINI_API_KEY",
    "GOOGLE_API_KEY",
    "GOOGLE_AI_API_KEY",
    "GOOGLE_GENERATIVE_AI_API_KEY",
    "GENAI_API_KEY",
    "GEMINI_KEY",
    "GEMINI_MODEL",
    "AI_PROVIDER",
    "OPENAI_API_KEY",
    "OPENAI_MODEL",
    "MONGO_URI",
    "USERNAME",
    "PASSWORD",
    "PASSWORD_HASH",
]
BOOTSTRAP_SECRET_KEYS = {"MONGO_URI", "USERNAME", "PASSWORD", "PASSWORD_HASH"}

# =====================================================
# MONGODB
# =====================================================

def safe_secret(key: str, default: str = ""):
    try:
        return st.secrets.get(key, default)
    except Exception:
        return default

def get_gemini_key() -> str:
    try:
        managed = (
            get_managed_secret("GEMINI_API_KEY")
            or get_managed_secret("GOOGLE_API_KEY")
            or get_managed_secret("GOOGLE_AI_API_KEY")
            or get_managed_secret("GOOGLE_GENERATIVE_AI_API_KEY")
            or get_managed_secret("GENAI_API_KEY")
            or get_managed_secret("GEMINI_KEY")
        )
        if managed:
            return managed
    except Exception:
        pass
    return (
        safe_secret("GEMINI_API_KEY", "")
        or safe_secret("GOOGLE_API_KEY", "")
        or safe_secret("GOOGLE_AI_API_KEY", "")
        or safe_secret("GOOGLE_GENERATIVE_AI_API_KEY", "")
        or safe_secret("GENAI_API_KEY", "")
        or safe_secret("GEMINI_KEY", "")
        or safe_secret("gemini_api_key", "")
        or os.getenv("GEMINI_API_KEY", "")
        or os.getenv("GOOGLE_API_KEY", "")
        or os.getenv("GOOGLE_AI_API_KEY", "")
        or os.getenv("GOOGLE_GENERATIVE_AI_API_KEY", "")
        or os.getenv("GENAI_API_KEY", "")
        or os.getenv("GEMINI_KEY", "")
    )

def get_gemini_model() -> str:
    override = str(st.session_state.get("gemini_model", "") or "").strip()
    if override:
        return override
    try:
        managed = get_managed_secret("GEMINI_MODEL")
        if managed:
            return managed
    except Exception:
        pass
    return safe_secret("GEMINI_MODEL", "") or os.getenv("GEMINI_MODEL", "gemini-2.5-flash")

def get_openai_key() -> str:
    try:
        managed = get_managed_secret("OPENAI_API_KEY")
        if managed:
            return managed
    except Exception:
        pass
    return safe_secret("OPENAI_API_KEY", "") or os.getenv("OPENAI_API_KEY", "")

def get_openai_model() -> str:
    override = str(st.session_state.get("openai_model", "") or "").strip()
    if override:
        return override
    try:
        managed = get_managed_secret("OPENAI_MODEL")
        if managed:
            return managed
    except Exception:
        pass
    return safe_secret("OPENAI_MODEL", "") or os.getenv("OPENAI_MODEL", "gpt-4o-mini")

def get_ai_provider() -> str:
    override = str(st.session_state.get("ai_provider", "") or "").strip().lower()
    if override and override != "auto":
        return override
    try:
        managed = get_managed_secret("AI_PROVIDER").strip().lower()
        if managed:
            return managed
    except Exception:
        pass
    configured = (safe_secret("AI_PROVIDER", "") or os.getenv("AI_PROVIDER", "")).strip().lower()
    if configured:
        return configured
    if get_gemini_key():
        return "gemini"
    if get_openai_key():
        return "openai"
    return "gemini"

def llm_is_configured() -> bool:
    provider = get_ai_provider()
    if provider in {"gemini", "google"}:
        return bool(get_gemini_key())
    if provider == "openai":
        return bool(get_openai_key())
    return False

def ai_setup_message() -> str:
    return (
        "AI is not configured. Set GEMINI_API_KEY in credentials/.env or .streamlit/secrets.toml. "
        "Optional: set AI_PROVIDER=gemini and GEMINI_MODEL=gemini-2.5-flash."
    )

def trim_for_ai(text: str, limit: int = 12000) -> str:
    text = str(text or "")
    if len(text) <= limit:
        return text
    return text[:limit] + "\n\n[Context trimmed]"

def df_for_ai(df: pd.DataFrame, columns: list[str] | None = None, limit: int = 40) -> str:
    if df is None or df.empty:
        return "No rows."
    view = df.copy()
    if columns:
        view = view[[col for col in columns if col in view.columns]].copy()
    for col in view.columns:
        if pd.api.types.is_datetime64_any_dtype(view[col]):
            view[col] = view[col].dt.strftime("%Y-%m-%d")
    return view.head(limit).to_csv(index=False)

def ask_llm(task: str, context: str, temperature: float = 0.2) -> str:
    prompt = f"""
You are an assistant for {BRAND_NAME}, a boutique sales and passbook tracking app.
Use only the provided app context. If the context is insufficient, say what is missing.
Keep the answer practical, concise, and in rupees where amounts are present.

Task:
{task}

Context:
{trim_for_ai(context)}
""".strip()

    provider = get_ai_provider()
    if provider in {"gemini", "google"}:
        api_key = get_gemini_key()
        if not api_key:
            raise RuntimeError("GEMINI_API_KEY is not configured.")
        try:
            from google import genai
            from google.genai import types
        except ImportError as exc:
            raise RuntimeError("Install the google-genai package to use Gemini AI features.") from exc

        client = genai.Client(api_key=api_key)
        response = client.models.generate_content(
            model=get_gemini_model(),
            contents=prompt,
            config=types.GenerateContentConfig(temperature=temperature),
        )
        return getattr(response, "text", "") or str(response)

    if provider == "openai":
        api_key = get_openai_key()
        if not api_key:
            raise RuntimeError("OPENAI_API_KEY is not configured.")
        try:
            from openai import OpenAI
        except ImportError as exc:
            raise RuntimeError("Install the openai package to use OpenAI features.") from exc

        client = OpenAI(api_key=api_key)
        response = client.responses.create(
            model=get_openai_model(),
            input=prompt,
            temperature=temperature,
        )
        return getattr(response, "output_text", "") or str(response)

    raise RuntimeError(f"Unsupported AI_PROVIDER: {provider}. Use gemini or openai.")

def render_ai_panel(title: str, context: str, key: str, default_task: str, expanded: bool = False):
    with st.expander(title, expanded=expanded):
        if not llm_is_configured():
            st.info(ai_setup_message())
            return
        task = st.text_area("Ask AI", value=default_task, height=90, key=f"{key}_task")
        if st.button("Run AI", key=f"{key}_run", width="stretch"):
            try:
                with st.spinner("Thinking..."):
                    st.session_state[f"{key}_answer"] = ask_llm(task, context)
            except Exception as exc:
                st.error(str(exc))
        if st.session_state.get(f"{key}_answer"):
            st.markdown(st.session_state[f"{key}_answer"])

def render_ai_action_panel(title: str, context: str, key: str, actions: dict[str, str], expanded: bool = False):
    with st.expander(title, expanded=expanded):
        if not llm_is_configured():
            st.info(ai_setup_message())
            return
        action_names = list(actions.keys()) + ["Custom"]
        action_name = st.selectbox("AI Action", action_names, key=f"{key}_action")
        default_task = "" if action_name == "Custom" else actions.get(action_name, "")
        action_key = re.sub(r"[^0-9A-Za-z]+", "_", action_name).strip("_").lower() or "custom"
        task = st.text_area("Instruction", value=default_task, height=100, key=f"{key}_task_{action_key}")
        if st.button("Run AI", key=f"{key}_run", width="stretch"):
            if not task.strip():
                st.warning("Enter what you want AI to do.")
                return
            try:
                with st.spinner("Thinking..."):
                    st.session_state[f"{key}_answer"] = ask_llm(task, context)
            except Exception as exc:
                st.error(str(exc))
        if st.session_state.get(f"{key}_answer"):
            st.markdown(st.session_state[f"{key}_answer"])

def parse_json_from_ai(text: str) -> dict:
    raw = str(text or "").strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.IGNORECASE)
        raw = re.sub(r"\s*```$", "", raw)
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        start = raw.find("{")
        end = raw.rfind("}")
        if start == -1 or end == -1 or end <= start:
            raise ValueError("AI did not return valid JSON.")
        parsed = json.loads(raw[start:end + 1])
    if not isinstance(parsed, dict):
        raise ValueError("AI response must be a JSON object.")
    return parsed

def match_option(value, options: list[str]) -> str | None:
    text = str(value or "").strip()
    for option in options:
        if text.casefold() == option.casefold():
            return option
    compact = re.sub(r"[^a-z0-9]+", "", text.casefold())
    for option in options:
        if compact and compact == re.sub(r"[^a-z0-9]+", "", option.casefold()):
            return option
    return None

def bool_from_ai(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value or "").strip().casefold() in {"1", "true", "yes", "y", "paid", "done", "complete"}

def coerce_ai_sale_updates(row: pd.Series, updates: dict) -> tuple[dict, list[str]]:
    allowed = {
        "customer_name", "customer_phone", "sale_date", "product_category", "vendor",
        "product_description", "quantity", "buying_price", "selling_price", "amount_paid",
        "payment_method", "delay_status", "notes", "payment_received", "last_payment_date",
        "last_payment_method", "last_payment_received_by",
    }
    set_fields = {}
    warnings = []
    for key, value in (updates or {}).items():
        if value is None or str(value).strip() == "":
            continue
        if key not in allowed:
            warnings.append(f"Ignored unsupported field: {key}")
            continue
        if key == "customer_name":
            set_fields[key] = str(value).strip()[:120]
        elif key == "customer_phone":
            set_fields[key] = normalize_phone(str(value))[:20]
        elif key == "sale_date":
            parsed_date = pd.to_datetime(value, errors="coerce")
            if pd.isna(parsed_date):
                warnings.append("Ignored invalid sale_date.")
            else:
                set_fields[key] = str(parsed_date.date())
        elif key == "product_category":
            matched = match_option(value, CATEGORIES)
            if matched:
                set_fields[key] = matched
            else:
                warnings.append(f"Ignored invalid category: {value}")
        elif key == "vendor":
            set_fields[key] = str(value).strip()[:100]
        elif key == "product_description":
            set_fields[key] = str(value).strip()[:500]
        elif key == "quantity":
            try:
                qty = int(float(value))
                if qty < 1:
                    raise ValueError
                set_fields[key] = qty
            except Exception:
                warnings.append("Ignored invalid quantity.")
        elif key in {"buying_price", "selling_price", "amount_paid"}:
            amount = money_value(value, default=-1)
            if amount < 0:
                warnings.append(f"Ignored invalid {key}.")
            else:
                set_fields[key] = round(float(amount), 2)
        elif key == "payment_method":
            matched = match_option(value, PAYMENT_METHODS)
            if matched:
                set_fields[key] = matched
            else:
                warnings.append(f"Ignored invalid payment_method: {value}")
        elif key == "delay_status":
            set_fields[key] = int(bool_from_ai(value))
        elif key == "notes":
            set_fields[key] = str(value).strip()[:500]
        elif key == "payment_received":
            set_fields[key] = int(bool_from_ai(value))
        elif key == "last_payment_date":
            parsed_date = pd.to_datetime(value, errors="coerce")
            if pd.isna(parsed_date):
                warnings.append("Ignored invalid last_payment_date.")
            else:
                set_fields[key] = str(parsed_date.date())
        elif key == "last_payment_method":
            matched = match_option(value, PAYMENT_COLLECTION_METHODS)
            if matched:
                set_fields[key] = matched
            else:
                warnings.append(f"Ignored invalid last_payment_method: {value}")
        elif key == "last_payment_received_by":
            set_fields[key] = str(value).strip()[:80]

    current_buy = money_value(row.get("buying_price"))
    current_sell = money_value(row.get("selling_price"))
    current_paid = money_value(row.get("amount_paid"))
    buy = money_value(set_fields.get("buying_price", current_buy))
    sell = money_value(set_fields.get("selling_price", current_sell))
    paid = money_value(set_fields.get("amount_paid", current_paid))

    if set_fields.get("payment_received") == 1 and "amount_paid" not in set_fields:
        paid = sell
        set_fields["amount_paid"] = round(paid, 2)
    if paid > sell:
        warnings.append("Amount paid cannot exceed selling price; ignored AI update.")
        return {}, warnings
    if sell < buy:
        warnings.append("AI update creates a loss because selling price is below buying price.")

    if {"buying_price", "selling_price", "amount_paid", "payment_received"} & set(set_fields):
        pending = max(round(sell - paid, 2), 0.0)
        set_fields["pending_amount"] = pending
        set_fields["payment_received"] = 1 if pending == 0 else 0

    if set_fields:
        set_fields["updated_at"] = str(datetime.now())
    return set_fields, warnings

def coerce_ai_new_sale_draft(draft: dict) -> tuple[dict, list[str]]:
    fields = {}
    warnings = []
    existing_customers = get_existing_customers_with_phone()
    customer_by_key = {
        re.sub(r"[^a-z0-9]+", "", str(customer.get("_id", "")).casefold()): customer
        for customer in existing_customers
    }

    name = str(draft.get("customer_name", "") or "").strip()
    name_key = re.sub(r"[^a-z0-9]+", "", name.casefold())
    matched_customer = customer_by_key.get(name_key) if name_key else None
    if matched_customer:
        name = str(matched_customer.get("_id", "") or name).strip()
    fields["customer_name"] = name[:120]

    phone = str(draft.get("customer_phone", "") or "").strip()
    if not phone and matched_customer:
        phone = str(matched_customer.get("phone", "") or "")
    fields["customer_phone"] = normalize_phone(phone)[:20]

    parsed_date = pd.to_datetime(draft.get("sale_date") or date.today(), errors="coerce")
    fields["sale_date"] = str(parsed_date.date() if pd.notna(parsed_date) else date.today())

    category = match_option(draft.get("product_category"), CATEGORIES)
    if not category:
        category = "Other"
        if draft.get("product_category"):
            warnings.append(f"Category was not recognized, using Other instead of {draft.get('product_category')}.")
    fields["product_category"] = category

    fields["vendor"] = str(draft.get("vendor", "") or "").strip()[:100]
    fields["product_description"] = str(draft.get("product_description", "") or "").strip()[:500]

    try:
        qty = int(float(draft.get("quantity", 1) or 1))
        fields["quantity"] = max(qty, 1)
    except Exception:
        fields["quantity"] = 1
        warnings.append("Quantity was invalid, using 1.")

    fields["buying_price"] = round(money_value(draft.get("buying_price"), default=0), 2)
    fields["selling_price"] = round(money_value(draft.get("selling_price"), default=0), 2)
    fields["amount_paid"] = round(money_value(draft.get("amount_paid"), default=0), 2)

    if bool_from_ai(draft.get("payment_received")) and fields["amount_paid"] <= 0 and fields["selling_price"] > 0:
        fields["amount_paid"] = fields["selling_price"]

    if fields["amount_paid"] > fields["selling_price"] and fields["selling_price"] > 0:
        fields["amount_paid"] = fields["selling_price"]
        warnings.append("Amount paid was above selling price, capped to selling price.")

    payment_method = match_option(draft.get("payment_method"), PAYMENT_METHODS) or "UPI"
    fields["payment_method"] = payment_method
    fields["notes"] = str(draft.get("notes", "") or "").strip()[:500]

    fields["pending_amount"] = max(round(fields["selling_price"] - fields["amount_paid"], 2), 0.0)
    fields["payment_received"] = 1 if fields["pending_amount"] == 0 and fields["selling_price"] > 0 else 0
    fields["delay_status"] = int(bool_from_ai(draft.get("delay_status")))

    if not fields["customer_name"]:
        warnings.append("Customer name is missing.")
    if fields["buying_price"] <= 0:
        warnings.append("Buying price is missing or zero.")
    if fields["selling_price"] <= 0:
        warnings.append("Selling price is missing or zero.")
    if fields["selling_price"] and fields["buying_price"] and fields["selling_price"] < fields["buying_price"]:
        warnings.append("Selling price is below buying price.")

    return fields, warnings

def save_sale_record(fields: dict, source: str = "manual") -> int:
    sale_id = get_next_id()
    get_col().insert_one({
        "id":                  sale_id,
        "customer_name":       str(fields.get("customer_name", "")).strip()[:120],
        "customer_phone":      normalize_phone(fields.get("customer_phone", "")),
        "sale_date":           str(fields.get("sale_date") or date.today()),
        "vendor":              str(fields.get("vendor", "")).strip()[:100],
        "product_category":    fields.get("product_category") or "Other",
        "product_description": str(fields.get("product_description", "")).strip()[:500],
        "quantity":            int(fields.get("quantity") or 1),
        "buying_price":        round(money_value(fields.get("buying_price")), 2),
        "selling_price":       round(money_value(fields.get("selling_price")), 2),
        "amount_paid":         round(money_value(fields.get("amount_paid")), 2),
        "pending_amount":      round(money_value(fields.get("pending_amount")), 2),
        "payment_received":    int(fields.get("payment_received") or 0),
        "delay_status":        int(fields.get("delay_status") or 0),
        "payment_method":      fields.get("payment_method") if fields.get("payment_method") in PAYMENT_METHODS else "UPI",
        "notes":               str(fields.get("notes", "")).strip()[:500],
        "created_via":         source,
        "created_at":          str(datetime.now()),
    })
    invalidate_cache()
    return sale_id

def render_ai_sale_entry_assistant():
    if not llm_is_configured():
        with st.expander("AI Add Sale", expanded=False):
            st.info(ai_setup_message())
        return

    sec("AI Add Sale")
    st.caption("Type the sale in one line. AI will extract a draft, then you can check and save it.")
    brief = st.text_area(
        "Sale brief",
        placeholder="Example: Ramya bought saree from RUPALI. Buying 1495, selling 1895, paid 500 by UPI, balance pending.",
        height=95,
        key="ai_sale_brief",
    )
    b1, b2 = st.columns([1, 1])
    with b1:
        if st.button("Create Sale Draft", key="ai_sale_create_draft", type="primary", width="stretch"):
            if not brief.strip():
                st.warning("Type the sale details first.")
            else:
                existing_customers = get_existing_customers_with_phone()
                existing_context = pd.DataFrame([
                    {"customer_name": c.get("_id", ""), "phone": c.get("phone", ""), "visits": c.get("visits", 0)}
                    for c in existing_customers[:80]
                ]).to_csv(index=False)
                recent_sales = fetch_all()
                context = "\n".join([
                    f"Today is {date.today()}.",
                    "The user is describing one boutique sale. Extract fields from the text. Do not invent unknown prices.",
                    "If payment method is not mentioned, use UPI. If paid/full paid, amount_paid should equal selling_price.",
                    "If only pending/balance is mentioned, calculate amount_paid as selling_price minus pending amount.",
                    "Use an existing customer name exactly if the brief matches one.",
                    f"Allowed categories: {', '.join(CATEGORIES)}",
                    f"Allowed payment methods: {', '.join(PAYMENT_METHODS)}",
                    "Existing customers:\n" + existing_context,
                    "Recent sales:\n" + df_for_ai(
                        recent_sales.sort_values("sale_date", ascending=False) if not recent_sales.empty else recent_sales,
                        ["sale_date","customer_name","vendor","product_category","product_description","buying_price","selling_price","payment_method"],
                        30,
                    ),
                    "Return only JSON in this shape:",
                    '{"sale": {"customer_name": "", "customer_phone": "", "sale_date": "", "product_category": "", "vendor": "", "product_description": "", "quantity": 1, "buying_price": 0, "selling_price": 0, "amount_paid": 0, "payment_method": "UPI", "notes": ""}, "reason": "short explanation"}',
                ])
                try:
                    with st.spinner("Creating sale draft..."):
                        raw = ask_llm(brief, context, temperature=0.0)
                    parsed = parse_json_from_ai(raw)
                    fields, warnings = coerce_ai_new_sale_draft(parsed.get("sale", {}))
                    st.session_state.ai_sale_draft = {
                        "fields": fields,
                        "warnings": warnings,
                        "reason": str(parsed.get("reason", "") or ""),
                    }
                except Exception as exc:
                    st.error(str(exc))
    with b2:
        if st.button("Clear AI Draft", key="ai_sale_clear_draft", width="stretch"):
            st.session_state.pop("ai_sale_draft", None)
            st.rerun()

    draft = st.session_state.get("ai_sale_draft")
    if not draft:
        return

    fields = draft.get("fields", {})
    if draft.get("reason"):
        st.info(draft["reason"])
    for warning in draft.get("warnings", []):
        st.warning(warning)

    with st.form("ai_sale_save_form"):
        c1, c2, c3 = st.columns(3)
        with c1:
            customer_name = st.text_input("Customer Name *", value=str(fields.get("customer_name", "")), key="ai_sale_customer_name")
        with c2:
            customer_phone = st.text_input("Phone", value=str(fields.get("customer_phone", "")), key="ai_sale_customer_phone")
        with c3:
            sale_dt = pd.to_datetime(fields.get("sale_date"), errors="coerce")
            sale_date_value = sale_dt.date() if pd.notna(sale_dt) else date.today()
            sale_date = st.date_input("Sale Date", value=sale_date_value, key="ai_sale_date")

        p1, p2, p3 = st.columns(3)
        with p1:
            category_index = CATEGORIES.index(fields.get("product_category")) if fields.get("product_category") in CATEGORIES else CATEGORIES.index("Other")
            product_category = st.selectbox("Category *", CATEGORIES, index=category_index, key="ai_sale_category")
        with p2:
            vendor = st.text_input("Vendor / Supplier", value=str(fields.get("vendor", "")), key="ai_sale_vendor")
        with p3:
            quantity = st.number_input("Quantity", min_value=1, step=1, value=int(fields.get("quantity") or 1), key="ai_sale_quantity")

        product_description = st.text_area("Description", value=str(fields.get("product_description", "")), height=65, key="ai_sale_description")

        pr1, pr2, pr3, pr4 = st.columns(4)
        with pr1:
            buying_price = st.number_input("Buying Price (₹) *", min_value=0.0, step=1.0, value=float(fields.get("buying_price") or 0), key="ai_sale_buying_price")
        with pr2:
            selling_price = st.number_input("Selling Price (₹) *", min_value=0.0, step=1.0, value=float(fields.get("selling_price") or 0), key="ai_sale_selling_price")
        with pr3:
            amount_paid = st.number_input("Amount Paid (₹)", min_value=0.0, step=1.0, value=float(fields.get("amount_paid") or 0), key="ai_sale_amount_paid")
        with pr4:
            pm_index = PAYMENT_METHODS.index(fields.get("payment_method")) if fields.get("payment_method") in PAYMENT_METHODS else PAYMENT_METHODS.index("UPI")
            payment_method = st.selectbox("Payment Method", PAYMENT_METHODS, index=pm_index, key="ai_sale_payment_method")

        pending_amount = max(round(float(selling_price) - float(amount_paid), 2), 0.0)
        profit_amount = round((float(selling_price) - float(buying_price)) * int(quantity), 2)
        am1, am2, am3 = st.columns(3)
        am1.metric("Pending", f"₹{pending_amount:,.2f}")
        am2.metric("Profit", f"₹{profit_amount:,.2f}")
        am3.metric("Total Value", f"₹{float(selling_price) * int(quantity):,.2f}")

        notes = st.text_area("Notes", value=str(fields.get("notes", "")), height=60, key="ai_sale_notes")
        save_ai_sale = st.form_submit_button("Save AI Sale", type="primary", width="stretch")

        if save_ai_sale:
            errs = []
            if not customer_name.strip():
                errs.append("Customer name is required.")
            if buying_price <= 0:
                errs.append("Buying price must be > 0.")
            if selling_price <= 0:
                errs.append("Selling price must be > 0.")
            if amount_paid > selling_price:
                errs.append("Amount paid cannot exceed selling price.")
            if selling_price and buying_price and selling_price < buying_price:
                st.warning("Selling price is below buying price.")
            if errs:
                for err in errs:
                    st.error(err)
            else:
                sale_fields = {
                    "customer_name": customer_name,
                    "customer_phone": customer_phone,
                    "sale_date": str(sale_date),
                    "vendor": vendor,
                    "product_category": product_category,
                    "product_description": product_description,
                    "quantity": int(quantity),
                    "buying_price": float(buying_price),
                    "selling_price": float(selling_price),
                    "amount_paid": float(amount_paid),
                    "pending_amount": pending_amount,
                    "payment_received": 1 if pending_amount == 0 else 0,
                    "delay_status": 0,
                    "payment_method": payment_method,
                    "notes": notes,
                }
                sale_id = save_sale_record(sale_fields, source="ai_add_sale")
                st.session_state.pop("ai_sale_draft", None)
                st.success(f"AI sale saved as Sale #{sale_id}.")
                st.rerun()

def render_ai_update_assistant(row: pd.Series, sale_id: int):
    with st.expander("AI Update Assistant", expanded=False):
        if not llm_is_configured():
            st.info(ai_setup_message())
            return
        st.caption("Type updates in plain English. You will see a preview before anything is saved.")
        instruction = st.text_area(
            "What should I update?",
            placeholder="Example: change selling price to 4200, amount paid to 2000, payment method UPI, and notes customer will pay balance next week",
            height=100,
            key=f"ai_update_instruction_{sale_id}",
        )
        preview_key = f"ai_update_preview_{sale_id}"
        if st.button("Preview AI Update", key=f"ai_update_preview_btn_{sale_id}", width="stretch"):
            if not instruction.strip():
                st.warning("Type what needs to be updated.")
            else:
                current_record = row.to_dict()
                context = "\n".join([
                    f"Sale ID: {sale_id}",
                    "Current sale record JSON:",
                    json.dumps(current_record, default=str, ensure_ascii=False),
                    "Allowed fields:",
                    ", ".join([
                        "customer_name", "customer_phone", "sale_date", "product_category", "vendor",
                        "product_description", "quantity", "buying_price", "selling_price",
                        "amount_paid", "payment_method", "delay_status", "notes", "payment_received",
                        "last_payment_date", "last_payment_method", "last_payment_received_by",
                    ]),
                    f"Allowed categories: {', '.join(CATEGORIES)}",
                    f"Allowed sale payment methods: {', '.join(PAYMENT_METHODS)}",
                    f"Allowed received payment methods: {', '.join(PAYMENT_COLLECTION_METHODS)}",
                    "If the user says mark paid, set payment_received true and amount_paid equal to selling_price.",
                    "If the user gives a partial payment amount, set amount_paid to that amount.",
                    "Return only JSON in this shape: {\"updates\": {\"field\": \"value\"}, \"reason\": \"short explanation\"}.",
                ])
                try:
                    with st.spinner("Reading your update..."):
                        raw = ask_llm(instruction, context, temperature=0.0)
                    parsed = parse_json_from_ai(raw)
                    set_fields, warnings = coerce_ai_sale_updates(row, parsed.get("updates", {}))
                    st.session_state[preview_key] = {
                        "updates": set_fields,
                        "warnings": warnings,
                        "reason": str(parsed.get("reason", "") or ""),
                        "raw": raw,
                    }
                except Exception as exc:
                    st.error(str(exc))

        preview = st.session_state.get(preview_key)
        if preview:
            updates = preview.get("updates", {})
            warnings = preview.get("warnings", [])
            if preview.get("reason"):
                st.info(preview["reason"])
            for warning in warnings:
                st.warning(warning)
            if not updates:
                st.info("No safe update fields were found.")
                return
            change_rows = []
            for field, new_value in updates.items():
                if field == "updated_at":
                    continue
                change_rows.append({
                    "Field": field,
                    "Current": row.get(field, ""),
                    "New": new_value,
                })
            st.dataframe(pd.DataFrame(change_rows), hide_index=True, width="stretch")
            a1, a2 = st.columns(2)
            with a1:
                if st.button("Apply AI Update", key=f"ai_update_apply_{sale_id}", type="primary", width="stretch"):
                    get_col().update_one({"id": int(sale_id)}, {"$set": updates})
                    invalidate_cache()
                    st.session_state.pop(preview_key, None)
                    st.success("AI update applied.")
                    st.rerun()
            with a2:
                if st.button("Clear AI Preview", key=f"ai_update_clear_{sale_id}", width="stretch"):
                    st.session_state.pop(preview_key, None)
                    st.rerun()

@st.cache_resource
def get_mongo_client():
    try:
        uri = safe_secret("MONGO_URI", os.getenv("MONGO_URI", ""))
        if not uri:
            st.error("⚠️ MONGO_URI not configured.")
            st.stop()
        client = MongoClient(uri, serverSelectionTimeoutMS=5000)
        client.admin.command("ping")
        return client
    except Exception as e:
        st.error(f"MongoDB connection failed: {e}")
        st.stop()

def get_db():
    return get_mongo_client()["boutique_db"]

def get_col():
    return get_db()["sales"]

def managed_secrets_collection():
    return get_db()["managed_secrets"]

def _secret_encryption_material() -> str:
    return (
        safe_secret("SECRET_ENCRYPTION_KEY", "")
        or safe_secret("TECHNICAL_SECRET_KEY", "")
        or os.getenv("SECRET_ENCRYPTION_KEY", "")
        or os.getenv("TECHNICAL_SECRET_KEY", "")
        or safe_secret("PASSWORD_HASH", "")
        or os.getenv("PASSWORD_HASH", "")
        or safe_secret("PASSWORD", "")
        or os.getenv("PASSWORD", "")
    )

def _secret_cipher():
    material = _secret_encryption_material()
    if not material:
        raise RuntimeError("Set PASSWORD, PASSWORD_HASH, or SECRET_ENCRYPTION_KEY before saving managed secrets.")
    try:
        from cryptography.fernet import Fernet
    except ImportError as exc:
        raise RuntimeError("Install cryptography to save encrypted managed secrets.") from exc
    key = base64.urlsafe_b64encode(hashlib.sha256(str(material).encode("utf-8")).digest())
    return Fernet(key)

def encrypt_managed_secret(value: str) -> str:
    return _secret_cipher().encrypt(str(value or "").encode("utf-8")).decode("utf-8")

def decrypt_managed_secret(token: str) -> str:
    if not token:
        return ""
    return _secret_cipher().decrypt(str(token).encode("utf-8")).decode("utf-8")

def mask_secret_value(value: str, visible: int = 4) -> str:
    value = str(value or "")
    if not value:
        return "Not set"
    if len(value) <= visible * 2:
        return "•" * len(value)
    return f"{value[:visible]}{'•' * 8}{value[-visible:]}"

def mask_mongo_uri(uri: str) -> str:
    uri = str(uri or "")
    if not uri:
        return "Not set"
    return re.sub(r"(mongodb(?:\+srv)?://)([^:@/]+):([^@/]+)@", r"\1\2:••••••••@", uri)

def get_managed_secret_doc(key: str) -> dict | None:
    key = str(key or "").strip().upper()
    if not key:
        return None
    try:
        return managed_secrets_collection().find_one({"_id": key})
    except Exception:
        return None

def get_managed_secret(key: str) -> str:
    doc = get_managed_secret_doc(key)
    if not doc:
        return ""
    try:
        return decrypt_managed_secret(doc.get("value_enc", ""))
    except Exception:
        return ""

def save_managed_secret(key: str, value: str):
    key = str(key or "").strip().upper()
    value = str(value or "")
    if key not in MANAGED_SECRET_KEYS:
        raise ValueError("Unsupported secret key.")
    if not value:
        raise ValueError("Secret value cannot be empty.")
    doc = {
        "_id": key,
        "value_enc": encrypt_managed_secret(value),
        "fingerprint": hashlib.sha256(value.encode("utf-8")).hexdigest()[:16],
        "masked": mask_mongo_uri(value) if key == "MONGO_URI" else mask_secret_value(value),
        "updated_at": datetime.now(),
        "updated_by": st.session_state.get("username", "Admin"),
        "active": True,
    }
    managed_secrets_collection().update_one({"_id": key}, {"$set": doc}, upsert=True)

def delete_managed_secret(key: str):
    key = str(key or "").strip().upper()
    if key:
        managed_secrets_collection().delete_one({"_id": key})

def secret_presence(key: str) -> dict:
    key = str(key or "").strip().upper()
    managed_doc = get_managed_secret_doc(key)
    env_value = os.getenv(key, "")
    streamlit_value = safe_secret(key, "")
    source = []
    if managed_doc:
        source.append("Managed")
    if streamlit_value:
        source.append("Streamlit")
    if env_value:
        source.append("Environment")
    raw_value = get_managed_secret(key) if managed_doc else (streamlit_value or env_value)
    return {
        "key": key,
        "source": " + ".join(source) if source else "Not set",
        "managed": bool(managed_doc),
        "masked": mask_mongo_uri(raw_value) if key == "MONGO_URI" else mask_secret_value(raw_value),
        "updated_at": managed_doc.get("updated_at") if managed_doc else "",
        "updated_by": managed_doc.get("updated_by", "") if managed_doc else "",
    }

def get_next_id():
    counter = get_db()["counters"].find_one_and_update(
        {"_id": "sales_id"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    return counter["seq"]

# =====================================================
# DATA HELPERS
# =====================================================

@st.cache_data(ttl=30)
def fetch_all() -> pd.DataFrame:
    docs = list(get_col().find({}, {"_id": 0}))
    if not docs:
        return pd.DataFrame()
    df = pd.DataFrame(docs)
    for c in ["buying_price", "selling_price", "amount_paid", "pending_amount"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    for c in ["payment_received", "delay_status"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype(int)
    if "sale_date" in df.columns:
        df["sale_date"] = pd.to_datetime(df["sale_date"], errors="coerce")
    df["profit"] = df["selling_price"] - df["buying_price"]
    df["margin"] = (df["profit"] / df["selling_price"].replace(0, 1) * 100).round(2)
    for col in ["vendor", "product_description", "notes", "customer_phone", "last_payment_date", "payment_date", "last_payment_method", "last_payment_received_by"]:
        if col not in df.columns:
            df[col] = ""
    return df

def invalidate_cache():
    fetch_all.clear()
    try:
        get_existing_vendors.clear()
    except NameError:
        pass

def metrics(df: pd.DataFrame) -> dict:
    if df.empty:
        return dict(sales=0, revenue=0, profit=0, pending=0, delayed=0, margin=0, customers=0)
    return dict(
        sales     = len(df),
        revenue   = df["selling_price"].sum(),
        profit    = df["profit"].sum(),
        pending   = df["pending_amount"].sum(),
        delayed   = int((df["delay_status"] == 1).sum()),
        margin    = df["margin"].mean(),
        customers = df["customer_name"].nunique(),
    )

def to_excel(df: pd.DataFrame) -> BytesIO:
    out = BytesIO()
    ex = df.copy()
    if "sale_date" in ex.columns:
        ex["sale_date"] = ex["sale_date"].astype(str)
    ex["profit"]        = (ex["selling_price"] - ex["buying_price"]).round(2)
    ex["profit_margin"] = (ex["profit"] / ex["selling_price"].replace(0, 1) * 100).round(2)
    ex["status"]  = ex["payment_received"].map({0: "Pending", 1: "Received"})
    ex["delayed"] = ex["delay_status"].map({0: "No", 1: "Yes"})
    ordered = ["id","customer_name","customer_phone","sale_date","vendor","product_category","product_description","buying_price","selling_price","profit","profit_margin","amount_paid","pending_amount","status","delayed","payment_method","last_payment_method","last_payment_date","last_payment_received_by","payment_date","notes","created_at"]
    cols = [c for c in ordered if c in ex.columns]
    ex = ex[cols]
    ex.columns = [c.replace("_", " ").title() for c in ex.columns]
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        ex.to_excel(w, index=False)
        ws = w.sheets["Sheet1"]
        for i, col in enumerate(ex.columns, 1):
            ml = max(ex.iloc[:, i-1].astype(str).str.len().max(), len(col)) + 4
            ws.column_dimensions[ws.cell(1, i).column_letter].width = min(ml, 45)
        from openpyxl.styles import Font, PatternFill, Alignment
        blue_fill = PatternFill("solid", fgColor="2E6FD8")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="E8EEF8")
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal="center")
    out.seek(0)
    return out

def get_existing_customers():
    pipeline = [
        {"$match": {"customer_name": {"$ne": None, "$ne": ""}}},
        {"$group": {
            "_id": "$customer_name",
            "visits": {"$sum": 1},
            "last_sale": {"$max": "$sale_date"},
        }},
        {"$sort": {"_id": 1}},
    ]
    return list(get_col().aggregate(pipeline))

def get_existing_customers_with_phone():
    pipeline = [
        {"$match": {"customer_name": {"$ne": None, "$ne": ""}}},
        {"$sort": {"sale_date": -1, "created_at": -1}},
        {"$group": {
            "_id": "$customer_name",
            "phones": {"$push": "$customer_phone"},
            "visits": {"$sum": 1},
            "last_sale": {"$max": "$sale_date"},
        }},
        {"$sort": {"_id": 1}},
    ]
    customers = list(get_col().aggregate(pipeline))
    for customer in customers:
        customer["phone"] = next((str(p).strip() for p in customer.get("phones", []) if str(p or "").strip()), "")
        customer.pop("phones", None)
    return customers

@st.cache_data(ttl=60)
def get_existing_vendors():
    vendors = set()
    for collection_name in ("sales", "inventory"):
        try:
            for vendor in get_db()[collection_name].distinct("vendor"):
                vendor = str(vendor or "").strip()
                if vendor:
                    vendors.add(vendor)
        except Exception:
                pass
    return sorted(vendors, key=str.casefold)

def app_settings_collection():
    return get_db()["app_settings"]

def normalize_app_settings(raw: dict | None = None) -> dict:
    settings = APP_SETTING_DEFAULTS.copy()
    settings.update(raw or {})

    settings["theme"] = str(settings.get("theme", "light")).strip().lower()
    if settings["theme"] not in {"light", "dark"}:
        settings["theme"] = APP_SETTING_DEFAULTS["theme"]

    allowed_pages = [p for p in dict.fromkeys(ADMIN_NAV_OPTIONS + MEMBER_NAV_OPTIONS) if p != "Logout"]
    if settings.get("default_page") not in allowed_pages:
        settings["default_page"] = APP_SETTING_DEFAULTS["default_page"]

    if settings.get("default_payment_method") not in PAYMENT_METHODS:
        settings["default_payment_method"] = APP_SETTING_DEFAULTS["default_payment_method"]

    if settings.get("default_bill_scope") not in BILL_SCOPE_OPTIONS:
        settings["default_bill_scope"] = APP_SETTING_DEFAULTS["default_bill_scope"]

    try:
        settings["default_review_days"] = max(7, min(int(settings.get("default_review_days", 90)), 365))
    except Exception:
        settings["default_review_days"] = APP_SETTING_DEFAULTS["default_review_days"]

    inventory_options = ["All"] + CATEGORIES
    if settings.get("default_inventory_category") not in inventory_options:
        settings["default_inventory_category"] = APP_SETTING_DEFAULTS["default_inventory_category"]

    settings["ai_provider"] = str(settings.get("ai_provider", "auto")).strip().lower()
    if settings["ai_provider"] not in {"auto", "gemini", "google", "openai"}:
        settings["ai_provider"] = APP_SETTING_DEFAULTS["ai_provider"]
    if settings["ai_provider"] == "google":
        settings["ai_provider"] = "gemini"

    settings["gemini_model"] = str(settings.get("gemini_model", "") or "").strip()[:120]
    settings["openai_model"] = str(settings.get("openai_model", "") or "").strip()[:120]
    return settings

@st.cache_data(ttl=60)
def load_app_settings() -> dict:
    try:
        doc = app_settings_collection().find_one({"_id": "global"}, {"_id": 0}) or {}
    except Exception:
        doc = {}
    return normalize_app_settings(doc)

def save_app_settings(updates: dict):
    current = load_app_settings()
    current.update(updates or {})
    clean = normalize_app_settings(current)
    clean.update({
        "updated_at": str(datetime.now()),
        "updated_by": st.session_state.get("username", "Admin"),
    })
    app_settings_collection().update_one({"_id": "global"}, {"$set": clean}, upsert=True)
    load_app_settings.clear()
    for key in APP_SETTING_DEFAULTS:
        st.session_state[key] = clean[key]

def apply_persistent_app_settings():
    if st.session_state.get("app_settings_loaded"):
        return
    settings = load_app_settings()
    for key, value in settings.items():
        st.session_state[key] = value
    st.session_state.app_settings_loaded = True

def app_pref(key: str, default=None):
    if key in st.session_state:
        return st.session_state[key]
    return load_app_settings().get(key, APP_SETTING_DEFAULTS.get(key, default))

def nav_options_for_current_user(include_logout: bool = True) -> list[str]:
    options = ADMIN_NAV_OPTIONS if _is_admin() else MEMBER_NAV_OPTIONS
    if include_logout:
        return list(options)
    return [page for page in options if page != "Logout"]

def render_top_settings():
    notice = st.session_state.pop("settings_notice", None)
    if notice:
        st.toast(notice)

    nav_choices = nav_options_for_current_user(include_logout=False)
    default_page = app_pref("default_page", nav_choices[0])
    if default_page not in nav_choices:
        default_page = nav_choices[0]

    payment_default = app_pref("default_payment_method", "UPI")
    if payment_default not in PAYMENT_METHODS:
        payment_default = "UPI" if "UPI" in PAYMENT_METHODS else PAYMENT_METHODS[0]

    bill_default = app_pref("default_bill_scope", BILL_SCOPE_ALL)
    if bill_default not in BILL_SCOPE_OPTIONS:
        bill_default = BILL_SCOPE_ALL

    inventory_options = ["All"] + CATEGORIES
    inventory_default = app_pref("default_inventory_category", "All")
    if inventory_default not in inventory_options:
        inventory_default = "All"

    provider_options = ["auto", "gemini", "openai"]
    provider_labels = {
        "auto": "Auto",
        "gemini": "Gemini",
        "openai": "OpenAI",
    }
    provider_default = app_pref("ai_provider", "auto")
    if provider_default not in provider_options:
        provider_default = "auto"

    st.markdown("<div class='settings-top-pad'></div>", unsafe_allow_html=True)
    with st.popover("Settings"):
        app_tab, ai_tab, account_tab = st.tabs(["App", "AI", "Account"])

        with app_tab:
            theme_choice = st.radio(
                "Appearance",
                ["Light", "Dark"],
                horizontal=True,
                index=1 if app_pref("theme", "light") == "dark" else 0,
                key="settings_theme_choice",
            )
            default_page_choice = st.selectbox(
                "Default page after sign-in",
                nav_choices,
                index=nav_choices.index(default_page),
                key="settings_default_page_choice",
            )
            default_payment_choice = st.selectbox(
                "Default payment method",
                PAYMENT_METHODS,
                index=PAYMENT_METHODS.index(payment_default),
                key="settings_default_payment_choice",
            )
            default_bill_choice = st.selectbox(
                "Default bill type",
                BILL_SCOPE_OPTIONS,
                index=BILL_SCOPE_OPTIONS.index(bill_default),
                key="settings_default_bill_choice",
            )
            review_days_choice = st.number_input(
                "Review date range",
                min_value=7,
                max_value=365,
                value=int(app_pref("default_review_days", 90)),
                step=7,
                key="settings_review_days_choice",
            )
            inventory_category_choice = st.selectbox(
                "Default inventory category",
                inventory_options,
                index=inventory_options.index(inventory_default),
                key="settings_inventory_category_choice",
            )
            if st.button("Save App Settings", key="settings_save_app", width="stretch"):
                save_app_settings({
                    "theme": theme_choice.lower(),
                    "default_page": default_page_choice,
                    "default_payment_method": default_payment_choice,
                    "default_bill_scope": default_bill_choice,
                    "default_review_days": int(review_days_choice),
                    "default_inventory_category": inventory_category_choice,
                })
                st.session_state.settings_force_page = default_page_choice
                st.session_state.settings_notice = "Settings saved."
                st.rerun()

        with ai_tab:
            ai_status = "Configured" if llm_is_configured() else "Not configured"
            gemini_status = "Gemini key found" if get_gemini_key() else "Gemini key missing"
            st.markdown(
                f"<div class='settings-status'><b>AI status:</b> {html_escape(ai_status)}<br>{html_escape(gemini_status)}</div>",
                unsafe_allow_html=True,
            )
            provider_choice = st.selectbox(
                "AI provider",
                provider_options,
                format_func=lambda value: provider_labels.get(value, value.title()),
                index=provider_options.index(provider_default),
                key="settings_ai_provider_choice",
            )
            gemini_model_choice = st.text_input(
                "Gemini model override",
                value=str(app_pref("gemini_model", "") or ""),
                placeholder=safe_secret("GEMINI_MODEL", "") or os.getenv("GEMINI_MODEL", "gemini-2.5-flash"),
                key="settings_gemini_model_choice",
            )
            openai_model_choice = st.text_input(
                "OpenAI model override",
                value=str(app_pref("openai_model", "") or ""),
                placeholder=safe_secret("OPENAI_MODEL", "") or os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
                key="settings_openai_model_choice",
            )
            if st.button("Save AI Settings", key="settings_save_ai", width="stretch"):
                save_app_settings({
                    "ai_provider": provider_choice,
                    "gemini_model": gemini_model_choice,
                    "openai_model": openai_model_choice,
                })
                st.session_state.settings_notice = "AI settings saved."
                st.rerun()

        with account_tab:
            username = st.session_state.get("username", "Admin")
            role = st.session_state.get("user_role", "admin").title()
            method = st.session_state.get("auth_method", "password").replace("_", " ").title()
            st.markdown(
                f"<div class='settings-status'><b>{html_escape(username)}</b><br>Role: {html_escape(role)}<br>Login: {html_escape(method)}</div>",
                unsafe_allow_html=True,
            )
            if _is_admin():
                if st.button("Open Security & Devices", key="settings_open_security", width="stretch"):
                    st.session_state.settings_force_page = "Security & Devices"
                    st.rerun()
            if st.button("Refresh Data Cache", key="settings_refresh_cache", width="stretch"):
                invalidate_cache()
                load_app_settings.clear()
                st.session_state.settings_notice = "Data cache refreshed."
                st.rerun()
            if st.button("Sign Out", key="settings_sign_out", width="stretch"):
                logout_current_session()
                st.rerun()

# =====================================================
# HELPERS
# =====================================================

def page_header(title, sub):
    if st.session_state.get("logged_in", False):
        left, right = st.columns([0.78, 0.22])
        with left:
            st.markdown(f"<div class='page-title'>{html_escape(title)}</div>", unsafe_allow_html=True)
            st.markdown(f"<div class='page-sub'>{html_escape(sub)}</div>", unsafe_allow_html=True)
        with right:
            render_top_settings()
    else:
        st.markdown(f"<div class='page-title'>{html_escape(title)}</div>", unsafe_allow_html=True)
        st.markdown(f"<div class='page-sub'>{html_escape(sub)}</div>", unsafe_allow_html=True)

def sec(label):
    st.markdown(f"<div class='sec-head'>{html_escape(label)}</div>", unsafe_allow_html=True)

def rule():
    st.markdown("<hr class='rule'>", unsafe_allow_html=True)

def rule_sm():
    st.markdown("<hr class='rule-sm'>", unsafe_allow_html=True)

def is_admin():
    return st.session_state.get("logged_in", False)

def normalize_phone(phone: str) -> str:
    digits = re.sub(r"\D", "", str(phone or ""))
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    elif len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    return digits[:20]

def first_nonempty(values) -> str:
    for value in values:
        value = str(value or "").strip()
        if value:
            return value
    return ""

def parse_currency(raw: str) -> tuple[float, bool]:
    text = str(raw or "").strip()
    if not text:
        return 0.0, True
    cleaned = re.sub(r"[₹,\s]", "", text)
    if not cleaned:
        return 0.0, True
    try:
        value = float(cleaned)
    except ValueError:
        return 0.0, False
    return value, value >= 0

def money_value(value, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default

def currency_input(label: str, key: str, value: float | None = None) -> tuple[str, float, bool]:
    default = "" if value in (None, 0, 0.0) else f"{float(value):.2f}"
    raw = st.text_input(label, value=default, placeholder="0.00", key=key)
    parsed, valid = parse_currency(raw)
    return raw, parsed, valid

def clean_text_cell(value) -> str:
    text = str(value or "").strip()
    return "" if text in ("—", "None", "NaT", "nan") else text

def date_cell_to_iso(value, required: bool = False) -> tuple[str, bool]:
    if value is None or str(value) in ("", "NaT", "nan", "None"):
        return "", not required
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return "", False
    return str(parsed.date()), True

def default_receiver_name() -> str:
    return str(st.session_state.get("username") or "Admin").strip().title()

def record_payment(row: pd.Series, payment_amount: float, payment_date: date, payment_method: str, received_by: str) -> tuple[bool, str]:
    pending = round(max(money_value(row.get("pending_amount")), 0.0), 2)
    amount = round(payment_amount, 2)
    payment_method = str(payment_method or "").strip()
    received_by = str(received_by or "").strip()
    if amount <= 0:
        return False, "Payment amount must be greater than 0."
    if amount > pending:
        return False, "Payment amount cannot exceed the pending amount."
    if not payment_method:
        return False, "Select how the customer paid."
    if not received_by:
        return False, "Enter who received the payment."

    new_pending = round(max(pending - amount, 0.0), 2)
    new_paid = round(money_value(row.get("amount_paid")) + amount, 2)
    payment_entry = {
        "amount": amount,
        "date": str(payment_date),
        "method": payment_method,
        "received_by": received_by[:80],
        "recorded_at": str(datetime.now()),
        "recorded_by": st.session_state.get("username", "Admin"),
    }
    set_fields = {
        "amount_paid": new_paid,
        "pending_amount": new_pending,
        "payment_received": 1 if new_pending == 0 else 0,
        "last_payment_date": str(payment_date),
        "last_payment_method": payment_method,
        "last_payment_received_by": received_by[:80],
        "payment_method": payment_method,
        "updated_at": str(datetime.now()),
    }
    if new_pending == 0:
        set_fields["payment_date"] = str(payment_date)

    get_col().update_one(
        {"id": int(row["id"])},
        {"$set": set_fields, "$push": {"payment_history": payment_entry}},
    )
    invalidate_cache()
    status = "Payment completed." if new_pending == 0 else f"Partial payment saved. ₹{new_pending:,.2f} still pending."
    return True, status

def save_account_editor_changes(original_df: pd.DataFrame, edited_df: pd.DataFrame) -> tuple[int, list[str]]:
    originals = original_df.set_index("id", drop=False)
    changed = 0
    errors = []

    for _, row in edited_df.iterrows():
        try:
            row_id = int(row["ID"])
        except (TypeError, ValueError):
            errors.append("One edited row has an invalid ID.")
            continue
        if row_id not in originals.index:
            errors.append(f"Sale #{row_id} was not found.")
            continue

        sale_date, sale_date_ok = date_cell_to_iso(row.get("Date"), required=True)
        paid_date, paid_date_ok = date_cell_to_iso(row.get("Paid Date"), required=False)
        buy = round(money_value(row.get("Buy ₹")), 2)
        sell = round(money_value(row.get("Sell ₹")), 2)
        paid = round(money_value(row.get("Paid ₹")), 2)
        if not sale_date_ok:
            errors.append(f"Sale #{row_id}: enter a valid sale date.")
            continue
        if not paid_date_ok:
            errors.append(f"Sale #{row_id}: enter a valid paid date or leave it blank.")
            continue
        if buy < 0 or sell < 0 or paid < 0:
            errors.append(f"Sale #{row_id}: amounts cannot be negative.")
            continue
        if paid > sell:
            errors.append(f"Sale #{row_id}: paid amount cannot exceed selling price.")
            continue

        pending = round(max(sell - paid, 0.0), 2)
        set_fields = {
            "customer_name": clean_text_cell(row.get("Customer"))[:120],
            "customer_phone": normalize_phone(row.get("Phone")),
            "sale_date": sale_date,
            "vendor": clean_text_cell(row.get("Vendor"))[:100],
            "product_category": clean_text_cell(row.get("Category")) or CATEGORIES[0],
            "buying_price": buy,
            "selling_price": sell,
            "amount_paid": paid,
            "pending_amount": pending,
            "payment_received": 1 if pending == 0 else 0,
            "payment_method": clean_text_cell(row.get("Sale Method")) or PAYMENT_METHODS[0],
            "last_payment_method": clean_text_cell(row.get("Paid Method")),
            "last_payment_date": paid_date,
            "last_payment_received_by": clean_text_cell(row.get("Received By"))[:80],
            "payment_date": paid_date if pending == 0 else "",
            "updated_at": str(datetime.now()),
        }
        if set_fields["product_category"] not in CATEGORIES:
            errors.append(f"Sale #{row_id}: choose a valid category.")
            continue
        if set_fields["payment_method"] not in PAYMENT_METHODS:
            errors.append(f"Sale #{row_id}: choose a valid sale method.")
            continue
        if set_fields["last_payment_method"] and set_fields["last_payment_method"] not in PAYMENT_COLLECTION_METHODS:
            errors.append(f"Sale #{row_id}: choose a valid paid method.")
            continue

        current = originals.loc[row_id]
        comparable = {
            "customer_name": str(current.get("customer_name", "") or "").strip()[:120],
            "customer_phone": normalize_phone(current.get("customer_phone", "")),
            "sale_date": date_cell_to_iso(current.get("sale_date"), required=False)[0],
            "vendor": str(current.get("vendor", "") or "").strip()[:100],
            "product_category": str(current.get("product_category", "") or "").strip(),
            "buying_price": round(money_value(current.get("buying_price")), 2),
            "selling_price": round(money_value(current.get("selling_price")), 2),
            "amount_paid": round(money_value(current.get("amount_paid")), 2),
            "pending_amount": round(money_value(current.get("pending_amount")), 2),
            "payment_received": int(money_value(current.get("payment_received"))),
            "payment_method": str(current.get("payment_method", "") or "").strip(),
            "last_payment_method": str(current.get("last_payment_method", "") or "").strip(),
            "last_payment_date": date_cell_to_iso(current.get("last_payment_date"), required=False)[0],
            "last_payment_received_by": str(current.get("last_payment_received_by", "") or "").strip()[:80],
            "payment_date": date_cell_to_iso(current.get("payment_date"), required=False)[0] if pending == 0 else "",
        }
        if any(set_fields[field] != comparable.get(field) for field in comparable):
            get_col().update_one({"id": row_id}, {"$set": set_fields})
            changed += 1

    if changed:
        invalidate_cache()
    return changed, errors

def bill_file_name(customer_name: str, bill_id: str = "") -> str:
    safe_name = re.sub(r"[^0-9A-Za-z]+", "_", str(customer_name or "customer")).strip("_").lower()
    safe_bill_id = re.sub(r"[^0-9A-Za-z]+", "_", str(bill_id or "")).strip("_").lower()
    prefix = f"{safe_bill_id}_" if safe_bill_id else ""
    return f"{prefix}bill_{safe_name or 'customer'}_{date.today()}.pdf"

def normalize_bill_scope(bill_scope: str | None) -> str:
    return bill_scope if bill_scope in BILL_SCOPE_OPTIONS else BILL_SCOPE_OPTIONS[0]

def normalize_bill_limit(bill_limit: int | float | str | None) -> int:
    try:
        return max(int(bill_limit), 1)
    except (TypeError, ValueError):
        return 5

def bill_scope_label(bill_scope: str | None, bill_limit: int | float | str | None = None) -> str:
    bill_scope = normalize_bill_scope(bill_scope)
    if bill_scope == BILL_SCOPE_LAST:
        return f"Last {normalize_bill_limit(bill_limit)} Transactions"
    return bill_scope

def get_customer_bill_data(df: pd.DataFrame, customer_name: str, bill_scope: str | None = None, bill_limit: int | float | str | None = None) -> pd.DataFrame:
    if df.empty or not customer_name:
        return pd.DataFrame()
    mask = df["customer_name"].astype(str).str.casefold().eq(str(customer_name).casefold())
    bill_df = df[mask].copy()
    bill_scope = normalize_bill_scope(bill_scope)
    if bill_scope == BILL_SCOPE_PENDING:
        bill_df = bill_df[bill_df["pending_amount"].map(money_value) > 0]
    elif bill_scope == BILL_SCOPE_LAST:
        bill_df = bill_df.sort_values(["sale_date", "id"], ascending=[False, False]).head(normalize_bill_limit(bill_limit))
    return bill_df.sort_values(["sale_date", "id"], ascending=[True, True]).copy()

def bill_status(row: pd.Series) -> str:
    pending = money_value(row.get("pending_amount"))
    return "PAID [x]" if pending <= 0 else "PENDING"

def bill_paid_date(row: pd.Series) -> str:
    for field in ("last_payment_date", "payment_date"):
        value = str(row.get(field, "") or "").strip()
        if value:
            parsed = pd.to_datetime(value, errors="coerce")
            return parsed.strftime("%d %b %Y") if pd.notna(parsed) else value
    return "-"

def bill_totals(hist: pd.DataFrame) -> dict:
    if hist.empty:
        return {"total_bill": 0.0, "total_paid": 0.0, "total_pending": 0.0}
    return {
        "total_bill": float(hist["selling_price"].map(money_value).sum()),
        "total_paid": float(hist["amount_paid"].map(money_value).sum()),
        "total_pending": float(hist["pending_amount"].map(money_value).sum()),
    }

def get_next_bill_id(bill_date: date | None = None) -> str:
    bill_date = bill_date or date.today()
    day_key = bill_date.strftime("%Y%m%d")
    counter = get_db()["counters"].find_one_and_update(
        {"_id": f"bill_id_{day_key}"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    return f"SKB-{day_key}-{int(counter['seq']):04d}"

def bill_history_collection():
    return get_db()["bill_history"]

def build_bill_history_doc(df: pd.DataFrame, customer_name: str, bill_date: date | None = None, bill_scope: str | None = None, bill_limit: int | float | str | None = None) -> dict:
    bill_date = bill_date or date.today()
    bill_scope = normalize_bill_scope(bill_scope)
    bill_limit = normalize_bill_limit(bill_limit)
    hist = get_customer_bill_data(df, customer_name, bill_scope=bill_scope, bill_limit=bill_limit)
    if hist.empty:
        if bill_scope == BILL_SCOPE_PENDING:
            raise ValueError("No pending transactions found for this customer.")
        raise ValueError("No purchases found for this customer.")

    totals = bill_totals(hist)
    bill_id = get_next_bill_id(bill_date)
    customer_phone = first_nonempty(hist.get("customer_phone", pd.Series(dtype=str)).tolist())
    rows = []
    for _, row in hist.iterrows():
        sale_dt = pd.to_datetime(row.get("sale_date"), errors="coerce")
        rows.append({
            "sale_id": int(row.get("id", 0) or 0),
            "sale_date": str(sale_dt.date()) if pd.notna(sale_dt) else "",
            "category": clean_text_cell(row.get("product_category")),
            "description": clean_text_cell(row.get("product_description")),
            "bill_amount": round(money_value(row.get("selling_price")), 2),
            "paid_amount": round(money_value(row.get("amount_paid")), 2),
            "pending_amount": round(money_value(row.get("pending_amount")), 2),
            "paid_date": bill_paid_date(row),
            "status": bill_status(row),
        })

    return {
        "bill_id": bill_id,
        "bill_date": str(bill_date),
        "customer_name": str(customer_name),
        "customer_phone": customer_phone,
        "bill_scope": bill_scope,
        "bill_limit": bill_limit if bill_scope == BILL_SCOPE_LAST else None,
        "bill_scope_label": bill_scope_label(bill_scope, bill_limit),
        "purchase_count": len(hist),
        "purchase_ids": [int(v) for v in hist["id"].dropna().tolist()],
        "items": rows,
        "total_bill": round(totals["total_bill"], 2),
        "total_paid": round(totals["total_paid"], 2),
        "total_pending": round(totals["total_pending"], 2),
        "upi_id": "9176619942@ybl",
        "generated_at": str(datetime.now()),
        "generated_by": st.session_state.get("username", "Admin"),
    }

def create_bill_history_record(df: pd.DataFrame, customer_name: str, bill_date: date | None = None, bill_scope: str | None = None, bill_limit: int | float | str | None = None) -> dict:
    doc = build_bill_history_doc(df, customer_name, bill_date=bill_date, bill_scope=bill_scope, bill_limit=bill_limit)
    bill_history_collection().insert_one(doc.copy())
    return doc

def get_bill_history(search: str = "", limit: int = 100) -> list[dict]:
    query = {}
    search = str(search or "").strip()
    if search:
        escaped = re.escape(search)
        query = {"$or": [
            {"bill_id": {"$regex": escaped, "$options": "i"}},
            {"customer_name": {"$regex": escaped, "$options": "i"}},
            {"customer_phone": {"$regex": escaped, "$options": "i"}},
        ]}
    return list(bill_history_collection().find(query, {"_id": 0}).sort("generated_at", -1).limit(limit))

def extract_passbook_pdf_text(file_bytes: bytes) -> str:
    try:
        import pdfplumber
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            return "\n".join(page.extract_text(x_tolerance=1, y_tolerance=3) or "" for page in pdf.pages)
    except ImportError:
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise RuntimeError("Install pdfplumber or pypdf to read passbook PDFs.") from exc
        reader = PdfReader(BytesIO(file_bytes))
        return "\n".join(page.extract_text() or "" for page in reader.pages)

def extract_passbook_pdf_tables(file_bytes: bytes) -> list[list[str]]:
    try:
        import pdfplumber
    except ImportError:
        return []

    rows = []
    try:
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                for table in page.extract_tables() or []:
                    for row in table or []:
                        cleaned = [clean_text_cell(cell) for cell in (row or [])]
                        if any(cleaned):
                            rows.append(cleaned)
    except Exception:
        return []
    return rows

def passbook_field(text: str, pattern: str) -> str:
    match = re.search(pattern, text, flags=re.IGNORECASE)
    return clean_text_cell(match.group(1)) if match else ""

def passbook_amount(value: str) -> float:
    try:
        return float(str(value or "0").replace(",", ""))
    except ValueError:
        return 0.0

def passbook_date_value(value: str) -> date:
    parsed = pd.to_datetime(value, format="%d/%m/%Y", errors="coerce")
    if pd.isna(parsed):
        parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    return parsed.date() if pd.notna(parsed) else date.today()

def passbook_transaction_name(description: str) -> str:
    desc = clean_text_cell(description)
    desc = re.sub(r"^(TO|BY)\s+", "", desc, flags=re.IGNORECASE).strip()
    parts = [clean_text_cell(part) for part in desc.split("/")]
    if len(parts) >= 4 and parts[1].upper() in {"DR", "CR"}:
        return parts[3] or "Unknown"
    if ":" in desc:
        return clean_text_cell(desc.split(":", 1)[0]) or "Unknown"
    return desc[:60] or "Unknown"

def passbook_vendor_key(name: str) -> str:
    return re.sub(r"\s+", " ", clean_text_cell(name)).strip().casefold()

def passbook_vendor_collection():
    return get_db()["passbook_vendors"]

def get_saved_passbook_vendors() -> list[str]:
    docs = list(passbook_vendor_collection().find({}, {"_id": 0, "name": 1}).sort("name", 1))
    return [clean_text_cell(doc.get("name")) for doc in docs if clean_text_cell(doc.get("name"))]

def save_passbook_vendor(name: str) -> bool:
    clean_name = clean_text_cell(name)
    key = passbook_vendor_key(clean_name)
    if not key:
        return False
    passbook_vendor_collection().update_one(
        {"_id": key},
        {
            "$set": {
                "name": clean_name,
                "updated_at": str(datetime.now()),
                "updated_by": st.session_state.get("username", "Admin"),
            },
            "$setOnInsert": {"created_at": str(datetime.now())},
        },
        upsert=True,
    )
    return True

def remove_passbook_vendor(name: str) -> bool:
    key = passbook_vendor_key(name)
    if not key:
        return False
    passbook_vendor_collection().delete_one({"_id": key})
    return True

def work_notes_collection():
    return get_db()["work_notes"]

def get_work_notes(limit: int = 200) -> list[dict]:
    return list(work_notes_collection().find({}, {"_id": 0}).sort([("work_date", -1), ("created_at", -1)]).limit(limit))

def save_work_note(work_date: date, note: str) -> int:
    note = clean_text_cell(note)[:1000]
    existing = work_notes_collection().find_one({}, sort=[("id", -1)], projection={"id": 1, "_id": 0})
    note_id = int(existing.get("id", 0)) + 1 if existing else 1
    work_notes_collection().insert_one({
        "id": note_id,
        "work_date": str(work_date),
        "note": note,
        "created_at": str(datetime.now()),
        "created_by": st.session_state.get("username", "Admin"),
    })
    return note_id

def delete_work_note(note_id: int):
    work_notes_collection().delete_one({"id": int(note_id)})

def passbook_transaction_row(txn_date: str, description: str, debit: str = "", credit: str = "", balance: str = "") -> dict | None:
    txn_date = clean_text_cell(txn_date)
    description = clean_text_cell(description)
    if not re.match(r"^\d{2}/\d{2}/\d{4}$", txn_date):
        return None
    if not re.match(r"^(TO|BY)\b", description, flags=re.IGNORECASE):
        return None
    debit_value = passbook_amount(debit)
    credit_value = passbook_amount(credit)
    balance_value = passbook_amount(balance)
    return {
        "Date": txn_date,
        "Name": passbook_transaction_name(description),
        "Description": description,
        "Debit": debit_value,
        "Credit": credit_value,
        "Balance": balance_value,
    }

def parse_passbook_table_transactions(table_rows: list[list[str]]) -> list[dict]:
    amount_re = re.compile(r"^\d{1,3}(?:,\d{3})*(?:\.\d{2})$|^\d+(?:\.\d{2})$")
    rows = []
    for row in table_rows:
        cells = [clean_text_cell(cell) for cell in row]
        if len(cells) < 2:
            continue
        if cells[0].upper() == "DATE" or cells[1].upper() == "DESCRIPTION":
            continue
        amount_cells = [cell for cell in cells[2:] if amount_re.match(cell)]
        txn_amount = amount_cells[0] if len(amount_cells) >= 2 else ""
        balance = amount_cells[-1] if amount_cells else ""
        direction = cells[1].split(" ", 1)[0].upper() if cells[1] else ""
        parsed = passbook_transaction_row(
            cells[0],
            cells[1],
            txn_amount if direction == "TO" else "",
            txn_amount if direction == "BY" else "",
            balance,
        )
        if parsed:
            rows.append(parsed)
    return rows

def parse_passbook_transactions(text: str) -> list[dict]:
    amount = r"\d{1,3}(?:,\d{3})*(?:\.\d{2})|\d+(?:\.\d{2})"
    row_re = re.compile(rf"^(\d{{2}}/\d{{2}}/\d{{4}})\s+(TO|BY)\s+(.+?)\s+({amount})\s+({amount})$", re.IGNORECASE)
    rows = []
    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", str(raw_line or "")).strip()
        match = row_re.match(line)
        if not match:
            continue
        txn_date, direction, description, txn_amount, balance = match.groups()
        direction = direction.upper()
        full_description = f"{direction} {description}".strip()
        parsed = passbook_transaction_row(
            txn_date,
            full_description,
            txn_amount if direction == "TO" else "",
            txn_amount if direction == "BY" else "",
            balance,
        )
        if parsed:
            rows.append(parsed)
    return rows

def merge_passbook_transactions(*groups: list[dict]) -> list[dict]:
    merged = []
    seen = set()
    for group in groups:
        for row in group:
            key = (
                row.get("Date", ""),
                row.get("Description", ""),
                round(money_value(row.get("Debit")), 2),
                round(money_value(row.get("Credit")), 2),
                round(money_value(row.get("Balance")), 2),
            )
            if key in seen:
                continue
            seen.add(key)
            merged.append(row)
    return merged

@st.cache_data(show_spinner=False)
def parse_passbook_pdf(file_bytes: bytes, filename: str) -> dict:
    text = extract_passbook_pdf_text(file_bytes)
    table_rows = extract_passbook_pdf_tables(file_bytes)
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines() if str(line).strip()]

    customer_name = passbook_field(text, r"CUSTOMER\s+DETAILS\s*:\s*(.+)")
    customer_name = re.sub(r"\s*\.$", "", customer_name).strip()
    address_lines = []
    for idx, line in enumerate(lines):
        if re.search(r"CUSTOMER\s+DETAILS\s*:", line, flags=re.IGNORECASE):
            for next_line in lines[idx + 1:]:
                if re.search(r"^(Statement Date|STATEMENT OF ACCOUNT|DATE DESCRIPTION)", next_line, flags=re.IGNORECASE):
                    break
                address_lines.append(next_line)
            break

    branch_address = ""
    for idx, line in enumerate(lines):
        if re.search(r"^BRANCH\s*:", line, flags=re.IGNORECASE) and idx + 1 < len(lines):
            candidate = lines[idx + 1]
            if not re.search(r"^(ACCOUNT|IFSC|CUSTOMER|STATEMENT)", candidate, flags=re.IGNORECASE):
                branch_address = candidate
            break

    table_transactions = parse_passbook_table_transactions(table_rows)
    text_transactions = parse_passbook_transactions(text)
    transactions = merge_passbook_transactions(table_transactions, text_transactions)
    total_debit = sum(row["Debit"] for row in transactions)
    total_credit = sum(row["Credit"] for row in transactions)
    latest_balance = transactions[-1]["Balance"] if transactions else 0.0

    return {
        "filename": filename,
        "bank": lines[0] if lines else "",
        "branch": passbook_field(text, r"BRANCH\s*:\s*(.+)"),
        "branch_address": branch_address,
        "account_no": passbook_field(text, r"ACCOUNT\s+NO\s*:\s*([^\n]+)"),
        "account_no_15": passbook_field(text, r"ACCOUNT\s+NO\(15\s+DIGIT\)\s*:\s*([^\n]+)"),
        "ifsc": passbook_field(text, r"IFSC\s*:\s*([^\n]+)"),
        "account_type": passbook_field(text, r"ACCOUNT\s+TYPE\s*:\s*([^\n]*)"),
        "customer_name": customer_name or filename,
        "address": ", ".join(address_lines),
        "statement_date": passbook_field(text, r"Statement\s+Date\s*:\s*([^\n]+)"),
        "statement_period": passbook_field(text, r"STATEMENT\s+OF\s+ACCOUNT\s+from\s+(.+)"),
        "transactions": transactions,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "latest_balance": latest_balance,
        "raw_text": text,
    }

def page_passbook_reader():
    page_header("Passbook Reader", "PDF Statement Extractor")
    uploads = st.file_uploader(
        "Upload Passbook PDF",
        type=["pdf"],
        accept_multiple_files=True,
        key="passbook_pdf_uploads",
    )
    if not uploads:
        st.info("Upload one or more passbook PDFs. The app will read account details and build a name filter from the transaction table.")
        return

    passbooks = []
    for uploaded in uploads:
        try:
            passbooks.append(parse_passbook_pdf(uploaded.getvalue(), uploaded.name))
        except RuntimeError as exc:
            st.error(str(exc))
        except Exception as exc:
            st.error(f"Could not read {uploaded.name}: {exc}")

    if not passbooks:
        return

    all_rows = []
    for idx, pb in enumerate(passbooks):
        for row in pb.get("transactions", []):
            row_copy = dict(row)
            row_copy["Passbook"] = pb.get("customer_name", "Unknown")
            row_copy["Statement Date"] = pb.get("statement_date", "")
            row_copy["Account"] = pb.get("account_no_15") or pb.get("account_no") or ""
            row_copy["_passbook_idx"] = idx
            all_rows.append(row_copy)

    txn_df = pd.DataFrame(all_rows)
    saved_vendors = get_saved_passbook_vendors()
    saved_vendor_keys = {passbook_vendor_key(name) for name in saved_vendors}

    f1, f2, f3 = st.columns([1.4, 1.4, 1])
    with f2:
        selected_idx = st.selectbox(
            "Passbook",
            list(range(len(passbooks))),
            format_func=lambda idx: f"{passbooks[idx].get('customer_name') or 'Unknown'} - {passbooks[idx].get('statement_date') or passbooks[idx].get('account_no_15') or passbooks[idx].get('filename')}",
            key="passbook_file_filter",
        )
    passbook_rows = txn_df[txn_df["_passbook_idx"] == selected_idx].copy() if not txn_df.empty else pd.DataFrame()
    names = []
    if not passbook_rows.empty and "Name" in passbook_rows.columns:
        names = sorted([name for name in passbook_rows["Name"].dropna().astype(str).unique() if name.strip()], key=str.casefold)
    with f1:
        selected_name = st.selectbox("Name Filter", ["All Names", "Saved Vendors"] + names, key="passbook_transaction_name_filter")
    with f3:
        direction_filter = st.selectbox("Type", ["All", "Credit", "Debit"], key="passbook_direction_filter")

    pb = passbooks[selected_idx]
    filtered = passbook_rows.copy()
    if selected_name == "Saved Vendors" and not filtered.empty:
        filtered = filtered[filtered["Name"].astype(str).map(passbook_vendor_key).isin(saved_vendor_keys)]
    elif selected_name != "All Names" and not filtered.empty:
        filtered = filtered[filtered["Name"].astype(str).str.casefold().eq(str(selected_name).casefold())]
    if direction_filter == "Credit" and not filtered.empty:
        filtered = filtered[filtered["Credit"].map(money_value) > 0]
    elif direction_filter == "Debit" and not filtered.empty:
        filtered = filtered[filtered["Debit"].map(money_value) > 0]

    total_credit = float(filtered["Credit"].map(money_value).sum()) if not filtered.empty else 0.0
    total_debit = float(filtered["Debit"].map(money_value).sum()) if not filtered.empty else 0.0
    latest_balance = float(filtered["Balance"].map(money_value).iloc[-1]) if not filtered.empty else 0.0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Rows", len(filtered))
    m2.metric("Credits", f"₹{total_credit:,.0f}")
    m3.metric("Debits", f"₹{total_debit:,.0f}")
    m4.metric("Last Balance", f"₹{latest_balance:,.2f}")

    a1, a2, a3 = st.columns([1, 1, 2])
    selected_name_key = passbook_vendor_key(selected_name)
    selected_is_vendor = selected_name_key in saved_vendor_keys
    with a1:
        if selected_name not in ["All Names", "Saved Vendors"]:
            if selected_is_vendor:
                if st.button("Remove Vendor", key="passbook_remove_vendor", width="stretch"):
                    remove_passbook_vendor(selected_name)
                    st.success(f"{selected_name} removed from saved vendors.")
                    st.rerun()
            else:
                if st.button("Mark as Vendor", key="passbook_mark_vendor", width="stretch"):
                    if save_passbook_vendor(selected_name):
                        st.success(f"{selected_name} saved as vendor.")
                        st.rerun()
        else:
            st.button("Mark as Vendor", key="passbook_mark_vendor_disabled", disabled=True, width="stretch")
    with a2:
        if saved_vendors:
            st.metric("Saved Vendors", len(saved_vendors))
        else:
            st.metric("Saved Vendors", 0)
    with a3:
        if selected_name == "Saved Vendors" and not saved_vendors:
            st.info("No vendors saved yet. Select a name, then click Mark as Vendor.")
        elif selected_name == "Saved Vendors":
            st.caption("Showing transactions for saved vendors only.")

    rule_sm()
    sec("Transactions")
    if filtered.empty:
        st.info("No transaction rows match this filter.")
    else:
        show_cols = ["Date", "Name", "Description", "Debit", "Credit", "Balance", "Passbook", "Statement Date", "Account"]
        show = filtered[[c for c in show_cols if c in filtered.columns]].copy()
        st.dataframe(show, width="stretch", hide_index=True, height=460)
        safe_name = re.sub(r"[^0-9A-Za-z]+", "_", selected_name if selected_name != "All Names" else pb.get("customer_name", "passbook")).strip("_").lower()
        st.download_button(
            "Download Filtered CSV",
            data=show.to_csv(index=False),
            file_name=f"passbook_{safe_name or 'transactions'}_{date.today()}.csv",
            mime="text/csv",
            width="stretch",
            key="passbook_filtered_csv",
        )
        passbook_context = "\n\n".join([
            f"Selected passbook customer: {pb.get('customer_name', 'Unknown')}",
            f"Statement date: {pb.get('statement_date', '')}",
            f"Account: {pb.get('account_no_15') or pb.get('account_no') or ''}",
            f"Selected name filter: {selected_name}",
            f"Selected type filter: {direction_filter}",
            "Account details:\n" + pd.DataFrame([
                {"Field": k, "Value": v}
                for k, v in pb.items()
                if k not in {"transactions", "raw_text"}
            ]).to_csv(index=False),
            "Filtered passbook transactions:\n" + show.to_csv(index=False),
        ])
        render_ai_action_panel(
            "AI PDF Extraction Assistant",
            passbook_context,
            "passbook_ai",
            {
                "Summarize transactions": "Summarize these filtered passbook transactions. Give total debit, total credit, repeated names, and anything that needs attention.",
                "Find vendor payments": "Identify which rows look like boutique vendor purchases. Group by name, total the debit amount, and suggest which names I should save as vendors.",
                "Suggest sale entries": "Suggest which debit transactions can be converted into boutique sale records. For each one, suggest vendor, category, buying price, and notes.",
                "Check extraction quality": "Check whether names, dates, debits, credits, and balances look extracted correctly. List suspicious or missing values only.",
                "Categorize names": "Group the transaction names into likely categories such as boutique vendor, household expense, bank charge, food/grocery, transfer, or unknown.",
            },
        )

        rule_sm()
        sec("Add Sale From Transaction")
        selectable = show.reset_index(drop=True).copy()
        selectable["_option"] = selectable.apply(
            lambda row: f"{row.get('Date', '-')} - {row.get('Name', '-')} - Debit ₹{money_value(row.get('Debit')):,.2f} - Credit ₹{money_value(row.get('Credit')):,.2f}",
            axis=1,
        )
        selected_txn_idx = st.selectbox(
            "Select Transaction",
            selectable.index.tolist(),
            format_func=lambda idx: selectable.loc[idx, "_option"],
            key="passbook_sale_txn_select",
        )
        txn = selectable.loc[selected_txn_idx]
        txn_amount = money_value(txn.get("Debit")) if money_value(txn.get("Debit")) > 0 else money_value(txn.get("Credit"))

        sale_customer_type = st.radio(
            "Customer Type",
            ["Existing Customer", "New Customer"],
            horizontal=True,
            key="passbook_sale_customer_type",
        )
        existing_customers = get_existing_customers_with_phone()
        existing_customer_map = {str(customer.get("_id", "")): customer for customer in existing_customers}
        selected_existing_customer = ""
        selected_existing_phone = ""
        if sale_customer_type == "Existing Customer":
            if existing_customers:
                selected_existing_customer = st.selectbox(
                    "Search Existing Customer",
                    [str(customer.get("_id", "")) for customer in existing_customers],
                    format_func=lambda name: f"{name} - {existing_customer_map.get(name, {}).get('phone') or 'No phone'}",
                    key="passbook_sale_existing_customer",
                )
                selected_existing_phone = existing_customer_map.get(selected_existing_customer, {}).get("phone", "")
            else:
                st.warning("No existing customers found. Use New Customer.")
                sale_customer_type = "New Customer"

        with st.form("passbook_add_sale_form"):
            st.caption(f"Vendor, buying price, and sale date are filled from: {txn.get('Description', '')}")
            c1, c2, c3 = st.columns(3)
            with c1:
                if sale_customer_type == "Existing Customer":
                    sale_customer = st.text_input("Customer Name *", value=selected_existing_customer, disabled=True, key="passbook_sale_customer_existing")
                else:
                    sale_customer = st.text_input("Customer Name *", key="passbook_sale_customer_new")
            with c2:
                sale_phone = st.text_input("Phone", value=selected_existing_phone, placeholder="+91 XXXXXXXXXX", key=f"passbook_sale_phone_{sale_customer_type.replace(' ', '_').lower()}")
            with c3:
                sale_date = st.date_input("Sale Date", value=passbook_date_value(txn.get("Date")), key="passbook_sale_date")

            p1, p2, p3 = st.columns(3)
            with p1:
                sale_category = st.selectbox("Category *", CATEGORIES, key="passbook_sale_category")
            with p2:
                sale_vendor = st.text_input("Vendor", value=str(txn.get("Name", "") or ""), key="passbook_sale_vendor")
            with p3:
                sale_qty = st.number_input("Quantity", min_value=1, step=1, value=1, key="passbook_sale_qty")

            sale_desc = st.text_area("Description", value=str(txn.get("Description", "") or ""), height=70, key="passbook_sale_desc")

            pr1, pr2, pr3, pr4 = st.columns(4)
            with pr1:
                sale_buy = st.number_input("Buying Price (₹) *", min_value=0.0, step=1.0, value=float(txn_amount), key="passbook_sale_buy")
            with pr2:
                sale_sell = st.number_input("Selling Price (₹) *", min_value=0.0, step=1.0, value=float(txn_amount), key="passbook_sale_sell")
            with pr3:
                sale_paid = st.number_input("Amount Paid (₹)", min_value=0.0, step=1.0, value=0.0, key="passbook_sale_paid")
            with pr4:
                sale_pm = st.selectbox("Payment Method", PAYMENT_METHODS, index=PAYMENT_METHODS.index("UPI") if "UPI" in PAYMENT_METHODS else 0, key="passbook_sale_payment_method")

            sale_pending = max(round(sale_sell - sale_paid, 2), 0.0)
            sale_profit = round((sale_sell - sale_buy) * sale_qty, 2)
            sm1, sm2, sm3 = st.columns(3)
            sm1.metric("Pending", f"₹{sale_pending:,.2f}")
            sm2.metric("Profit", f"₹{sale_profit:,.2f}")
            sm3.metric("Total Value", f"₹{sale_sell * sale_qty:,.2f}")

            sale_notes = st.text_area("Notes", value=f"From passbook transaction: {txn.get('Description', '')}", height=60, key="passbook_sale_notes")
            save_sale = st.form_submit_button("Save Sale", width="stretch")

        if sale_customer_type == "Existing Customer":
            sale_customer = selected_existing_customer
            sale_phone = selected_existing_phone

        if save_sale:
            errs = []
            if not str(sale_customer or "").strip():
                errs.append("Customer name is required.")
            if sale_buy <= 0:
                errs.append("Buying price must be > 0.")
            if sale_sell <= 0:
                errs.append("Selling price must be > 0.")
            if sale_paid > sale_sell:
                errs.append("Amount paid cannot exceed selling price.")
            if not str(sale_vendor or "").strip():
                errs.append("Vendor is required.")
            if len(str(sale_customer)) > 120:
                errs.append("Customer name must be under 120 characters.")
            if len(str(sale_phone)) > 20:
                errs.append("Phone number must be under 20 characters.")
            if len(str(sale_vendor)) > 100:
                errs.append("Vendor name must be under 100 characters.")
            if errs:
                for err in errs:
                    st.error(err)
            else:
                get_col().insert_one({
                    "id": get_next_id(),
                    "customer_name": str(sale_customer).strip()[:120],
                    "customer_phone": normalize_phone(sale_phone),
                    "sale_date": str(sale_date),
                    "vendor": str(sale_vendor).strip()[:100],
                    "product_category": sale_category,
                    "product_description": str(sale_desc).strip()[:500],
                    "quantity": int(sale_qty),
                    "buying_price": round(float(sale_buy), 2),
                    "selling_price": round(float(sale_sell), 2),
                    "amount_paid": round(float(sale_paid), 2),
                    "pending_amount": sale_pending,
                    "payment_received": 1 if sale_pending == 0 else 0,
                    "delay_status": 0,
                    "payment_method": sale_pm,
                    "notes": str(sale_notes).strip()[:500],
                    "passbook_source": {
                        "date": str(txn.get("Date", "")),
                        "name": str(txn.get("Name", "")),
                        "description": str(txn.get("Description", "")),
                        "debit": money_value(txn.get("Debit")),
                        "credit": money_value(txn.get("Credit")),
                        "balance": money_value(txn.get("Balance")),
                    },
                    "created_at": str(datetime.now()),
                })
                invalidate_cache()
                st.success(f"Sale added for {str(sale_customer).strip()} from passbook transaction.")
                st.rerun()

    with st.expander("Saved Vendors", expanded=False):
        if saved_vendors:
            vendor_df = pd.DataFrame({"Vendor Name": saved_vendors})
            st.dataframe(vendor_df, width="stretch", hide_index=True, height=220)
            remove_name = st.selectbox("Remove Saved Vendor", saved_vendors, key="passbook_saved_vendor_remove_select")
            if st.button("Remove Selected Vendor", key="passbook_saved_vendor_remove", width="stretch"):
                remove_passbook_vendor(remove_name)
                st.success(f"{remove_name} removed from saved vendors.")
                st.rerun()
        else:
            st.info("No saved vendors yet.")

    with st.expander("All Names Found", expanded=False):
        if names:
            name_counts = txn_df[txn_df["_passbook_idx"] == selected_idx]["Name"].value_counts().reset_index()
            name_counts.columns = ["Name", "Transactions"]
            st.dataframe(name_counts, width="stretch", hide_index=True, height=320)
        else:
            st.info("No names found in the uploaded passbook transactions.")

def make_upi_qr_png(amount: float = 0.0) -> BytesIO:
    try:
        import qrcode
    except ImportError as exc:
        raise RuntimeError("Install qrcode[pil] to generate payment QR codes.") from exc

    from urllib.parse import quote
    upi_id = "9176619942@ybl"
    uri = f"upi://pay?pa={quote(upi_id)}&pn={quote(BRAND_NAME)}&cu=INR"
    if amount > 0:
        uri += f"&am={amount:.2f}"
    qr = qrcode.QRCode(box_size=8, border=2)
    qr.add_data(uri)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    out = BytesIO()
    img.save(out, format="PNG")
    out.seek(0)
    return out

def generate_customer_bill_pdf(df: pd.DataFrame, customer_name: str, bill_date: date | None = None, bill_id: str = "", bill_scope: str | None = None, bill_limit: int | float | str | None = None) -> BytesIO:
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise RuntimeError("Install reportlab to generate PDF bills.") from exc

    bill_date = bill_date or date.today()
    bill_scope = normalize_bill_scope(bill_scope)
    bill_limit = normalize_bill_limit(bill_limit)
    hist = get_customer_bill_data(df, customer_name, bill_scope=bill_scope, bill_limit=bill_limit)
    if hist.empty:
        if bill_scope == BILL_SCOPE_PENDING:
            raise ValueError("No pending transactions found for this customer.")
        raise ValueError("No purchases found for this customer.")

    customer_phone = first_nonempty(hist.get("customer_phone", pd.Series(dtype=str)).tolist())
    totals = bill_totals(hist)
    total_bill = totals["total_bill"]
    total_paid = totals["total_paid"]
    total_pending = totals["total_pending"]

    out = BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"Bill - {customer_name}",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("BillTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=20, leading=24, textColor=colors.HexColor("#0F172A"), spaceAfter=4)
    sub_style = ParagraphStyle("BillSub", parent=styles["Normal"], fontSize=9, leading=12, textColor=colors.HexColor("#475569"))
    right_style = ParagraphStyle("Right", parent=styles["Normal"], fontSize=9, alignment=TA_RIGHT, textColor=colors.HexColor("#475569"))
    center_style = ParagraphStyle("Center", parent=styles["Normal"], fontSize=9, alignment=TA_CENTER, textColor=colors.HexColor("#0F172A"))

    story = []
    if os.path.exists(BILL_SYMBOL_PATH):
        brand_symbol = Image(BILL_SYMBOL_PATH, width=20 * mm, height=20 * mm)
    else:
        brand_symbol = Paragraph("<b>SK</b>", center_style)
    brand_header = Table(
        [[
            brand_symbol,
            [Paragraph(BRAND_NAME, title_style), Paragraph("Customer Purchase Bill", sub_style)],
        ]],
        colWidths=[24 * mm, 88 * mm],
    )
    brand_header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    header = Table(
        [[
            brand_header,
            [
                Paragraph(f"<b>Bill ID:</b> {html_escape(bill_id or '-')}", right_style),
                Paragraph(f"<b>Bill Date:</b> {bill_date.strftime('%d %b %Y')}", right_style),
                Paragraph(f"<b>UPI:</b> 9176619942@ybl", right_style),
            ],
        ]],
        colWidths=[112 * mm, 56 * mm],
    )
    header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW", (0, 0), (-1, -1), 0.6, colors.HexColor("#CBD5E1")),
    ]))
    story.append(header)
    story.append(Spacer(1, 8))

    customer_block = Table(
        [[
            Paragraph(f"<b>Customer:</b> {html_escape(str(customer_name))}", sub_style),
            Paragraph(f"<b>Phone:</b> {html_escape(customer_phone or '-')}", sub_style),
            Paragraph(f"<b>Bill ID:</b> {html_escape(bill_id or '-')}", sub_style),
            Paragraph(f"<b>Type:</b> {html_escape(bill_scope_label(bill_scope, bill_limit))}", sub_style),
        ]],
        colWidths=[54 * mm, 36 * mm, 42 * mm, 36 * mm],
    )
    customer_block.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#DBEAFE")),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
        ("PADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(customer_block)
    story.append(Spacer(1, 10))
    story.append(Paragraph("Bill Contents", ParagraphStyle("Section", parent=styles["Heading2"], fontSize=12, textColor=colors.HexColor("#0F172A"), spaceAfter=6)))

    table_data = [["Date", "Item / Category", "Bill", "Paid", "Paid Date", "Status"]]
    for _, row in hist.iterrows():
        sale_dt = pd.to_datetime(row.get("sale_date"), errors="coerce")
        date_text = sale_dt.strftime("%d %b %Y") if pd.notna(sale_dt) else "-"
        desc = clean_text_cell(row.get("product_description")) or clean_text_cell(row.get("product_category")) or "-"
        category = clean_text_cell(row.get("product_category"))
        item = f"{html_escape(desc)}<br/><font color='#64748B'>{html_escape(category)}</font>"
        table_data.append([
            date_text,
            Paragraph(item, sub_style),
            f"Rs {money_value(row.get('selling_price')):,.2f}",
            f"Rs {money_value(row.get('amount_paid')):,.2f}",
            bill_paid_date(row),
            bill_status(row),
        ])

    purchase_table = Table(table_data, colWidths=[22 * mm, 76 * mm, 22 * mm, 22 * mm, 25 * mm, 22 * mm], repeatRows=1)
    purchase_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAF1FF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (2, 1), (3, -1), "RIGHT"),
        ("ALIGN", (5, 1), (5, -1), "CENTER"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("PADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(purchase_table)
    story.append(Spacer(1, 10))

    qr_buf = make_upi_qr_png(total_pending)
    totals = Table(
        [[
            Image(qr_buf, width=34 * mm, height=34 * mm),
            [
                Paragraph("<b>PhonePe / UPI Payment</b>", sub_style),
                Paragraph("UPI ID: 9176619942@ybl", sub_style),
                Paragraph(f"QR amount: Rs {total_pending:,.2f}" if total_pending > 0 else "No pending amount", sub_style),
            ],
            [
                Paragraph(f"<b>Total Pending:</b> Rs {total_pending:,.2f}", right_style),
            ],
        ]],
        colWidths=[38 * mm, 62 * mm, 68 * mm],
    )
    totals.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#CBD5E1")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("PADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(totals)
    story.append(Spacer(1, 8))
    if total_pending > 0:
        story.append(Paragraph(f"Pending amount to be paid: <b>Rs {total_pending:,.2f}</b>", ParagraphStyle("Pending", parent=center_style, fontSize=10, textColor=colors.HexColor("#B91C1C"))))
    else:
        story.append(Paragraph("All listed purchases are paid.", ParagraphStyle("Paid", parent=center_style, fontSize=10, textColor=colors.HexColor("#047857"))))

    doc.build(story)
    out.seek(0)
    return out

def render_customer_bill_download(df: pd.DataFrame, customer_name: str, key: str, label: str = "Generate Bill PDF", bill_date: date | None = None, bill_scope: str | None = None, bill_limit: int | float | str | None = None):
    if bill_scope is None:
        chosen_scope = st.selectbox("Bill Type", BILL_SCOPE_OPTIONS, key=f"{key}_scope")
    else:
        chosen_scope = normalize_bill_scope(bill_scope)
    if bill_limit is None and chosen_scope == BILL_SCOPE_LAST:
        chosen_limit = st.number_input("Last Transactions", min_value=1, max_value=100, value=5, step=1, key=f"{key}_limit")
    else:
        chosen_limit = normalize_bill_limit(bill_limit)
    scope_key = re.sub(r"[^0-9A-Za-z]+", "_", bill_scope_label(chosen_scope, chosen_limit)).strip("_").lower()
    state_key = f"{key}_{scope_key}_bill_download"
    if st.button(label, key=f"{key}_create", width="stretch"):
        try:
            bill_doc = create_bill_history_record(df, customer_name, bill_date=bill_date or date.today(), bill_scope=chosen_scope, bill_limit=chosen_limit)
            bill_pdf = generate_customer_bill_pdf(df, customer_name, bill_date=bill_date or date.today(), bill_id=bill_doc["bill_id"], bill_scope=chosen_scope, bill_limit=chosen_limit)
            st.session_state[state_key] = {
                "bill_id": bill_doc["bill_id"],
                "customer_name": customer_name,
                "bill_scope": chosen_scope,
                "bill_limit": chosen_limit if chosen_scope == BILL_SCOPE_LAST else None,
                "pdf": bill_pdf.getvalue(),
            }
            st.success(f"Bill {bill_doc['bill_id']} generated and saved to history.")
        except RuntimeError as exc:
            st.error(str(exc))
            return
        except ValueError as exc:
            st.info(str(exc))
            return

    payload = st.session_state.get(state_key)
    if payload:
        st.download_button(
            f"Download {payload['bill_id']}",
            data=payload["pdf"],
            file_name=bill_file_name(payload["customer_name"], payload["bill_id"]),
            mime="application/pdf",
            key=f"{key}_download",
            width="stretch",
        )

def vendor_picker(label: str, key_prefix: str, current: str = "") -> str:
    current = str(current or "").strip()
    vendors = get_existing_vendors()
    options = [""] + vendors
    if current and current not in options:
        options.insert(1, current)
    options.append(VENDOR_MANUAL_OPTION)
    selected = st.selectbox(
        label,
        options,
        index=options.index(current) if current in options else 0,
        format_func=lambda value: "Select vendor" if value == "" else value,
        key=f"{key_prefix}_select",
    )
    if selected == VENDOR_MANUAL_OPTION:
        return st.text_input("New Vendor", value="" if current in vendors else current, key=f"{key_prefix}_manual").strip()
    return selected.strip()

# =====================================================
# PUBLIC ADD SALE PAGE
# =====================================================

def page_add_sale(public=False):
    if public:
        st.markdown("""
        <div class='pub-banner'>
            <div class='pub-banner-title'>Shree Krishna Boutique</div>
            <div class='pub-banner-sub'>◆ Record a New Sale</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        page_header("New Sale", "Record a Transaction")

    if not public:
        render_ai_sale_entry_assistant()
        rule()

    ctype = st.radio("Customer type", ["New Customer", "Existing Customer"], horizontal=True, label_visibility="collapsed")
    rule_sm()

    cname, cphone = "", ""

    if ctype == "Existing Customer":
        if public:
            existing = get_existing_customers()
        else:
            existing = get_existing_customers_with_phone()

        if existing:
            if public:
                opts = [r["_id"] for r in existing]
                sel  = st.selectbox("Select Customer", opts)
                cname = sel
                rec   = next((r for r in existing if r["_id"] == cname), {})
                ca, cb = st.columns(2)
                ca.info(f"**Name:** {cname}")
                cb.info(f"**Visits:** {rec.get('visits', '—')}")
                cphone = ""
            else:
                opts  = [f"{r['_id']}  —  {r.get('phone','') or 'No phone'}" for r in existing]
                sel   = st.selectbox("Select Customer", opts)
                cname = sel.split("  —  ")[0].strip()
                rec   = next((r for r in existing if r["_id"] == cname), {})
                cphone = rec.get("phone", "")
                ca, cb, cc = st.columns(3)
                ca.info(f"**Name:** {cname}")
                cb.info(f"**Phone:** {cphone or 'N/A'}")
                cc.info(f"**Visits:** {rec.get('visits', '—')}")
        else:
            st.warning("No existing customers found.")
            ctype = "New Customer"

    with st.form("sale_form", clear_on_submit=True):
        sec("Customer")
        c1, c2, c3 = st.columns(3)
        with c1:
            cname  = st.text_input("Customer Name *", value=cname,
                                   disabled=(ctype == "Existing Customer"))
        with c2:
            if public and ctype == "Existing Customer":
                cphone = ""
                st.text_input("Phone", value="", placeholder="(Admin access required)", disabled=True)
            else:
                phone_key_seed = "new" if ctype != "Existing Customer" else re.sub(r"[^0-9A-Za-z]+", "_", cname) or "existing"
                cphone = st.text_input("Phone", value=cphone, placeholder="+91 XXXXXXXXXX",
                                       key=f"sale_phone_{phone_key_seed}")
        with c3:
            sdate = st.date_input("Sale Date", date.today())

        sec("Product")
        p1, p2, p3 = st.columns(3)
        with p1: cat  = st.selectbox("Category *", CATEGORIES)
        with p2: vend = vendor_picker("Vendor / Supplier", "sale_vendor")
        with p3: qty  = st.number_input("Quantity", min_value=1, step=1, value=1)
        desc = st.text_area("Description", placeholder="Fabric, colour, design details…", height=70)

        sec("Pricing & Payment")
        pr1, pr2, pr3, pr4 = st.columns(4)
        with pr1: _, buy, buy_ok           = currency_input("Buying Price (₹) *", "sale_buying_price")
        with pr2: _, sell, sell_ok         = currency_input("Selling Price (₹) *", "sale_selling_price")
        with pr3: _, paid_amt, paid_ok     = currency_input("Amount Paid (₹)", "sale_amount_paid")
        with pr4:
            default_payment = "UPI" if public else app_pref("default_payment_method", "UPI")
            if default_payment not in PAYMENT_METHODS:
                default_payment = "UPI" if "UPI" in PAYMENT_METHODS else PAYMENT_METHODS[0]
            pm = st.selectbox(
                "Payment Method",
                PAYMENT_METHODS,
                index=PAYMENT_METHODS.index(default_payment),
                key="sale_payment_method",
            )

        pending_amt = max(round(sell - paid_amt, 2), 0.0)
        profit_amt  = round((sell - buy) * qty, 2)
        margin_pct  = round(profit_amt / (sell * qty) * 100, 2) if sell > 0 else 0.0

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Pending",        f"₹{pending_amt:,.2f}")
        m2.metric("Profit (Total)", f"₹{profit_amt:,.2f}")
        m3.metric("Margin",         f"{margin_pct:.1f}%")
        m4.metric("Total Value",    f"₹{sell * qty:,.2f}")

        notes = st.text_area("Notes", placeholder="Special instructions…", height=60)

        submitted = st.form_submit_button("Save Sale", width="stretch")

        if submitted:
            # ── Public form rate limiting ────────────────────────────────
            if public:
                now = time.time()
                window_start = st.session_state.get("pub_window_start", now)
                pub_count    = st.session_state.get("pub_submit_count", 0)
                # Reset counter every 60 seconds
                if now - window_start > 60:
                    st.session_state.pub_window_start  = now
                    st.session_state.pub_submit_count  = 0
                    pub_count = 0
                pub_count += 1
                st.session_state.pub_submit_count = pub_count
                if pub_count > 5:
                    st.error("Too many submissions. Please wait a minute before trying again.")
                    st.stop()

            # ── Input length guards ──────────────────────────────────────
            MAX = {"name": 120, "phone": 20, "vendor": 100, "desc": 500, "notes": 500}
            errs = []
            if len(cname) > MAX["name"]:   errs.append(f"Customer name must be under {MAX['name']} characters.")
            if len(cphone) > MAX["phone"]:  errs.append(f"Phone number must be under {MAX['phone']} characters.")
            if len(vend) > MAX["vendor"]:   errs.append(f"Vendor name must be under {MAX['vendor']} characters.")
            if len(desc) > MAX["desc"]:     errs.append(f"Description must be under {MAX['desc']} characters.")
            if len(notes) > MAX["notes"]:   errs.append(f"Notes must be under {MAX['notes']} characters.")

            if not cname.strip():  errs.append("Customer name is required.")
            if not buy_ok:         errs.append("Buying price must be a valid number.")
            elif buy  <= 0:        errs.append("Buying price must be > 0.")
            if not sell_ok:        errs.append("Selling price must be a valid number.")
            elif sell <= 0:        errs.append("Selling price must be > 0.")
            if not paid_ok:        errs.append("Amount paid must be a valid number.")
            elif sell_ok and paid_amt > sell:
                errs.append("Amount paid cannot exceed selling price.")
            if buy_ok and sell_ok and sell < buy:
                st.warning("Selling price is below buying price — this sale will be a loss.")

            if errs:
                for e in errs: st.error(e)
            else:
                get_col().insert_one({
                    "id":                  get_next_id(),
                    "customer_name":       cname.strip()[:120],
                    "customer_phone":      normalize_phone(cphone),
                    "sale_date":           str(sdate),
                    "vendor":              vend.strip()[:100],
                    "product_category":    cat,
                    "product_description": desc.strip()[:500],
                    "quantity":            qty,
                    "buying_price":        round(buy, 2),
                    "selling_price":       round(sell, 2),
                    "amount_paid":         round(paid_amt, 2),
                    "pending_amount":      pending_amt,
                    "payment_received":    1 if pending_amt == 0 else 0,
                    "delay_status":        0,
                    "payment_method":      pm,
                    "notes":               notes.strip()[:500],
                    "created_at":          str(datetime.now()),
                })
                invalidate_cache()
                st.success(f"✓ Sale recorded for {cname.strip()}.")
                st.balloons()
                st.rerun()

# =====================================================
# AUTH HELPERS
# =====================================================

_MAX_ATTEMPTS   = 5
_LOCKOUT_SECS   = 300   # 5-minute lockout after max attempts
_BACKOFF_BASE   = 1.5   # seconds — doubles each attempt after 1st failure
_FACE_MATCH_TOLERANCE = 0.46
_TEMP_QR_TYPE = "boutique_temp_login"
_FACE_VECTOR_VERSION = "dlib_128d_v1"
_KEY_DEFAULT_DAYS = 30
_KEY_MAX_DAYS = 365

def _bcrypt_lib():
    import bcrypt
    return bcrypt

def auth_faces_collection():
    return get_db()["auth_faces"]

def auth_qr_collection():
    return get_db()["auth_temp_qr_logins"]

def auth_devices_collection():
    return get_db()["auth_login_devices"]

def auth_keys_collection():
    return get_db()["auth_encrypted_key_logins"]

def technical_settings_collection():
    return get_db()["technical_settings"]

@st.cache_resource
def ensure_auth_indexes():
    try:
        auth_faces_collection().create_index([("active", 1), ("name", 1)])
        auth_qr_collection().create_index([("active", 1), ("expires_at", 1)])
        auth_devices_collection().create_index([("active", 1), ("last_login_at", -1)])
        auth_keys_collection().create_index([("active", 1), ("expires_at", 1)])
        auth_keys_collection().create_index("public_key_hash", unique=True)
        technical_settings_collection().create_index("updated_at")
        managed_secrets_collection().create_index("updated_at")
    except Exception:
        pass

def _new_auth_id(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex}"

def _get_face_match_tolerance() -> float:
    raw = safe_secret("FACE_MATCH_TOLERANCE", "") or os.getenv("FACE_MATCH_TOLERANCE", "")
    try:
        if raw:
            return min(max(float(raw), 0.35), 0.60)
    except Exception:
        pass
    return _FACE_MATCH_TOLERANCE

@st.cache_resource
def _load_cv_libs():
    try:
        import cv2
        import numpy as np
        return cv2, np, ""
    except Exception as exc:
        return None, None, str(exc)

@st.cache_resource
def _load_face_tools():
    try:
        cv2, np, error = _load_cv_libs()
        if error:
            return None, None, None, None, error
        try:
            import face_recognition
            return cv2, np, face_recognition, "face_recognition", ""
        except Exception:
            import dlib
            import face_recognition_models
            detector = dlib.get_frontal_face_detector()
            pose_predictor = dlib.shape_predictor(face_recognition_models.pose_predictor_model_location())
            face_encoder = dlib.face_recognition_model_v1(face_recognition_models.face_recognition_model_location())
            tools = {
                "detector": detector,
                "pose_predictor": pose_predictor,
                "face_encoder": face_encoder,
                "dlib": dlib,
            }
            return cv2, np, tools, "dlib", ""
    except Exception as exc:
        return None, None, None, None, str(exc)

def _cv_dependency_message(error: str) -> str:
    return (
        "Camera and QR processing need OpenCV and numpy installed. "
        f"Dependency error: {error}"
    )

def _vision_dependency_message(error: str) -> str:
    return (
        "Camera login needs OpenCV, numpy, dlib-bin, and face-recognition-models installed. "
        f"Dependency error: {error}"
    )

def _uploaded_file_to_bgr(uploaded_file):
    cv2, np, error = _load_cv_libs()
    if error:
        return None, _cv_dependency_message(error)
    if uploaded_file is None:
        return None, "Capture an image first."
    raw = uploaded_file.getvalue()
    if not raw:
        return None, "The captured image is empty."
    image_array = np.frombuffer(raw, dtype=np.uint8)
    image_bgr = cv2.imdecode(image_array, cv2.IMREAD_COLOR)
    if image_bgr is None:
        return None, "Could not read the captured image."
    return image_bgr, ""

def extract_face_encoding(uploaded_file):
    cv2, np, face_tools, backend, error = _load_face_tools()
    if error:
        return None, _vision_dependency_message(error)
    image_bgr, error = _uploaded_file_to_bgr(uploaded_file)
    if error:
        return None, error
    image_rgb = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2RGB)
    if backend == "face_recognition":
        locations = face_tools.face_locations(image_rgb, model="hog")
        if not locations:
            return None, "No face was found. Try brighter light and keep your face centered."
        if len(locations) > 1:
            return None, "Multiple faces were found. Capture only one person."
        encodings = face_tools.face_encodings(image_rgb, known_face_locations=locations)
        if not encodings:
            return None, "Could not convert this face into a Face ID vector."
        return [float(v) for v in encodings[0]], ""

    dlib = face_tools["dlib"]
    rectangles = list(face_tools["detector"](image_rgb, 1))
    if not rectangles:
        return None, "No face was found. Try brighter light and keep your face centered."
    if len(rectangles) > 1:
        return None, "Multiple faces were found. Capture only one person."
    shape = face_tools["pose_predictor"](image_rgb, rectangles[0])
    encoding = face_tools["face_encoder"].compute_face_descriptor(image_rgb, shape)
    return [float(v) for v in np.array(encoding, dtype="float64")], ""

def _find_matching_face(encoding: list[float]):
    _, np, error = _load_cv_libs()
    if error:
        return None, None, _cv_dependency_message(error)

    docs = list(auth_faces_collection().find(
        {"active": {"$ne": False}},
        {"encoding": 1, "name": 1, "role": 1, "created_at": 1, "last_login_at": 1},
    ))
    if not docs:
        return None, None, "No Face IDs are enrolled yet."

    known = []
    candidates = []
    for doc in docs:
        stored = doc.get("encoding")
        if isinstance(stored, list) and len(stored) == len(encoding):
            known.append(stored)
            candidates.append(doc)

    if not candidates:
        return None, None, "No valid Face ID vectors are saved."

    distances = np.linalg.norm(
        np.array(known, dtype="float64") - np.array(encoding, dtype="float64"),
        axis=1,
    )
    best_idx = int(np.argmin(distances))
    best_distance = float(distances[best_idx])
    if best_distance <= _get_face_match_tolerance():
        return candidates[best_idx], best_distance, ""
    return None, best_distance, "Face ID not recognized."

def save_face_profile(name: str, role: str, encoding: list[float], created_via: str, qr_token_id: str | None = None) -> str:
    face_id = _new_auth_id("face")
    auth_faces_collection().insert_one({
        "_id": face_id,
        "name": name.strip()[:120],
        "role": role if role in {"admin", "member"} else "member",
        "encoding": encoding,
        "vector_version": _FACE_VECTOR_VERSION,
        "active": True,
        "created_via": created_via,
        "qr_token_id": qr_token_id,
        "created_by": st.session_state.get("username", "System"),
        "created_at": datetime.now(),
        "last_login_at": None,
        "login_count": 0,
    })
    return face_id

def _hash_temp_qr_secret(secret_value: str) -> str:
    return hashlib.sha256(f"{_TEMP_QR_TYPE}:{secret_value}".encode()).hexdigest()

def _hash_pin(pin: str) -> str:
    bcrypt = _bcrypt_lib()
    return bcrypt.hashpw(pin.encode(), bcrypt.gensalt()).decode()

def _check_pin(pin: str, pin_hash: str) -> bool:
    try:
        bcrypt = _bcrypt_lib()
        return bcrypt.checkpw(pin.encode(), pin_hash.encode())
    except Exception:
        return False

def make_temp_login_qr_png(payload: str) -> bytes:
    try:
        import qrcode
    except ImportError as exc:
        raise RuntimeError("Install qrcode[pil] to generate temporary login QR codes.") from exc
    qr = qrcode.QRCode(box_size=8, border=2)
    qr.add_data(payload)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    out = BytesIO()
    img.save(out, format="PNG")
    return out.getvalue()

def create_temp_qr_invite(member_name: str, pin: str, expires_hours: int) -> dict:
    token_id = _new_auth_id("qr")
    token_secret = secrets.token_urlsafe(32)
    expires_at = datetime.now() + timedelta(hours=int(expires_hours))
    payload = {
        "type": _TEMP_QR_TYPE,
        "token_id": token_id,
        "secret": token_secret,
    }
    auth_qr_collection().insert_one({
        "_id": token_id,
        "member_name": member_name.strip()[:120],
        "secret_hash": _hash_temp_qr_secret(token_secret),
        "pin_hash": _hash_pin(pin),
        "active": True,
        "expires_at": expires_at,
        "created_at": datetime.now(),
        "created_by": st.session_state.get("username", "Admin"),
        "pin_failures": 0,
        "used_at": None,
        "face_id": None,
    })
    payload_json = json.dumps(payload, separators=(",", ":"))
    return {
        "token_id": token_id,
        "member_name": member_name.strip()[:120],
        "expires_at": expires_at,
        "payload_json": payload_json,
        "qr_png": make_temp_login_qr_png(payload_json),
    }

def _decode_qr_payload(uploaded_file):
    cv2, _, error = _load_cv_libs()
    if error:
        return None, _cv_dependency_message(error)
    image_bgr, error = _uploaded_file_to_bgr(uploaded_file)
    if error:
        return None, error
    detector = cv2.QRCodeDetector()
    decoded, _, _ = detector.detectAndDecode(image_bgr)
    if not decoded:
        try:
            ok, decoded_info, _, _ = detector.detectAndDecodeMulti(image_bgr)
            if ok:
                decoded = next((item for item in decoded_info if item), "")
        except Exception:
            decoded = ""
    if not decoded:
        return None, "No QR code was detected. Hold the QR clearly in the camera frame."
    try:
        payload = json.loads(decoded)
    except Exception:
        return None, "This QR code is not a boutique login QR."
    if payload.get("type") != _TEMP_QR_TYPE:
        return None, "This QR code is not a boutique login QR."
    if not payload.get("token_id") or not payload.get("secret"):
        return None, "This QR code is missing login data."
    return payload, ""

def _lookup_temp_qr(payload: dict):
    doc = auth_qr_collection().find_one({"_id": payload.get("token_id")})
    if not doc:
        return None, "This QR login has not been found."
    expected = _hash_temp_qr_secret(str(payload.get("secret", "")))
    if not hmac.compare_digest(str(doc.get("secret_hash", "")), expected):
        return None, "This QR login is not valid."
    if not doc.get("active", True) or doc.get("used_at"):
        return None, "This QR login has already been used or revoked."
    expires_at = doc.get("expires_at")
    if isinstance(expires_at, datetime) and expires_at < datetime.now():
        return None, "This QR login has expired."
    return doc, ""

def _verify_temp_qr_pin(payload: dict, pin: str):
    doc, error = _lookup_temp_qr(payload)
    if error:
        return None, error
    if _check_pin(pin, str(doc.get("pin_hash", ""))):
        return doc, ""

    failures = int(doc.get("pin_failures", 0)) + 1
    update = {"$set": {"pin_failures": failures}}
    if failures >= _MAX_ATTEMPTS:
        update["$set"]["active"] = False
        update["$set"]["revoked_reason"] = "too_many_pin_failures"
    auth_qr_collection().update_one({"_id": doc["_id"]}, update)
    if failures >= _MAX_ATTEMPTS:
        return None, "Too many wrong PIN attempts. This QR login is now revoked."
    return None, f"Invalid PIN. {max(_MAX_ATTEMPTS - failures, 0)} attempt(s) remaining."

def _mark_temp_qr_used(token_id: str, face_id: str | None = None):
    auth_qr_collection().update_one(
        {"_id": token_id},
        {"$set": {
            "active": False,
            "used_at": datetime.now(),
            "face_id": face_id,
        }},
    )

def _load_key_crypto():
    try:
        from cryptography.hazmat.primitives import serialization
        from cryptography.hazmat.primitives.asymmetric import rsa
    except ImportError as exc:
        raise RuntimeError("Install cryptography to use encrypted key login.") from exc
    return serialization, rsa

def _clean_key_label(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()[:120]

def _private_key_public_hash(private_key) -> str:
    serialization, _ = _load_key_crypto()
    public_der = private_key.public_key().public_bytes(
        encoding=serialization.Encoding.DER,
        format=serialization.PublicFormat.SubjectPublicKeyInfo,
    )
    return hashlib.sha256(public_der).hexdigest()

def _format_key_fingerprint(public_key_hash: str) -> str:
    short_hash = str(public_key_hash or "")[:32]
    return ":".join(short_hash[i:i + 2] for i in range(0, len(short_hash), 2))

def _load_private_key_from_pem(pem_text: str, passphrase: str = ""):
    serialization, _ = _load_key_crypto()
    raw = str(pem_text or "").strip().encode("utf-8")
    if not raw:
        raise ValueError("Upload or paste a private key first.")
    password = str(passphrase or "").encode("utf-8") or None
    try:
        return serialization.load_pem_private_key(raw, password=password)
    except TypeError as exc:
        if password:
            raise ValueError("Private key passphrase is incorrect.") from exc
        raise ValueError("This private key is encrypted. Enter its passphrase.") from exc
    except ValueError as exc:
        raise ValueError("Invalid private key. Upload or paste a valid PEM private key.") from exc

def _encrypted_key_status(doc: dict) -> str:
    expires_at = doc.get("expires_at")
    uses = int(doc.get("uses", 0) or 0)
    max_uses = int(doc.get("max_uses", 0) or 0)
    if not doc.get("active", True) or doc.get("revoked_at"):
        return "Revoked"
    if isinstance(expires_at, datetime) and expires_at <= datetime.now():
        return "Expired"
    if max_uses and uses >= max_uses:
        return "Used"
    return "Active"

def create_encrypted_key_login(for_user: str, role: str, valid_days: int, max_uses: int = 0, note: str = "") -> tuple[bytes, dict]:
    serialization, rsa = _load_key_crypto()
    for_user = _clean_key_label(for_user)
    if not for_user:
        raise ValueError("Enter who this encrypted key is for.")
    role = role if role in {"admin", "member"} else "member"
    valid_days = max(1, min(int(valid_days or _KEY_DEFAULT_DAYS), _KEY_MAX_DAYS))
    max_uses = max(0, min(int(max_uses or 0), 1000))
    private_key = rsa.generate_private_key(public_exponent=65537, key_size=4096)
    private_pem = private_key.private_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PrivateFormat.TraditionalOpenSSL,
        encryption_algorithm=serialization.NoEncryption(),
    )
    public_key_hash = _private_key_public_hash(private_key)
    now = datetime.now()
    doc = {
        "_id": _new_auth_id("key"),
        "public_key_hash": public_key_hash,
        "fingerprint": _format_key_fingerprint(public_key_hash),
        "for_user": for_user,
        "role": role,
        "active": True,
        "created_by": st.session_state.get("username", "Admin"),
        "created_at": now,
        "expires_at": now + timedelta(days=valid_days),
        "uses": 0,
        "max_uses": max_uses,
        "note": str(note or "").strip()[:200],
    }
    auth_keys_collection().insert_one(doc.copy())
    return private_pem, doc

def consume_encrypted_key_login(pem_text: str, passphrase: str = "") -> tuple[bool, str, dict | None]:
    try:
        private_key = _load_private_key_from_pem(pem_text, passphrase=passphrase)
    except Exception as exc:
        return False, str(exc), None
    public_key_hash = _private_key_public_hash(private_key)
    doc = auth_keys_collection().find_one({"public_key_hash": public_key_hash})
    if not doc:
        return False, "Encrypted key was not found. Generate a valid key first.", None
    status = _encrypted_key_status(doc)
    if status == "Revoked":
        return False, "This encrypted key has been revoked.", None
    if status == "Expired":
        return False, "This encrypted key has expired. Generate a fresh key.", None
    if status == "Used":
        return False, "This encrypted key has reached its allowed login limit.", None
    uses = int(doc.get("uses", 0) or 0)
    result = auth_keys_collection().update_one(
        {
            "_id": doc["_id"],
            "public_key_hash": public_key_hash,
            "uses": uses,
            "active": True,
            "expires_at": {"$gt": datetime.now()},
        },
        {
            "$inc": {"uses": 1},
            "$set": {"last_used_at": datetime.now(), "last_used_by": doc.get("for_user", "Key User")},
        },
    )
    if result.modified_count != 1:
        return False, "This encrypted key is no longer valid. Generate a fresh key.", None
    doc["uses"] = uses + 1
    doc["last_used_at"] = datetime.now()
    return True, "Encrypted key accepted.", doc

def _request_device_details() -> dict:
    headers = {}
    try:
        headers = dict(st.context.headers)
    except Exception:
        headers = {}
    user_agent = headers.get("user-agent", "") or headers.get("User-Agent", "")
    ip_addr = headers.get("x-forwarded-for", "") or headers.get("X-Forwarded-For", "")
    lower = user_agent.lower()
    if "iphone" in lower:
        device = "iPhone"
    elif "ipad" in lower:
        device = "iPad"
    elif "macintosh" in lower or "mac os" in lower:
        device = "Mac"
    elif "windows" in lower:
        device = "Windows"
    elif "android" in lower:
        device = "Android"
    else:
        device = "Browser"
    if "chrome" in lower:
        browser = "Chrome"
    elif "safari" in lower:
        browser = "Safari"
    elif "firefox" in lower:
        browser = "Firefox"
    elif "edge" in lower:
        browser = "Edge"
    else:
        browser = "Web"
    signature = hashlib.sha256(f"{user_agent}|{ip_addr}".encode()).hexdigest()[:18]
    return {
        "label": f"{browser} on {device}",
        "signature": signature,
        "user_agent": user_agent[:500],
        "ip": ip_addr[:120],
    }

def establish_login(username: str, role: str = "member", method: str = "face", face_id: str | None = None, qr_token_id: str | None = None):
    device_id = st.session_state.get("auth_device_id")
    if device_id:
        existing_device = auth_devices_collection().find_one({"_id": device_id}, {"active": 1})
        if existing_device and existing_device.get("active") is False:
            device_id = None
    if not device_id:
        device_id = _new_auth_id("device")
        st.session_state.auth_device_id = device_id

    details = _request_device_details()
    now = datetime.now()
    auth_devices_collection().update_one(
        {"_id": device_id},
        {
            "$set": {
                "user_name": username.strip()[:120] or "Member",
                "role": role,
                "method": method,
                "face_id": face_id,
                "qr_token_id": qr_token_id,
                "label": details["label"],
                "signature": details["signature"],
                "user_agent": details["user_agent"],
                "ip": details["ip"],
                "last_login_at": now,
                "active": True,
            },
            "$setOnInsert": {"created_at": now},
        },
        upsert=True,
    )
    if face_id:
        auth_faces_collection().update_one(
            {"_id": face_id},
            {"$set": {"last_login_at": now}, "$inc": {"login_count": 1}},
        )
    st.session_state.logged_in = True
    st.session_state.username = username.strip()[:120] or "Member"
    st.session_state.user_role = role
    st.session_state.auth_method = method
    st.session_state.face_id = face_id

def _current_auth_device_revoked() -> bool:
    device_id = st.session_state.get("auth_device_id")
    if not device_id:
        return False
    doc = auth_devices_collection().find_one({"_id": device_id}, {"active": 1})
    return bool(doc and doc.get("active") is False)

def _is_admin() -> bool:
    return st.session_state.get("user_role", "admin") == "admin"

def _clear_auth_state():
    for key in (
        "logged_in",
        "username",
        "user_role",
        "auth_method",
        "face_id",
        "pending_qr_payload",
        "pending_qr_member",
        "pending_qr_token_id",
        "pending_qr_pin_ok",
        "generated_temp_qr",
        "auth_device_id",
        "technical_unlocked",
        "technical_unlocked_at",
        "technical_auth_method",
        "generated_technical_password",
    ):
        st.session_state.pop(key, None)
    st.session_state.logged_in = False

def _get_stored_hash() -> bytes | None:
    """
    Return the bcrypt hash of the admin password.

    Priority:
      1. st.secrets["PASSWORD_HASH"]  — a bcrypt hash (preferred for production)
      2. st.secrets["PASSWORD"]       — plain text, hashed on the fly (migration path)
      3. env var PASSWORD_HASH        — bcrypt hash
      4. env var PASSWORD             — plain text, hashed on the fly

    If none of these are set the function returns None and login is blocked.
    """
    # 1. Pre-hashed secret
    h = safe_secret("PASSWORD_HASH", "") or os.getenv("PASSWORD_HASH", "")
    if h:
        return h.encode() if isinstance(h, str) else h

    # 2. Plain-text secret — hash on the fly (one-time cost per cold start)
    p = safe_secret("PASSWORD", "") or os.getenv("PASSWORD", "")
    if p:
        bcrypt = _bcrypt_lib()
        return bcrypt.hashpw(p.encode(), bcrypt.gensalt())

    # Nothing configured — fail closed
    return None


def _get_username() -> str | None:
    u = safe_secret("USERNAME", "") or os.getenv("USERNAME", "")
    return u.strip() or None


def _check_lockout() -> tuple[bool, int]:
    """Returns (is_locked, seconds_remaining)."""
    attempts  = st.session_state.get("login_attempts", 0)
    lock_time = st.session_state.get("login_lock_until", 0)
    now       = time.time()
    if lock_time and now < lock_time:
        return True, int(lock_time - now)
    if lock_time and now >= lock_time:
        # Reset after lockout expires
        st.session_state.login_attempts   = 0
        st.session_state.login_lock_until = 0
    return False, 0


def _record_failure():
    attempts = st.session_state.get("login_attempts", 0) + 1
    st.session_state.login_attempts = attempts
    if attempts >= _MAX_ATTEMPTS:
        st.session_state.login_lock_until = time.time() + _LOCKOUT_SECS
    else:
        # Progressive back-off delay (no await needed — this is server-side Streamlit)
        delay = _BACKOFF_BASE * (2 ** (attempts - 1))
        time.sleep(min(delay, 30))


def _verify_credentials(username: str, password: str) -> bool:
    stored_user = _get_username()
    stored_hash = _get_stored_hash()

    if stored_user is None or stored_hash is None:
        return False

    # Constant-time username compare
    user_ok = hmac.compare_digest(username.encode(), stored_user.encode())
    # bcrypt compare (constant-time internally)
    try:
        bcrypt = _bcrypt_lib()
        pass_ok = bcrypt.checkpw(password.encode(), stored_hash)
    except Exception:
        pass_ok = False

    return user_ok and pass_ok

def _verify_admin_password_only(password: str) -> bool:
    stored_hash = _get_stored_hash()
    if stored_hash is None:
        return False
    try:
        bcrypt = _bcrypt_lib()
        return bcrypt.checkpw(str(password or "").encode(), stored_hash)
    except Exception:
        return False

def get_technical_access_doc() -> dict:
    try:
        return technical_settings_collection().find_one({"_id": "technical_access"}) or {}
    except Exception:
        return {}

def technical_password_status() -> str:
    doc = get_technical_access_doc()
    if doc.get("password_hash") and doc.get("active", True):
        return "Generated technical password active"
    return "Fallback admin password only"

def verify_technical_password(password: str) -> tuple[bool, str]:
    password = str(password or "")
    doc = get_technical_access_doc()
    if doc.get("password_hash") and doc.get("active", True):
        if _check_pin(password, str(doc.get("password_hash", ""))):
            return True, "technical_password"
    if _verify_admin_password_only(password):
        return True, "admin_fallback"
    return False, ""

def generate_technical_password() -> str:
    password = secrets.token_urlsafe(24)
    technical_settings_collection().update_one(
        {"_id": "technical_access"},
        {"$set": {
            "password_hash": _hash_pin(password),
            "active": True,
            "updated_at": datetime.now(),
            "updated_by": st.session_state.get("username", "Admin"),
            "last_generated_at": datetime.now(),
        }},
        upsert=True,
    )
    return password

def clear_technical_unlock():
    for key in ("technical_unlocked", "technical_unlocked_at", "technical_auth_method"):
        st.session_state.pop(key, None)


# =====================================================
# ADMIN LOGIN
# =====================================================

def _clear_pending_qr_login():
    for key in ("pending_qr_payload", "pending_qr_member", "pending_qr_token_id", "pending_qr_pin_ok"):
        st.session_state.pop(key, None)

def _render_qr_member_login_panel():
    if st.session_state.get("pending_qr_payload"):
        member_name = st.session_state.get("pending_qr_member", "Member")
        st.success(f"QR accepted for {member_name}.")
        with st.form("qr_pin_form"):
            pin = st.text_input("PIN", type="password", key="qr_login_pin")
            submitted = st.form_submit_button("Verify PIN", width="stretch")
        if submitted:
            doc, error = _verify_temp_qr_pin(st.session_state.pending_qr_payload, pin)
            if error:
                st.error(error)
            else:
                member_name = doc.get("member_name", member_name)
                _mark_temp_qr_used(doc["_id"])
                _clear_pending_qr_login()
                establish_login(member_name, role="member", method="qr_pin", qr_token_id=doc["_id"])
                st.rerun()
        if st.button("Scan different QR", key="scan_different_qr"):
            _clear_pending_qr_login()
            st.rerun()
        return

    st.caption("Paste the QR payload from the generated temporary QR invite.")
    payload_text = st.text_area(
        "QR Payload",
        placeholder='{"type":"boutique_temp_login","token_id":"...","secret":"..."}',
        height=130,
        key="qr_login_payload_text",
    )
    if st.button("Use QR Payload", key="use_qr_payload", width="stretch"):
        try:
            payload = json.loads(payload_text.strip())
        except Exception:
            st.error("Paste a valid QR payload first.")
            return
        if payload.get("type") != _TEMP_QR_TYPE:
            st.error("This QR payload is not a boutique login QR.")
            return
        doc, error = _lookup_temp_qr(payload)
        if error:
            st.error(error)
            return
        st.session_state.pending_qr_payload = payload
        st.session_state.pending_qr_member = doc.get("member_name", "Member")
        st.session_state.pending_qr_token_id = doc["_id"]
        st.rerun()

def _render_encrypted_key_login_panel():
    st.caption("Upload or paste a registered PEM private key.")
    key_upload = st.file_uploader("Upload Private Key", type=["pem", "key", "txt"], key="login_key_upload")
    pasted_key = st.text_area(
        "Paste Private Key",
        placeholder="-----BEGIN RSA PRIVATE KEY-----",
        height=180,
        key="login_key_paste",
    )
    key_passphrase = st.text_input("Private Key Passphrase", type="password", placeholder="Optional", key="login_key_passphrase")
    if st.button("Sign in with encrypted key", key="login_key_submit", width="stretch"):
        key_text = ""
        if key_upload is not None:
            try:
                key_text = key_upload.getvalue().decode("utf-8")
            except UnicodeDecodeError:
                st.error("Private key file must be a text PEM file.")
                return
        if not key_text:
            key_text = pasted_key
        ok, message, doc = consume_encrypted_key_login(key_text, passphrase=key_passphrase)
        if ok and doc:
            establish_login(
                doc.get("for_user", "Key User"),
                role=doc.get("role", "member"),
                method="encrypted_key",
            )
            st.success(message)
            st.rerun()
        else:
            st.error(message)

def _render_password_fallback_panel():
    if _get_username() is None:
        st.info("Password fallback username is not configured.")
        return

    attempts_left = _MAX_ATTEMPTS - st.session_state.get("login_attempts", 0)
    if attempts_left < _MAX_ATTEMPTS:
        st.warning(f"{attempts_left} attempt(s) remaining before lockout.")

    with st.form("admin_login_form"):
        u = st.text_input("Username", placeholder="username", key="admin_u")
        p = st.text_input("Password", type="password", placeholder="••••••••", key="admin_p")
        submitted = st.form_submit_button("Sign in with password", width="stretch")

    if submitted:
        if _verify_credentials(u, p):
            st.session_state.login_attempts = 0
            st.session_state.login_lock_until = 0
            establish_login(u, role="admin", method="password")
            st.rerun()
        else:
            _record_failure()
            locked2, secs_left2 = _check_lockout()
            if locked2:
                st.error(f"Account locked for {secs_left2 // 60}m {secs_left2 % 60}s.")
            else:
                st.error("Invalid credentials.")

def render_admin_login_strip():
    st.markdown("<div class='admin-strip'>", unsafe_allow_html=True)
    st.markdown("<div class='admin-strip-label'>◆ Secure Access</div>", unsafe_allow_html=True)

    with st.expander("Sign in to Admin Dashboard", expanded=False):
        locked, secs_left = _check_lockout()
        if locked:
            st.error(f"Too many failed attempts. Try again in {secs_left // 60}m {secs_left % 60}s.")
            st.markdown("</div>", unsafe_allow_html=True)
            return

        key_tab, password_tab, qr_tab = st.tabs(["Encrypted Key", "Password fallback", "QR member"])
        with key_tab:
            _render_encrypted_key_login_panel()
        with password_tab:
            _render_password_fallback_panel()
        with qr_tab:
            _render_qr_member_login_panel()

    st.markdown("</div>", unsafe_allow_html=True)

# =====================================================
# ADMIN SIDEBAR
# =====================================================

def sidebar():
    with st.sidebar:
        st.markdown("""
        <div class='sb-brand'>
            <div class='sb-logo' style='font-family:"DM Serif Display",serif'>Shree Krishna</div>
            <div class='sb-mark'>Boutique Manager</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<div class='sb-sep'></div>", unsafe_allow_html=True)

        df = fetch_all()
        m  = metrics(df)

        c1, c2 = st.columns(2)
        c1.metric("Pending", f"₹{m['pending']:,.0f}")
        c2.metric("Profit",  f"₹{m['profit']:,.0f}")
        c1.metric("Sales",   m["sales"])
        c2.metric("Clients", m["customers"])

        st.markdown("<div class='sb-sep'></div>", unsafe_allow_html=True)

        nav_options = nav_options_for_current_user()
        forced_page = st.session_state.pop("settings_force_page", None)
        if forced_page in nav_options:
            st.session_state["sidebar_nav"] = forced_page
        default_page = app_pref("default_page", nav_options[0])
        if default_page not in nav_options:
            default_page = nav_options[0]
        if st.session_state.get("sidebar_nav") not in nav_options:
            st.session_state["sidebar_nav"] = default_page

        nav = st.radio("Navigation", nav_options, label_visibility="collapsed", key="sidebar_nav")

        st.markdown("<div class='sb-sep'></div>", unsafe_allow_html=True)
        role_label = st.session_state.get("user_role", "admin").title()
        st.markdown(
            f"<div class='sb-user'>◆ {st.session_state.get('username','Admin').title()} · {role_label}</div>",
            unsafe_allow_html=True,
        )
    return nav

# =====================================================
# ADMIN PAGES
# =====================================================

def _format_auth_dt(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if value:
        return str(value)[:16]
    return "Never"

def page_security_devices():
    if not _is_admin():
        st.error("Security settings are available only to admins.")
        return

    ensure_auth_indexes()
    page_header("Security & Devices", "Encrypted Keys, QR Invites & Login Sessions")
    key_tab, qr_tab, device_tab = st.tabs(["Encrypted Keys", "Temporary QR", "Login devices"])

    with key_tab:
        with st.container(border=True):
            st.subheader("Generate encrypted key")
            with st.form("generate_encrypted_key_form"):
                key_name = st.text_input("Key for whom", value=st.session_state.get("username", "Admin"), key="security_key_name")
                key_role = st.selectbox("Access role", ["admin", "member"], format_func=lambda item: item.title(), key="security_key_role")
                key_days = st.number_input("Valid days", min_value=1, max_value=_KEY_MAX_DAYS, value=_KEY_DEFAULT_DAYS, step=1)
                key_max_uses = st.number_input("Allowed logins", min_value=0, max_value=1000, value=0, step=1, help="Use 0 for unlimited logins until expiry or revocation.")
                key_note = st.text_input("Key note", placeholder="Optional", key="security_key_note")
                key_submitted = st.form_submit_button("Generate encrypted key", width="stretch")
            if key_submitted:
                try:
                    private_pem, key_doc = create_encrypted_key_login(key_name, key_role, int(key_days), int(key_max_uses), key_note)
                    safe_name = re.sub(r"[^0-9A-Za-z]+", "_", key_doc["for_user"]).strip("_").lower() or "encrypted_key"
                    st.session_state.generated_encrypted_key = {
                        "private_pem": private_pem.decode("utf-8"),
                        "file_name": f"{safe_name}_encrypted_key.pem",
                        "fingerprint": key_doc["fingerprint"],
                        "expires_at": key_doc["expires_at"],
                    }
                    st.success(f"Encrypted key generated for {key_doc['for_user']}.")
                except Exception as exc:
                    st.error(str(exc))

            generated_key = st.session_state.get("generated_encrypted_key")
            if generated_key:
                st.caption(f"Fingerprint: {generated_key['fingerprint']}")
                st.caption(f"Expires: {_format_auth_dt(generated_key['expires_at'])}")
                st.download_button(
                    "Download private key PEM",
                    data=generated_key["private_pem"].encode("utf-8"),
                    file_name=generated_key["file_name"],
                    mime="application/x-pem-file",
                    width="content",
                )
                with st.expander("Emergency copy private key", expanded=False):
                    st.code(generated_key["private_pem"])

        st.markdown("<div class='rule-sm'></div>", unsafe_allow_html=True)
        keys = list(auth_keys_collection().find({}, {"public_key_hash": 0}).sort("created_at", -1).limit(100))
        if not keys:
            st.info("No encrypted keys created yet.")
        for key_doc in keys:
            key_id = key_doc["_id"]
            status = _encrypted_key_status(key_doc)
            max_uses = int(key_doc.get("max_uses", 0) or 0)
            with st.container(border=True):
                c1, c2, c3, c4 = st.columns([2.2, 1.1, 1.6, 0.9])
                c1.markdown(f"**{html_escape(str(key_doc.get('for_user', 'Key User')))}**")
                c1.caption(str(key_doc.get("fingerprint", "-")))
                c2.write(str(key_doc.get("role", "member")).title())
                c2.caption(status)
                c3.write(f"Expires: {_format_auth_dt(key_doc.get('expires_at'))}")
                c3.caption(f"Logins: {int(key_doc.get('uses', 0) or 0)}/{max_uses if max_uses else 'unlimited'}")
                if status == "Active":
                    if c4.button("Revoke", key=f"revoke_key_{key_id}", width="stretch"):
                        auth_keys_collection().update_one(
                            {"_id": key_id},
                            {"$set": {
                                "active": False,
                                "revoked_at": datetime.now(),
                                "revoked_by": st.session_state.get("username", "Admin"),
                            }},
                        )
                        st.success("Encrypted key revoked.")
                        st.rerun()
                else:
                    if c4.button("Delete", key=f"delete_key_{key_id}", width="stretch"):
                        auth_keys_collection().delete_one({"_id": key_id})
                        st.success("Encrypted key deleted.")
                        st.rerun()

    with qr_tab:
        with st.container(border=True):
            st.subheader("Generate temporary QR")
            with st.form("generate_temp_qr_form"):
                member_name = st.text_input("Member name", key="qr_member_name")
                pin = st.text_input("PIN to share", type="password", key="qr_member_pin")
                pin2 = st.text_input("Confirm PIN", type="password", key="qr_member_pin2")
                expires_hours = st.number_input("Expires in hours", min_value=1, max_value=168, value=24, step=1)
                submitted = st.form_submit_button("Generate QR", width="stretch")
            if submitted:
                if not member_name.strip():
                    st.error("Enter the member name.")
                elif len(pin) < 4:
                    st.error("Use at least 4 characters for the PIN.")
                elif pin != pin2:
                    st.error("PIN values do not match.")
                else:
                    st.session_state.generated_temp_qr = create_temp_qr_invite(member_name, pin, int(expires_hours))
                    st.success("Temporary QR generated.")

            generated = st.session_state.get("generated_temp_qr")
            if generated:
                st.image(generated["qr_png"], width=260)
                st.caption(f"For {generated['member_name']} · expires {_format_auth_dt(generated['expires_at'])}")
                st.download_button(
                    "Download QR",
                    data=generated["qr_png"],
                    file_name=f"{generated['member_name'].replace(' ', '_')}_login_qr.png",
                    mime="image/png",
                    width="content",
                )
                with st.expander("Copy QR payload for login", expanded=False):
                    st.code(generated["payload_json"])

        st.markdown("<div class='rule-sm'></div>", unsafe_allow_html=True)
        invites = list(auth_qr_collection().find(
            {},
            {"secret_hash": 0, "pin_hash": 0},
        ).sort("created_at", -1).limit(50))
        if not invites:
            st.info("No QR invites created yet.")
        for invite in invites:
            invite_id = invite["_id"]
            active = bool(invite.get("active", True)) and not invite.get("used_at")
            if isinstance(invite.get("expires_at"), datetime) and invite["expires_at"] < datetime.now():
                active = False
            with st.container(border=True):
                c1, c2, c3, c4 = st.columns([2.2, 1.1, 1.4, 0.9])
                c1.markdown(f"**{html_escape(str(invite.get('member_name', 'Member')))}**")
                c1.caption(f"Created {_format_auth_dt(invite.get('created_at'))}")
                c2.write("Active" if active else "Closed")
                c3.write(f"Expires: {_format_auth_dt(invite.get('expires_at'))}")
                if invite.get("used_at"):
                    c3.caption(f"Used {_format_auth_dt(invite.get('used_at'))}")
                elif active and c4.button("Revoke", key=f"revoke_qr_{invite_id}", width="stretch"):
                    auth_qr_collection().update_one(
                        {"_id": invite_id},
                        {"$set": {
                            "active": False,
                            "revoked_at": datetime.now(),
                            "revoked_by": st.session_state.get("username", "Admin"),
                        }},
                    )
                    st.success("QR invite revoked.")
                    st.rerun()

    with device_tab:
        devices = list(auth_devices_collection().find({}).sort("last_login_at", -1).limit(100))
        if not devices:
            st.info("No login devices recorded yet.")
        for device in devices:
            device_id = device["_id"]
            current = device_id == st.session_state.get("auth_device_id")
            active = bool(device.get("active", True))
            with st.container(border=True):
                c1, c2, c3, c4 = st.columns([2.2, 1.1, 1.6, 0.9])
                label = device.get("label") or "Browser session"
                c1.markdown(f"**{html_escape(str(label))}**")
                c1.caption("This session" if current else str(device.get("signature", "")))
                c2.write(str(device.get("role", "member")).title())
                c2.caption(str(device.get("method", "login")).replace("_", " ").title())
                c3.write(str(device.get("user_name", "Member")))
                c3.caption(f"Last login {_format_auth_dt(device.get('last_login_at'))}")
                if active:
                    if c4.button("Revoke", key=f"revoke_device_{device_id}", width="stretch"):
                        auth_devices_collection().update_one(
                            {"_id": device_id},
                            {"$set": {
                                "active": False,
                                "revoked_at": datetime.now(),
                                "revoked_by": st.session_state.get("username", "Admin"),
                            }},
                        )
                        if current:
                            _clear_auth_state()
                        st.success("Login device revoked.")
                        st.rerun()
                else:
                    if c4.button("Delete", key=f"delete_device_{device_id}", width="stretch"):
                        auth_devices_collection().delete_one({"_id": device_id})
                        st.success("Login device deleted.")
                        st.rerun()

    security_context = "\n\n".join([
        "Encrypted keys:\n" + df_for_ai(pd.DataFrame(keys).drop(columns=["public_key_hash"], errors="ignore") if "keys" in locals() and keys else pd.DataFrame(), limit=60),
        "Temporary QR invites:\n" + df_for_ai(pd.DataFrame(list(auth_qr_collection().find({}, {"secret_hash": 0, "pin_hash": 0}).sort("created_at", -1).limit(60))), limit=60),
        "Login devices:\n" + df_for_ai(pd.DataFrame(list(auth_devices_collection().find({}, {"user_agent": 0}).sort("last_login_at", -1).limit(60))), limit=60),
    ])
    render_ai_panel(
        "AI Security Review",
        security_context,
        "security_ai",
        "Review login security. Flag risky active keys, stale devices, expired QR invites, and suggest what to revoke or rotate.",
    )

def _format_bytes(value) -> str:
    try:
        size = float(value or 0)
    except Exception:
        return "0 B"
    for unit in ["B", "KB", "MB", "GB", "TB"]:
        if size < 1024 or unit == "TB":
            return f"{size:,.1f} {unit}" if unit != "B" else f"{size:,.0f} B"
        size /= 1024
    return f"{size:,.1f} TB"

def _json_safe(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    return value

def _run_db_command(label: str, command):
    start = time.perf_counter()
    try:
        result = get_db().command(command)
        elapsed_ms = round((time.perf_counter() - start) * 1000, 2)
        return {"label": label, "ok": True, "elapsed_ms": elapsed_ms, "result": _json_safe(result), "error": ""}
    except Exception as exc:
        elapsed_ms = round((time.perf_counter() - start) * 1000, 2)
        return {"label": label, "ok": False, "elapsed_ms": elapsed_ms, "result": {}, "error": str(exc)}

def collect_mongo_collection_stats() -> pd.DataFrame:
    rows = []
    db = get_db()
    try:
        names = sorted(db.list_collection_names(), key=str.casefold)
    except Exception:
        names = []
    for name in names:
        col = db[name]
        row = {
            "Collection": name,
            "Documents": 0,
            "Indexes": 0,
            "Data Size": "—",
            "Storage Size": "—",
            "Avg Object": "—",
            "Capped": "—",
            "Status": "OK",
        }
        try:
            row["Documents"] = int(col.estimated_document_count())
        except Exception as exc:
            row["Status"] = str(exc)[:120]
        try:
            row["Indexes"] = len(col.index_information())
        except Exception:
            pass
        try:
            stats = db.command("collstats", name)
            row["Data Size"] = _format_bytes(stats.get("size"))
            row["Storage Size"] = _format_bytes(stats.get("storageSize"))
            row["Avg Object"] = _format_bytes(stats.get("avgObjSize"))
            row["Capped"] = "Yes" if stats.get("capped") else "No"
        except Exception:
            pass
        rows.append(row)
    return pd.DataFrame(rows)

def render_technical_gate() -> bool:
    if not _is_admin():
        page_header("Technical", "Protected System Console")
        st.error("Technical access is admin-only.")
        return False
    if st.session_state.get("technical_unlocked"):
        return True

    page_header("Technical", "Protected System Console")
    lock_until = st.session_state.get("technical_lock_until", 0)
    if lock_until and time.time() < lock_until:
        left = int(lock_until - time.time())
        st.error(f"Technical access locked for {left // 60}m {left % 60}s.")
        return False
    if lock_until and time.time() >= lock_until:
        st.session_state.technical_attempts = 0
        st.session_state.technical_lock_until = 0

    with st.container(border=True):
        st.subheader("Enter Technical Password")
        st.caption(f"Status: {technical_password_status()}. The admin password remains the fallback password.")
        with st.form("technical_access_form"):
            password = st.text_input("Technical Password", type="password", key="technical_password_input")
            submitted = st.form_submit_button("Unlock Technical", width="stretch")
        if submitted:
            ok, method = verify_technical_password(password)
            if ok:
                st.session_state.technical_unlocked = True
                st.session_state.technical_unlocked_at = datetime.now()
                st.session_state.technical_auth_method = method
                st.session_state.technical_attempts = 0
                st.session_state.technical_lock_until = 0
                st.rerun()
            else:
                attempts = int(st.session_state.get("technical_attempts", 0)) + 1
                st.session_state.technical_attempts = attempts
                if attempts >= _MAX_ATTEMPTS:
                    st.session_state.technical_lock_until = time.time() + _LOCKOUT_SECS
                    st.error("Too many wrong attempts. Technical access is temporarily locked.")
                else:
                    time.sleep(min(_BACKOFF_BASE * (2 ** (attempts - 1)), 30))
                    st.error(f"Invalid technical password. {max(_MAX_ATTEMPTS - attempts, 0)} attempt(s) remaining.")
    return False

def render_mongodb_monitoring_tab():
    uri = safe_secret("MONGO_URI", os.getenv("MONGO_URI", ""))
    commands = [
        _run_db_command("Ping", {"ping": 1}),
        _run_db_command("Database Stats", {"dbStats": 1}),
        _run_db_command("Build Info", {"buildInfo": 1}),
        _run_db_command("Connection Status", {"connectionStatus": 1}),
        _run_db_command("Server Status", {"serverStatus": 1}),
        _run_db_command("Host Info", {"hostInfo": 1}),
    ]
    command_map = {item["label"]: item for item in commands}
    collection_stats = collect_mongo_collection_stats()
    db_stats = command_map.get("Database Stats", {}).get("result", {})
    server_status = command_map.get("Server Status", {}).get("result", {})

    total_docs = int(collection_stats["Documents"].sum()) if not collection_stats.empty and "Documents" in collection_stats.columns else 0
    total_indexes = int(collection_stats["Indexes"].sum()) if not collection_stats.empty and "Indexes" in collection_stats.columns else 0

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Ping", f"{command_map['Ping']['elapsed_ms']} ms" if command_map["Ping"]["ok"] else "Failed")
    c2.metric("Collections", int(db_stats.get("collections", len(collection_stats) if not collection_stats.empty else 0) or 0))
    c3.metric("Documents", f"{total_docs:,}")
    c4.metric("Indexes", f"{total_indexes:,}")
    c5.metric("Data Size", _format_bytes(db_stats.get("dataSize")))
    c6.metric("Storage", _format_bytes(db_stats.get("storageSize")))

    st.markdown("<div class='settings-status'>"
                f"<b>Database:</b> {html_escape(get_db().name)}<br>"
                f"<b>Mongo URI:</b> {html_escape(mask_mongo_uri(uri))}<br>"
                f"<b>Server uptime:</b> {html_escape(str(server_status.get('uptime', 'Unavailable')))} seconds"
                "</div>", unsafe_allow_html=True)

    sec("Collections")
    if collection_stats.empty:
        st.info("No collections found or collection listing is not permitted.")
    else:
        st.dataframe(collection_stats, width="stretch", hide_index=True)
        selected_collection = st.selectbox(
            "Collection Details",
            collection_stats["Collection"].tolist(),
            key="technical_collection_detail",
        )
        detail = {"collection": selected_collection}
        try:
            detail["indexes"] = _json_safe(get_db()[selected_collection].index_information())
        except Exception as exc:
            detail["index_error"] = str(exc)
        try:
            detail["stats"] = _json_safe(get_db().command("collstats", selected_collection))
        except Exception as exc:
            detail["stats_error"] = str(exc)
        try:
            sample = get_db()[selected_collection].find_one({}, {"_id": 0}) or {}
            detail["sample_fields"] = sorted([str(key) for key in sample.keys()])
        except Exception as exc:
            detail["sample_error"] = str(exc)
        with st.expander("Selected collection technical details", expanded=False):
            st.json(detail)

    sec("Command Results")
    command_rows = pd.DataFrame([
        {
            "Command": item["label"],
            "Status": "OK" if item["ok"] else "Failed / Not permitted",
            "Time ms": item["elapsed_ms"],
            "Error": item["error"][:220],
        }
        for item in commands
    ])
    st.dataframe(command_rows, width="stretch", hide_index=True)
    for item in commands:
        with st.expander(f"{item['label']} raw result", expanded=False):
            if item["ok"]:
                st.json(item["result"])
            else:
                st.error(item["error"])

def render_gemini_api_details_tab():
    gemini_key = get_gemini_key()
    provider = get_ai_provider()
    model = get_gemini_model()
    rows = [
        secret_presence("GEMINI_API_KEY"),
        secret_presence("GOOGLE_API_KEY"),
        secret_presence("GEMINI_MODEL"),
        secret_presence("AI_PROVIDER"),
        secret_presence("OPENAI_API_KEY"),
        secret_presence("OPENAI_MODEL"),
    ]
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("AI Provider", provider.title() if provider else "Auto")
    c2.metric("Gemini Key", "Found" if gemini_key else "Missing")
    c3.metric("Gemini Model", model)
    c4.metric("AI Ready", "Yes" if llm_is_configured() else "No")

    try:
        import importlib.metadata as importlib_metadata
        genai_version = importlib_metadata.version("google-genai")
    except Exception:
        genai_version = "Not installed"
    st.markdown(
        "<div class='settings-status'>"
        f"<b>google-genai:</b> {html_escape(genai_version)}<br>"
        f"<b>Active Gemini key:</b> {html_escape(mask_secret_value(gemini_key))}<br>"
        f"<b>OpenAI key:</b> {html_escape(mask_secret_value(get_openai_key()))}"
        "</div>",
        unsafe_allow_html=True,
    )

    st.dataframe(pd.DataFrame(rows), width="stretch", hide_index=True)

    if st.button("Test Gemini API", key="technical_test_gemini", width="stretch"):
        if not gemini_key:
            st.error("Gemini API key is missing.")
        else:
            try:
                from google import genai
                from google.genai import types
                with st.spinner("Testing Gemini..."):
                    client = genai.Client(api_key=gemini_key)
                    response = client.models.generate_content(
                        model=model,
                        contents="Reply with exactly: Gemini OK",
                        config=types.GenerateContentConfig(temperature=0),
                    )
                st.success("Gemini API responded.")
                st.code((getattr(response, "text", "") or str(response))[:1200])
            except Exception as exc:
                st.error(str(exc))

def render_managed_secrets_tab():
    st.warning("Managed AI secrets are encrypted in MongoDB. Streamlit Cloud deployment secrets still control startup-only values such as MONGO_URI and password unless you update them in Streamlit secrets too.")
    rows = [secret_presence(key) for key in MANAGED_SECRET_KEYS]
    show = pd.DataFrame(rows).rename(columns={
        "key": "Secret",
        "source": "Source",
        "managed": "Managed Override",
        "masked": "Masked Value",
        "updated_at": "Managed Updated",
        "updated_by": "Managed By",
    })
    st.dataframe(show, width="stretch", hide_index=True)

    sec("Edit Secret")
    selected_key = st.selectbox("Secret Key", MANAGED_SECRET_KEYS, key="technical_secret_key")
    current = secret_presence(selected_key)
    st.caption(f"Current source: {current['source']} · Value: {current['masked']}")
    if selected_key in BOOTSTRAP_SECRET_KEYS:
        st.info("This is a startup/login secret. You can store a managed copy here, but the live app still needs the matching Streamlit Cloud secret for startup/auth behavior.")

    new_value = st.text_input("New Secret Value", type="password", key="technical_secret_value")
    save_col, delete_col = st.columns(2)
    with save_col:
        if st.button("Save Encrypted Secret", key="technical_save_secret", width="stretch"):
            try:
                save_managed_secret(selected_key, new_value)
                st.success(f"{selected_key} saved as an encrypted managed secret.")
                st.rerun()
            except Exception as exc:
                st.error(str(exc))
    with delete_col:
        if st.button("Delete Managed Override", key="technical_delete_secret", width="stretch"):
            delete_managed_secret(selected_key)
            st.success(f"Managed override removed for {selected_key}.")
            st.rerun()

def render_technical_settings_tab():
    doc = get_technical_access_doc()
    c1, c2, c3 = st.columns(3)
    c1.metric("Technical Password", "Active" if doc.get("password_hash") and doc.get("active", True) else "Fallback Only")
    c2.metric("Unlocked By", str(st.session_state.get("technical_auth_method", "—")).replace("_", " ").title())
    c3.metric("Generated", _format_auth_dt(doc.get("last_generated_at")) if doc else "—")

    if st.button("Generate New Technical Password", key="generate_technical_password", width="stretch"):
        st.session_state.generated_technical_password = generate_technical_password()
        st.success("Technical password generated. Copy it now; it will be shown only in this session.")

    generated = st.session_state.get("generated_technical_password")
    if generated:
        st.warning("Copy this password now. It is not shown again after sign-out.")
        st.code(generated)

    a, b = st.columns(2)
    with a:
        if st.button("Disable Generated Password", key="disable_technical_password", width="stretch"):
            technical_settings_collection().update_one(
                {"_id": "technical_access"},
                {"$set": {"active": False, "updated_at": datetime.now(), "updated_by": st.session_state.get("username", "Admin")}},
                upsert=True,
            )
            st.session_state.pop("generated_technical_password", None)
            st.success("Generated technical password disabled. Admin password fallback remains active.")
            st.rerun()
    with b:
        if st.button("Lock Technical Now", key="lock_technical_now", width="stretch"):
            clear_technical_unlock()
            st.rerun()

def page_technical():
    if not render_technical_gate():
        return

    page_header("Technical", "System Monitoring & Secrets")
    mongo_tab, gemini_tab, secrets_tab, settings_tab = st.tabs([
        "MongoDB Monitoring",
        "Gemini API Details",
        "Secrets",
        "Settings",
    ])
    with mongo_tab:
        render_mongodb_monitoring_tab()
    with gemini_tab:
        render_gemini_api_details_tab()
    with secrets_tab:
        render_managed_secrets_tab()
    with settings_tab:
        render_technical_settings_tab()

def page_dashboard():
    page_header("Dashboard", "Business Overview")
    df = fetch_all()
    m  = metrics(df)

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Sales",      m["sales"])
    c2.metric("Revenue",    f"₹{m['revenue']:,.0f}")
    c3.metric("Net Profit", f"₹{m['profit']:,.0f}")
    c4.metric("Pending",    f"₹{m['pending']:,.0f}")
    c5.metric("Avg Margin", f"{m['margin']:.1f}%")
    c6.metric("Customers",  m["customers"])

    rule()

    if df.empty:
        st.markdown("<div class='empty'><div class='empty-glyph'>◆</div><div>No sales yet.</div></div>", unsafe_allow_html=True)
        return

    df["month"] = df["sale_date"].dt.to_period("M").astype(str)

    cl, cr = st.columns([3, 2])
    with cl:
        monthly = df.groupby("month").agg(revenue=("selling_price","sum"), profit=("profit","sum"), sales=("id","count")).reset_index()
        fig = go.Figure()
        fig.add_trace(go.Bar(x=monthly["month"], y=monthly["revenue"], name="Revenue", marker_color="rgba(46,111,216,0.4)", marker_line_color="#2E6FD8", marker_line_width=1))
        fig.add_trace(go.Scatter(x=monthly["month"], y=monthly["profit"], name="Profit", mode="lines+markers", line=dict(color="#7ADFA0", width=2), marker=dict(size=5, color="#7ADFA0")))
        styled_fig(fig, 300).update_layout(title="Monthly Revenue & Profit", barmode="overlay", legend=dict(orientation="h", y=1.18, x=0))
        st.plotly_chart(fig, width="stretch")

    with cr:
        paid    = (df["payment_received"] == 1).sum()
        pending = (df["payment_received"] == 0).sum()
        fig2 = go.Figure(go.Pie(labels=["Collected","Pending"], values=[paid, pending], hole=0.72, marker=dict(colors=["#2E6FD8","#0F1A2E"]), textfont=dict(size=11), hovertemplate="%{label}: %{value}<extra></extra>"))
        fig2.add_annotation(text=f"<b>{paid+pending}</b>", x=0.5, y=0.52, showarrow=False, font=dict(color="#E8EEF8", family="Playfair Display", size=28))
        fig2.add_annotation(text="sales", x=0.5, y=0.38, showarrow=False, font=dict(color="#3D5478", family="Jost", size=11))
        styled_fig(fig2, 300).update_layout(title="Payment Status", showlegend=True, legend=dict(orientation="h", y=-0.05, x=0.25))
        st.plotly_chart(fig2, width="stretch")

    cl2, cr2 = st.columns(2)
    with cl2:
        cat_rev = df.groupby("product_category")["selling_price"].sum().reset_index()
        fig3 = px.pie(cat_rev, values="selling_price", names="product_category", title="Revenue by Category", hole=0.55, color_discrete_sequence=["#2E6FD8","#4D8AE8","#7ADFA0","#8BACD8","#E08090","#1A3D80","#3D9A6C","#4A9AC8","#9B9070","#A8C4F0"])
        styled_fig(fig3, 270); st.plotly_chart(fig3, width="stretch")

    with cr2:
        daily = df.set_index("sale_date")["selling_price"].resample("D").sum().reset_index()
        daily.columns = ["date","revenue"]
        daily["rolling"] = daily["revenue"].rolling(7, min_periods=1).mean()
        fig4 = go.Figure()
        fig4.add_trace(go.Bar(x=daily["date"], y=daily["revenue"], name="Daily", marker_color="rgba(46,111,216,0.25)", marker_line_width=0))
        fig4.add_trace(go.Scatter(x=daily["date"], y=daily["rolling"], name="7-day avg", line=dict(color="#2E6FD8", width=1.8)))
        styled_fig(fig4, 270).update_layout(title="Daily Revenue", legend=dict(orientation="h", y=1.18, x=0))
        st.plotly_chart(fig4, width="stretch")

    sec("Recent Transactions")
    recent = df.sort_values("sale_date", ascending=False).head(10).copy()
    recent["sale_date"] = recent["sale_date"].dt.strftime("%d %b %Y")
    recent["Status"]    = recent["payment_received"].map({1:"Paid", 0:"Pending"})
    recent["Delayed"]   = recent["delay_status"].map({0:"—", 1:"Yes"})
    show = recent[["id","customer_name","sale_date","product_category","selling_price","profit","pending_amount","Status","Delayed"]].copy()
    show.columns = ["ID","Customer","Date","Category","Amount ₹","Profit ₹","Pending ₹","Status","Delayed"]
    st.dataframe(show, width="stretch", hide_index=True)

    rule()
    da, db, _ = st.columns([1, 1, 2])
    with da:
        st.download_button("Export CSV", data=df.assign(sale_date=df["sale_date"].astype(str)).to_csv(index=False), file_name=f"boutique_{date.today()}.csv", mime="text/csv", width="stretch")
    with db:
        st.download_button("Export Excel", data=to_excel(df), file_name=f"boutique_{date.today()}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", width="stretch")

    render_ai_panel(
        "AI Dashboard Summary",
        build_ai_business_context(),
        "dashboard_ai",
        "Summarize the current business status, important risks, and the top 3 actions I should take next.",
    )


def page_review():
    page_header("Accounts", "All Transactions")
    df = fetch_all()
    if df.empty:
        st.markdown("<div class='empty'><div class='empty-glyph'>◆</div><div>No transactions yet.</div></div>", unsafe_allow_html=True)
        return

    with st.expander("Filter & Sort", expanded=True):
        c1, c2, c3, c4 = st.columns(4)
        with c1: srch  = st.text_input("Customer / Phone")
        with c2: catf  = st.selectbox("Category",   ["All"] + CATEGORIES)
        with c3: payf  = st.selectbox("Payment",    ["All","Paid","Pending"])
        with c4: dlayf = st.selectbox("Delay Flag", ["All","On Time","Delayed"])
        c5, c6, c7 = st.columns(3)
        with c5: sortby = st.selectbox("Sort By", ["Date ↓","Date ↑","Amount ↓","Pending ↓","Profit ↓"])
        default_review_days = int(app_pref("default_review_days", 90))
        with c6: d_from = st.date_input("From", value=date.today() - timedelta(days=default_review_days))
        with c7: d_to   = st.date_input("To",   value=date.today())

    fdf = df.copy()
    if srch:
        phone_text = fdf["customer_phone"].astype(str)
        mask = (
            fdf["customer_name"].str.contains(srch, case=False, na=False, regex=False)
            | phone_text.str.contains(srch, case=False, na=False, regex=False)
        )
        phone_digits = normalize_phone(srch)
        if phone_digits:
            mask = mask | phone_text.map(normalize_phone).str.contains(phone_digits, na=False, regex=False)
        fdf = fdf[mask]
    if catf  != "All": fdf = fdf[fdf["product_category"] == catf]
    if payf  == "Paid":     fdf = fdf[fdf["payment_received"] == 1]
    elif payf == "Pending": fdf = fdf[fdf["payment_received"] == 0]
    if dlayf == "On Time":  fdf = fdf[fdf["delay_status"] == 0]
    elif dlayf == "Delayed": fdf = fdf[fdf["delay_status"] == 1]
    fdf = fdf[(fdf["sale_date"] >= pd.Timestamp(d_from)) & (fdf["sale_date"] <= pd.Timestamp(d_to))]
    sm = {"Date ↓":("sale_date",False),"Date ↑":("sale_date",True),"Amount ↓":("selling_price",False),"Pending ↓":("pending_amount",False),"Profit ↓":("profit",False)}
    sc, sa = sm[sortby]
    fdf = fdf.sort_values(sc, ascending=sa)

    rule_sm()
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Transactions", len(fdf))
    m2.metric("Revenue",      f"₹{fdf['selling_price'].sum():,.0f}")
    m3.metric("Profit",       f"₹{fdf['profit'].sum():,.0f}")
    m4.metric("Pending",      f"₹{fdf['pending_amount'].sum():,.0f}")
    m5.metric("Avg Margin",   f"{fdf['margin'].mean():.1f}%" if not fdf.empty else "—")
    rule_sm()

    save_message = st.session_state.pop("accounts_save_message", None)
    if save_message:
        st.success(save_message)

    editor = fdf[["id","customer_name","customer_phone","sale_date","vendor","product_category","buying_price","selling_price","profit","amount_paid","pending_amount","payment_method","last_payment_method","last_payment_date","last_payment_received_by","payment_received"]].copy()
    editor["ID"] = editor["id"].astype(int)
    editor["Customer"] = editor["customer_name"].fillna("").astype(str)
    editor["Phone"] = editor["customer_phone"].fillna("").astype(str)
    editor["Date"] = editor["sale_date"].map(lambda value: pd.to_datetime(value, errors="coerce").date() if pd.notna(pd.to_datetime(value, errors="coerce")) else date.today())
    editor["Vendor"] = editor["vendor"].fillna("").astype(str)
    editor["Category"] = editor["product_category"].map(lambda value: value if value in CATEGORIES else CATEGORIES[0])
    editor["Buy ₹"] = editor["buying_price"].map(money_value)
    editor["Sell ₹"] = editor["selling_price"].map(money_value)
    editor["Profit ₹"] = editor["profit"].map(money_value)
    editor["Paid ₹"] = editor["amount_paid"].map(money_value)
    editor["Pending ₹"] = editor["pending_amount"].map(money_value)
    editor["Sale Method"] = editor["payment_method"].map(lambda value: value if value in PAYMENT_METHODS else PAYMENT_METHODS[0])
    editor["Paid Method"] = editor["last_payment_method"].map(lambda value: value if value in PAYMENT_COLLECTION_METHODS else "")
    editor["Paid Date"] = editor["last_payment_date"].map(lambda value: pd.to_datetime(value, errors="coerce").date() if pd.notna(pd.to_datetime(value, errors="coerce")) else None)
    editor["Received By"] = editor["last_payment_received_by"].fillna("").astype(str)
    editor["Status"] = editor["payment_received"].map({0:"Pending", 1:"Paid"}).fillna("Pending")
    editor = editor[["ID","Customer","Phone","Date","Vendor","Category","Buy ₹","Sell ₹","Profit ₹","Paid ₹","Pending ₹","Sale Method","Paid Method","Paid Date","Received By","Status"]]

    edited_accounts = st.data_editor(
        editor,
        width="stretch",
        hide_index=True,
        num_rows="fixed",
        disabled=["ID", "Profit ₹", "Pending ₹", "Status"],
        column_config={
            "ID": st.column_config.NumberColumn("ID", disabled=True),
            "Date": st.column_config.DateColumn("Date", format="YYYY-MM-DD", required=True),
            "Category": st.column_config.SelectboxColumn("Category", options=CATEGORIES, required=True),
            "Buy ₹": st.column_config.NumberColumn("Buy ₹", min_value=0.0, step=100.0, format="₹ %.2f"),
            "Sell ₹": st.column_config.NumberColumn("Sell ₹", min_value=0.0, step=100.0, format="₹ %.2f"),
            "Profit ₹": st.column_config.NumberColumn("Profit ₹", disabled=True, format="₹ %.2f"),
            "Paid ₹": st.column_config.NumberColumn("Paid ₹", min_value=0.0, step=100.0, format="₹ %.2f"),
            "Pending ₹": st.column_config.NumberColumn("Pending ₹", disabled=True, format="₹ %.2f"),
            "Sale Method": st.column_config.SelectboxColumn("Sale Method", options=PAYMENT_METHODS, required=True),
            "Paid Method": st.column_config.SelectboxColumn("Paid Method", options=[""] + PAYMENT_COLLECTION_METHODS),
            "Paid Date": st.column_config.DateColumn("Paid Date", format="YYYY-MM-DD"),
            "Status": st.column_config.TextColumn("Status", disabled=True),
        },
        key="accounts_inline_editor",
    )
    save_edit_col, _ = st.columns([1, 3])
    with save_edit_col:
        if st.button("Save Edited Rows", width="stretch"):
            changed, errors = save_account_editor_changes(fdf, edited_accounts)
            for err in errors:
                st.error(err)
            if changed:
                st.session_state.accounts_save_message = f"Saved {changed} edited row(s)."
                st.rerun()
            elif not errors:
                st.info("No edited rows to save.")

    dc, de, _ = st.columns([1,1,2])
    with dc:
        st.download_button("Export CSV", data=fdf.assign(sale_date=fdf["sale_date"].astype(str)).to_csv(index=False), file_name=f"accounts_{date.today()}.csv", mime="text/csv", width="stretch")
    with de:
        st.download_button("Export Excel", data=to_excel(fdf), file_name=f"accounts_{date.today()}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", width="stretch")

    render_ai_panel(
        "AI Review Current Filter",
        "Filtered account rows:\n" + df_for_ai(fdf, ["id","customer_name","customer_phone","sale_date","vendor","product_category","selling_price","amount_paid","pending_amount","payment_method"], 60),
        "review_ai",
        "Analyze these filtered rows. Call out pending collections, unusual entries, and practical follow-up actions.",
    )

    sec("Mark Payments")
    pend = fdf[fdf["pending_amount"] > 0].sort_values("pending_amount", ascending=False)
    if pend.empty:
        st.success("All payments received for current filter.")
    else:
        st.markdown(f"<span class='badge badge-gold'>{len(pend)} pending — ₹{pend['pending_amount'].sum():,.0f}</span>", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        for _, row in pend.iterrows():
            row_id = int(row["id"])
            sale_day = row["sale_date"].strftime("%d %b %Y") if pd.notna(row["sale_date"]) else "—"
            pending_value = money_value(row.get("pending_amount"))
            vendor_name = str(row.get("vendor", "") or "").strip() or "—"
            st.markdown(
                f"""
                <div class='pay-card'>
                    <div class='pay-grid'>
                        <div>
                            <div class='pay-name'>{html_escape(str(row['customer_name']))}</div>
                            <div class='pay-meta'>{html_escape(str(row.get('product_category', '—')))} · Sale #{row_id}</div>
                        </div>
                        <div>
                            <div class='pay-label'>Vendor</div>
                            <div class='pay-amount'>{html_escape(vendor_name)}</div>
                        </div>
                        <div>
                            <div class='pay-label'>Pending</div>
                            <div class='pay-amount'>₹{pending_value:,.2f}</div>
                        </div>
                        <div>
                            <div class='pay-label'>Sale Date</div>
                            <div class='pay-amount'>{sale_day}</div>
                        </div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            action_cols = st.columns([4, 1])
            with action_cols[1]:
                if st.button("Mark Paid", key=f"pay_open_{row_id}", width="stretch"):
                    st.session_state.payment_editor_id = row_id
                    st.rerun()

            if st.session_state.get("payment_editor_id") == row_id:
                st.markdown("<div class='pay-form-note'>Enter the received amount. Use the full pending amount for complete payment, or a smaller amount for partial payment.</div>", unsafe_allow_html=True)
                with st.form(f"payment_form_{row_id}"):
                    pc1, pc2, pc3, pc4 = st.columns([1, 1, 1, 1.2])
                    with pc1:
                        _, payment_amount, payment_ok = currency_input("Amount Received (₹)", f"payment_amount_{row_id}", pending_value)
                    with pc2:
                        payment_date = st.date_input("Paid Date", value=date.today(), key=f"payment_date_{row_id}")
                    with pc3:
                        payment_method = st.selectbox("Paid By", PAYMENT_COLLECTION_METHODS, key=f"payment_method_{row_id}")
                    with pc4:
                        received_by = st.text_input("Received By", value=default_receiver_name(), key=f"payment_received_by_{row_id}")
                    save_col, cancel_col = st.columns(2)
                    with save_col:
                        save_payment = st.form_submit_button("Save Payment", width="stretch")
                    with cancel_col:
                        cancel_payment = st.form_submit_button("Cancel", width="stretch")

                    if save_payment:
                        if not payment_ok:
                            st.error("Payment amount must be a valid number.")
                        else:
                            ok, message = record_payment(row, payment_amount, payment_date, payment_method, received_by)
                            if ok:
                                st.session_state.payment_editor_id = None
                                st.success(message)
                                st.rerun()
                            else:
                                st.error(message)
                    if cancel_payment:
                        st.session_state.payment_editor_id = None
                        st.rerun()


def page_update():
    page_header("Update", "Edit or Delete a Record")
    c1, c2 = st.columns([2,1])
    with c1: sname = st.text_input("Search by Customer Name / Phone")
    with c2: sid   = st.number_input("Or by Sale ID", min_value=0, step=1)

    if not sname and sid == 0:
        st.info("Enter a customer name, phone, or sale ID to search.")
        return

    if sname:
        search_text = sname.strip()
        phone_digits = normalize_phone(search_text)
        terms = [
            {"customer_name": {"$regex": re.escape(search_text), "$options":"i"}},
            {"customer_phone": {"$regex": re.escape(search_text), "$options":"i"}},
        ]
        if phone_digits and phone_digits != search_text:
            terms.append({"customer_phone": {"$regex": re.escape(phone_digits), "$options":"i"}})
        q = {"$or": terms}
    else:
        q = {"id": int(sid)}
    docs = list(get_col().find(q, {"_id":0}))
    if not docs:
        st.warning("No matching transaction found.")
        return

    df = pd.DataFrame(docs)
    for c in ["buying_price","selling_price","amount_paid","pending_amount"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    preview_cols = [c for c in ["id","customer_name","sale_date","product_category","selling_price","pending_amount","payment_received"] if c in df.columns]
    preview = df[preview_cols].copy()
    if "payment_received" in preview.columns:
        preview["payment_received"] = preview["payment_received"].map({0:"Pending",1:"Paid"})
    st.dataframe(preview, width="stretch", hide_index=True)

    sel = st.selectbox("Select ID to Edit", df["id"].tolist(), format_func=lambda x: f"#{x} — {df[df['id']==x]['customer_name'].values[0]}")
    row = df[df["id"] == sel].iloc[0]
    rule_sm()
    render_ai_update_assistant(row, int(sel))
    rule_sm()

    with st.form("update_form"):
        sec("Customer & Product")
        c1, c2, c3 = st.columns(3)
        with c1:
            nn = st.text_input("Customer Name", value=str(row.get("customer_name","")))
            np = st.text_input("Phone",          value=str(row.get("customer_phone","")))
        with c2:
            ci  = CATEGORIES.index(row["product_category"]) if row.get("product_category") in CATEGORIES else 0
            nc  = st.selectbox("Category", CATEGORIES, index=ci)
            nv  = vendor_picker("Vendor", f"update_vendor_{sel}", str(row.get("vendor","")))
        with c3:
            try:    existing_date = pd.to_datetime(row.get("sale_date")).date()
            except: existing_date = date.today()
            new_date = st.date_input("Sale Date", value=existing_date)
            nqty = st.number_input("Quantity", min_value=1, step=1, value=int(row.get("quantity",1)))

        ndesc = st.text_area("Description", value=str(row.get("product_description","")), height=60)
        sec("Pricing & Payment")
        pr1, pr2, pr3, pr4 = st.columns(4)
        with pr1: _, nb, nb_ok   = currency_input("Buying Price (₹)", f"update_buying_price_{sel}", float(row["buying_price"]))
        with pr2: _, ns, ns_ok   = currency_input("Selling Price (₹)", f"update_selling_price_{sel}", float(row["selling_price"]))
        with pr3: _, npa, npa_ok = currency_input("Amount Paid (₹)", f"update_amount_paid_{sel}", float(row["amount_paid"]))
        with pr4:
            pi  = PAYMENT_METHODS.index(row["payment_method"]) if row.get("payment_method") in PAYMENT_METHODS else 0
            npm = st.selectbox("Payment Method", PAYMENT_METHODS, index=pi)

        nd     = st.checkbox("Mark as Delayed", value=bool(row.get("delay_status",0)))
        nnotes = st.text_area("Notes", value=str(row.get("notes","")), height=60)

        npend   = max(round(ns - npa, 2), 0.0)
        nprofit = round(ns - nb, 2)
        m1, m2, m3 = st.columns(3)
        m1.metric("Updated Pending", f"₹{npend:,.2f}")
        m2.metric("Updated Profit",  f"₹{nprofit:,.2f}")
        m3.metric("Updated Margin",  f"{(nprofit/ns*100 if ns>0 else 0):.1f}%")

        bu, bd = st.columns(2)
        with bu: upd = st.form_submit_button("Save Changes",       width="stretch")
        with bd: dlt = st.form_submit_button("Delete Transaction",  width="stretch")

        if upd:
            errs = []
            if not nb_ok:
                errs.append("Buying price must be a valid number.")
            if not ns_ok:
                errs.append("Selling price must be a valid number.")
            if not npa_ok:
                errs.append("Amount paid must be a valid number.")
            if ns_ok and npa_ok and npa > ns:
                errs.append("Amount paid cannot exceed selling price.")
            if errs:
                for err in errs:
                    st.error(err)
            else:
                get_col().update_one({"id": sel}, {"$set": {
                    "customer_name": nn.strip(), "customer_phone": normalize_phone(np),
                    "sale_date": str(new_date), "product_category": nc,
                    "vendor": nv.strip(), "product_description": ndesc.strip(),
                    "quantity": nqty, "buying_price": round(nb,2),
                    "selling_price": round(ns,2), "amount_paid": round(npa,2),
                    "pending_amount": npend, "delay_status": int(nd),
                    "payment_method": npm, "notes": nnotes.strip(),
                    "payment_received": 1 if npend==0 else 0,
                    "updated_at": str(datetime.now()),
                }})
                invalidate_cache(); st.success("Transaction updated."); st.rerun()
        if dlt:
            get_col().delete_one({"id": sel})
            invalidate_cache(); st.success("Transaction deleted."); st.rerun()


def page_customers():
    page_header("Customers", "All Clients")
    df = fetch_all()
    if df.empty:
        st.markdown("<div class='empty'><div class='empty-glyph'>◆</div><div>No customers yet.</div></div>", unsafe_allow_html=True)
        return

    summ = (df.sort_values("sale_date", ascending=False).groupby("customer_name").agg(
        phone=("customer_phone", first_nonempty), transactions=("id","count"),
        spent=("selling_price","sum"), pending=("pending_amount","sum"),
        last_visit=("sale_date","max"), profit=("profit","sum"),
    ).reset_index())
    summ["last_visit"] = pd.to_datetime(summ["last_visit"]).dt.strftime("%d %b %Y")
    summ = summ.sort_values("spent", ascending=False).reset_index(drop=True)
    summ["tier"] = pd.cut(summ["spent"], bins=[0,5000,20000,50000,float("inf")], labels=["Bronze","Silver","Gold","Platinum"])

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Customers", len(summ))
    m2.metric("Avg Spend",       f"₹{summ['spent'].mean():,.0f}")
    m3.metric("With Pending",    len(summ[summ["pending"] > 0]))
    m4.metric("Total Revenue",   f"₹{summ['spent'].sum():,.0f}")

    rule_sm()
    c1, c2 = st.columns([2,1])
    with c1: srch   = st.text_input("Search Customer")
    with c2: tier_f = st.selectbox("Tier", ["All","Bronze","Silver","Gold","Platinum"])

    view = summ.copy()
    if srch:
        phone_digits = normalize_phone(srch)
        mask = (
            view["customer_name"].str.contains(srch, case=False, na=False, regex=False)
            | view["phone"].astype(str).str.contains(srch, case=False, na=False, regex=False)
        )
        if phone_digits:
            mask = mask | view["phone"].astype(str).map(normalize_phone).str.contains(phone_digits, na=False, regex=False)
        view = view[mask]
    if tier_f != "All": view = view[view["tier"] == tier_f]

    disp = view.rename(columns={"customer_name":"Customer","phone":"Phone","transactions":"Visits","spent":"Total Spent ₹","pending":"Pending ₹","last_visit":"Last Visit","profit":"Profit ₹","tier":"Tier"})
    st.dataframe(disp.style.format({"Total Spent ₹":"₹{:,.0f}","Pending ₹":"₹{:,.0f}","Profit ₹":"₹{:,.0f}"}), width="stretch", hide_index=True)

    dc, de = st.columns(2)
    with dc:
        st.download_button("Export CSV", data=disp.to_csv(index=False), file_name=f"customers_{date.today()}.csv", mime="text/csv", width="stretch")
    with de:
        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as w: disp.to_excel(w, index=False)
        out.seek(0)
        st.download_button("Export Excel", data=out, file_name=f"customers_{date.today()}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", width="stretch")

    sec("Purchase History")
    chosen = st.selectbox("Select Customer", summ["customer_name"].tolist())
    if chosen:
        hist = df[df["customer_name"] == chosen].sort_values("sale_date", ascending=False).copy()
        hist["status"]   = hist["payment_received"].map({0:"Pending",1:"Paid"})
        hist["sale_date"] = hist["sale_date"].dt.strftime("%d %b %Y")
        h1, h2, h3, h4 = st.columns(4)
        h1.metric("Visits",      len(hist))
        h2.metric("Total Spent", f"₹{hist['selling_price'].sum():,.0f}")
        h3.metric("Pending",     f"₹{hist['pending_amount'].sum():,.0f}")
        h4.metric("Profit",      f"₹{hist['profit'].sum():,.0f}")
        cols = [c for c in ["sale_date","product_category","product_description","selling_price","amount_paid","pending_amount","payment_method","status"] if c in hist.columns]
        show = hist[cols].copy()
        show.columns = ["Date","Category","Description","Price ₹","Paid ₹","Pending ₹","Method","Status"][:len(cols)]
        st.dataframe(show, width="stretch", hide_index=True)
        if len(hist) > 1:
            hs = df[df["customer_name"]==chosen].sort_values("sale_date").copy()
            hs["cumulative"] = hs["selling_price"].cumsum()
            fig = px.line(hs, x="sale_date", y="cumulative", title=f"Cumulative Spend — {chosen}", markers=True)
            fig.update_traces(line_color="#2E6FD8", marker_color="#4D8AE8", marker_size=5)
            styled_fig(fig, 230); st.plotly_chart(fig, width="stretch")
        render_ai_panel(
            "AI Customer Summary",
            f"Customer: {chosen}\nPurchase history:\n{df_for_ai(hist, limit=60)}",
            "customer_ai",
            "Summarize this customer's buying pattern, pending amount risk, and what I should do next.",
        )


def page_vendors():
    page_header("Vendors", "Supplier Directory")
    df = fetch_all()
    vendors = get_existing_vendors()

    try:
        inv_docs = list(get_db()["inventory"].find({}, {"_id": 0}))
    except Exception as exc:
        inv_docs = []
        st.warning(f"Inventory vendors could not be loaded: {exc}")

    inv_df = pd.DataFrame(inv_docs)
    sales_rows = pd.DataFrame(columns=["vendor"])
    inv_items = pd.DataFrame(columns=["vendor"])

    sales_summary = pd.DataFrame(columns=["vendor", "sales", "revenue", "profit", "pending", "last_sale"])
    if df is not None and not df.empty and "vendor" in df.columns:
        sales_rows = df.copy()
        sales_rows["vendor"] = sales_rows["vendor"].fillna("").astype(str).str.strip()
        sales_rows = sales_rows[sales_rows["vendor"] != ""].copy()
        if not sales_rows.empty:
            for col in ["selling_price", "profit", "pending_amount"]:
                if col not in sales_rows.columns:
                    sales_rows[col] = 0.0
                sales_rows[col] = pd.to_numeric(sales_rows[col], errors="coerce").fillna(0.0)
            sales_rows["_sale_date_sort"] = pd.to_datetime(sales_rows["sale_date"], errors="coerce") if "sale_date" in sales_rows.columns else pd.NaT
            sales_summary = (
                sales_rows.groupby("vendor").agg(
                    sales=("vendor", "count"),
                    revenue=("selling_price", "sum"),
                    profit=("profit", "sum"),
                    pending=("pending_amount", "sum"),
                    last_sale=("_sale_date_sort", "max"),
                ).reset_index()
            )

    inventory_summary = pd.DataFrame(columns=["vendor", "inventory_items", "stock_qty", "stock_value", "low_stock_items"])
    if not inv_df.empty and "vendor" in inv_df.columns:
        inv_items = inv_df.copy()
        inv_items["vendor"] = inv_items["vendor"].fillna("").astype(str).str.strip()
        inv_items = inv_items[inv_items["vendor"] != ""].copy()
        if not inv_items.empty:
            for col in ["quantity", "min_stock", "cost_price"]:
                if col not in inv_items.columns:
                    inv_items[col] = 0.0
                inv_items[col] = pd.to_numeric(inv_items[col], errors="coerce").fillna(0.0)
            inv_items["stock_value"] = inv_items["quantity"] * inv_items["cost_price"]
            inv_items["low_stock_item"] = inv_items["quantity"] <= inv_items["min_stock"]
            item_col = "name" if "name" in inv_items.columns else "vendor"
            inventory_summary = (
                inv_items.groupby("vendor").agg(
                    inventory_items=(item_col, "count"),
                    stock_qty=("quantity", "sum"),
                    stock_value=("stock_value", "sum"),
                    low_stock_items=("low_stock_item", "sum"),
                ).reset_index()
            )

    all_vendors = sorted(set(vendors) | set(sales_summary["vendor"].astype(str)) | set(inventory_summary["vendor"].astype(str)), key=str.casefold)
    if not all_vendors:
        st.markdown("<div class='empty'><div class='empty-glyph'>◆</div><div>No vendors yet.</div></div>", unsafe_allow_html=True)
        return

    summary = pd.DataFrame({"vendor": all_vendors})
    summary = summary.merge(sales_summary, on="vendor", how="left").merge(inventory_summary, on="vendor", how="left")
    for col in ["sales", "revenue", "profit", "pending", "inventory_items", "stock_qty", "stock_value", "low_stock_items"]:
        summary[col] = pd.to_numeric(summary[col], errors="coerce").fillna(0.0)
    summary["last_sale_label"] = pd.to_datetime(summary["last_sale"], errors="coerce").dt.strftime("%d %b %Y").fillna("—")
    summary["vendor_type"] = summary.apply(
        lambda r: "Sales + Inventory" if r["sales"] > 0 and r["inventory_items"] > 0 else ("Sales" if r["sales"] > 0 else ("Inventory" if r["inventory_items"] > 0 else "Saved")),
        axis=1,
    )
    summary = summary.sort_values(["revenue", "stock_value", "vendor"], ascending=[False, False, True]).reset_index(drop=True)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Vendors", len(summary))
    m2.metric("With Sales", int((summary["sales"] > 0).sum()))
    m3.metric("Total Revenue", f"₹{summary['revenue'].sum():,.0f}")
    m4.metric("Stock Value", f"₹{summary['stock_value'].sum():,.0f}")

    rule_sm()
    f1, f2, f3 = st.columns([2, 1, 1])
    with f1:
        search = st.text_input("Search Vendor", key="vendor_list_search")
    with f2:
        show_filter = st.selectbox("Show", ["All Vendors", "With Sales", "With Pending", "Inventory Vendors", "Low Stock"], key="vendor_list_filter")
    with f3:
        sort_by = st.selectbox("Sort By", ["Revenue", "Pending", "Sales", "Inventory Items", "Vendor Name"], key="vendor_list_sort")

    view = summary.copy()
    if search:
        view = view[view["vendor"].str.contains(search, case=False, na=False, regex=False)]
    if show_filter == "With Sales":
        view = view[view["sales"] > 0]
    elif show_filter == "With Pending":
        view = view[view["pending"] > 0]
    elif show_filter == "Inventory Vendors":
        view = view[view["inventory_items"] > 0]
    elif show_filter == "Low Stock":
        view = view[view["low_stock_items"] > 0]

    sort_map = {
        "Revenue": ("revenue", False),
        "Pending": ("pending", False),
        "Sales": ("sales", False),
        "Inventory Items": ("inventory_items", False),
        "Vendor Name": ("vendor", True),
    }
    sort_col, ascending = sort_map[sort_by]
    view = view.sort_values(sort_col, ascending=ascending)

    disp = view[["vendor", "vendor_type", "sales", "revenue", "profit", "pending", "last_sale_label", "inventory_items", "stock_qty", "low_stock_items"]].copy()
    disp = disp.rename(columns={
        "vendor": "Vendor",
        "vendor_type": "Source",
        "sales": "Sales",
        "revenue": "Revenue ₹",
        "profit": "Profit ₹",
        "pending": "Pending ₹",
        "last_sale_label": "Last Sale",
        "inventory_items": "Inventory Items",
        "stock_qty": "Stock Qty",
        "low_stock_items": "Low Stock",
    })
    st.dataframe(
        disp.style.format({
            "Sales": "{:,.0f}",
            "Revenue ₹": "₹{:,.0f}",
            "Profit ₹": "₹{:,.0f}",
            "Pending ₹": "₹{:,.0f}",
            "Inventory Items": "{:,.0f}",
            "Stock Qty": "{:,.0f}",
            "Low Stock": "{:,.0f}",
        }),
        width="stretch",
        hide_index=True,
    )

    dc, de = st.columns(2)
    with dc:
        st.download_button("Export CSV", data=disp.to_csv(index=False), file_name=f"vendors_{date.today()}.csv", mime="text/csv", width="stretch")
    with de:
        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as w:
            disp.to_excel(w, index=False)
        out.seek(0)
        st.download_button("Export Excel", data=out, file_name=f"vendors_{date.today()}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", width="stretch")

    sec("Vendor Details")
    detail_options = view["vendor"].tolist() if not view.empty else summary["vendor"].tolist()

    def vendor_label(name: str) -> str:
        row = summary[summary["vendor"].eq(name)].iloc[0]
        return f"{name} — {int(row['sales'])} sales / ₹{row['revenue']:,.0f}"

    chosen = st.selectbox("Select Vendor", detail_options, format_func=vendor_label, key="vendor_detail_select")
    vendor_sales = sales_rows[sales_rows["vendor"].eq(chosen)].copy() if not sales_rows.empty else pd.DataFrame()
    vendor_inventory = inv_items[inv_items["vendor"].eq(chosen)].copy() if not inv_items.empty else pd.DataFrame()

    d1, d2, d3, d4 = st.columns(4)
    d1.metric("Sales", len(vendor_sales))
    d2.metric("Revenue", f"₹{vendor_sales['selling_price'].sum():,.0f}" if "selling_price" in vendor_sales.columns else "₹0")
    d3.metric("Pending", f"₹{vendor_sales['pending_amount'].sum():,.0f}" if "pending_amount" in vendor_sales.columns else "₹0")
    d4.metric("Stock Items", len(vendor_inventory))

    sales_tab, inventory_tab = st.tabs(["Sales", "Inventory"])
    with sales_tab:
        if vendor_sales.empty:
            st.info("No sales found for this vendor.")
        else:
            vendor_sales = vendor_sales.sort_values("_sale_date_sort", ascending=False)
            show_cols = [c for c in ["sale_date", "customer_name", "product_category", "product_description", "selling_price", "profit", "pending_amount", "payment_method"] if c in vendor_sales.columns]
            show_sales = vendor_sales[show_cols].copy()
            if "sale_date" in show_sales.columns:
                show_sales["sale_date"] = pd.to_datetime(show_sales["sale_date"], errors="coerce").dt.strftime("%d %b %Y")
            show_sales = show_sales.rename(columns={
                "sale_date": "Date",
                "customer_name": "Customer",
                "product_category": "Category",
                "product_description": "Description",
                "selling_price": "Sell ₹",
                "profit": "Profit ₹",
                "pending_amount": "Pending ₹",
                "payment_method": "Method",
            })
            st.dataframe(show_sales, width="stretch", hide_index=True)

    with inventory_tab:
        if vendor_inventory.empty:
            st.info("No inventory items found for this vendor.")
        else:
            show_cols = [c for c in ["name", "sku", "category", "quantity", "min_stock", "cost_price", "sell_price", "notes", "updated_at"] if c in vendor_inventory.columns]
            show_inventory = vendor_inventory[show_cols].copy()
            show_inventory = show_inventory.rename(columns={
                "name": "Item",
                "sku": "SKU",
                "category": "Category",
                "quantity": "Qty",
                "min_stock": "Min Stock",
                "cost_price": "Cost ₹",
                "sell_price": "MRP ₹",
                "notes": "Notes",
                "updated_at": "Updated",
            })
            st.dataframe(show_inventory, width="stretch", hide_index=True)

    vendor_context = "\n\n".join([
        "Vendor summary:\n" + df_for_ai(summary, ["vendor", "vendor_type", "sales", "revenue", "profit", "pending", "last_sale_label", "inventory_items", "stock_qty", "stock_value", "low_stock_items"], 120),
        f"Selected vendor: {chosen}",
        "Selected vendor sales:\n" + df_for_ai(vendor_sales, limit=80),
        "Selected vendor inventory:\n" + df_for_ai(vendor_inventory, limit=80),
    ])
    render_ai_panel(
        "AI Vendor Insights",
        vendor_context,
        "vendor_list_ai",
        "Identify top vendors, pending risks, restock opportunities, and practical follow-up actions.",
    )


def page_generate_bill():
    page_header("Generate Bill", "Customer PDF Statement")
    df = fetch_all()
    if df.empty:
        st.info("No sales available to bill.")
        return

    customers = sorted([c for c in df["customer_name"].dropna().astype(str).unique() if c.strip()], key=str.casefold)
    if not customers:
        st.info("No customers found.")
        return

    def customer_label(name: str) -> str:
        hist = get_customer_bill_data(df, name)
        pending = float(hist["pending_amount"].map(money_value).sum()) if not hist.empty else 0.0
        total = float(hist["selling_price"].map(money_value).sum()) if not hist.empty else 0.0
        return f"{name} — Pending ₹{pending:,.0f} / Total ₹{total:,.0f}"

    c1, c2, c3, c4 = st.columns([2, 1, 1, 1])
    bill_scope_default = app_pref("default_bill_scope", BILL_SCOPE_ALL)
    if bill_scope_default not in BILL_SCOPE_OPTIONS:
        bill_scope_default = BILL_SCOPE_ALL
    with c1:
        selected = st.selectbox("Customer", customers, format_func=customer_label, key="bill_customer")
    with c2:
        bill_scope = st.selectbox("Bill Type", BILL_SCOPE_OPTIONS, index=BILL_SCOPE_OPTIONS.index(bill_scope_default), key="bill_scope")
    with c3:
        bill_limit = st.number_input("Last Transactions", min_value=1, max_value=100, value=5, step=1, key="bill_limit", disabled=bill_scope != BILL_SCOPE_LAST)
    with c4:
        bill_dt = st.date_input("Bill Date", value=date.today(), key="bill_date")

    hist = get_customer_bill_data(df, selected, bill_scope=bill_scope, bill_limit=bill_limit)
    total_bill = float(hist["selling_price"].map(money_value).sum())
    total_paid = float(hist["amount_paid"].map(money_value).sum())
    total_pending = float(hist["pending_amount"].map(money_value).sum())

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Purchases", len(hist))
    m2.metric("Total Bill", f"₹{total_bill:,.0f}")
    m3.metric("Paid", f"₹{total_paid:,.0f}")
    m4.metric("Pending", f"₹{total_pending:,.0f}")

    rule_sm()
    if hist.empty:
        st.info("No pending transactions found for this customer." if bill_scope == BILL_SCOPE_PENDING else "No purchases found for this customer.")
    else:
        st.caption(f"Bill preview: {bill_scope_label(bill_scope, bill_limit)}")
        preview = hist[["sale_date","product_category","product_description","selling_price","amount_paid","pending_amount","last_payment_date"]].copy()
        preview["sale_date"] = preview["sale_date"].dt.strftime("%d %b %Y")
        preview["status"] = hist.apply(bill_status, axis=1)
        preview["last_payment_date"] = hist.apply(bill_paid_date, axis=1)
        preview.columns = ["Date","Category","Description","Bill ₹","Paid ₹","Pending ₹","Paid Date","Status"]
        st.dataframe(preview, width="stretch", hide_index=True)

    dc, _ = st.columns([1, 3])
    with dc:
        render_customer_bill_download(df, selected, key=f"bill_page_download_{re.sub(r'[^0-9A-Za-z]+', '_', selected)}", label="Generate Bill PDF", bill_date=bill_dt, bill_scope=bill_scope, bill_limit=bill_limit)

    rule()
    sec("Bill History")
    h1, h2 = st.columns([2, 1])
    with h1:
        history_search = st.text_input("Search Bill ID / Customer / Phone", key="bill_history_search")
    with h2:
        history_limit = st.number_input("Show Last", min_value=10, max_value=500, value=100, step=10, key="bill_history_limit")

    history = get_bill_history(history_search, limit=int(history_limit))
    if not history:
        st.info("No generated bills found for this search.")
        return

    history_df = pd.DataFrame(history)
    if "bill_scope" not in history_df.columns:
        history_df["bill_scope"] = "All Transactions"
    history_df["bill_scope"] = history_df["bill_scope"].fillna("All Transactions")
    if "bill_scope_label" not in history_df.columns:
        history_df["bill_scope_label"] = history_df.apply(lambda row: bill_scope_label(row.get("bill_scope"), row.get("bill_limit")), axis=1)
    history_df["bill_scope_label"] = history_df["bill_scope_label"].fillna(history_df["bill_scope"])
    history_show = history_df[["bill_id","customer_name","customer_phone","bill_scope_label","bill_date","generated_at","purchase_count","total_bill","total_paid","total_pending","generated_by"]].copy()
    history_show["generated_at"] = pd.to_datetime(history_show["generated_at"], errors="coerce").dt.strftime("%d %b %Y, %I:%M %p").fillna(history_show["generated_at"])
    history_show.columns = ["Bill ID","Customer","Phone","Type","Bill Date","Generated On","Purchases","Total Bill ₹","Paid ₹","Pending ₹","Generated By"]
    st.dataframe(history_show, width="stretch", hide_index=True)

    selected_bill_id = st.selectbox("View Bill Details", history_show["Bill ID"].tolist(), key="bill_history_detail")
    selected_doc = next((doc for doc in history if doc.get("bill_id") == selected_bill_id), None)
    if selected_doc:
        d1, d2, d3, d4 = st.columns(4)
        d1.metric("Bill ID", selected_doc.get("bill_id", "—"))
        d2.metric("Generated", str(selected_doc.get("generated_at", "—"))[:19])
        d3.metric("Total", f"₹{money_value(selected_doc.get('total_bill')):,.0f}")
        d4.metric("Pending", f"₹{money_value(selected_doc.get('total_pending')):,.0f}")

        items = selected_doc.get("items", [])
        if items:
            items_df = pd.DataFrame(items)
            item_cols = ["sale_id","sale_date","category","description","bill_amount","paid_amount","pending_amount","paid_date","status"]
            items_df = items_df[[c for c in item_cols if c in items_df.columns]].copy()
            items_df.columns = ["Sale ID","Sale Date","Category","Description","Bill ₹","Paid ₹","Pending ₹","Paid Date","Status"][:len(items_df.columns)]
            st.dataframe(items_df, width="stretch", hide_index=True)
            render_ai_panel(
                "AI Bill Message",
                f"Bill document:\n{pd.DataFrame([selected_doc]).drop(columns=['items'], errors='ignore').to_csv(index=False)}\nItems:\n{items_df.to_csv(index=False)}",
                "bill_ai",
                "Draft a short polite customer message for this bill. Mention pending amount if any and keep it WhatsApp-friendly.",
            )


def page_analytics():
    page_header("Analytics", "Business Intelligence")
    df = fetch_all()
    if df.empty:
        st.info("No data available.")
        return

    df["month"] = df["sale_date"].dt.to_period("M").astype(str)
    df["dow"]   = df["sale_date"].dt.day_name()

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Revenue",       f"₹{df['selling_price'].sum():,.0f}")
    k2.metric("Profit",        f"₹{df['profit'].sum():,.0f}")
    k3.metric("Avg Order",     f"₹{df['selling_price'].mean():,.0f}")
    k4.metric("Avg Margin",    f"{df['margin'].mean():.1f}%")
    k5.metric("Delayed Count", int((df["delay_status"]==1).sum()))

    rule()
    t1, t2, t3, t4, t5 = st.tabs(["Trends","Customers","Categories","Payments","Top Items"])

    with t1:
        c1, c2 = st.columns(2)
        with c1:
            monthly = df.groupby("month").agg(revenue=("selling_price","sum"), profit=("profit","sum")).reset_index()
            fig = go.Figure()
            fig.add_trace(go.Bar(x=monthly["month"], y=monthly["revenue"], name="Revenue", marker_color="rgba(46,111,216,0.4)", marker_line_color="#2E6FD8", marker_line_width=1))
            fig.add_trace(go.Scatter(x=monthly["month"], y=monthly["profit"], name="Profit", mode="lines+markers", line=dict(color="#7ADFA0", width=2), marker=dict(size=5)))
            styled_fig(fig).update_layout(title="Revenue & Profit by Month", barmode="overlay", legend=dict(orientation="h", y=1.18, x=0))
            st.plotly_chart(fig, width="stretch")
        with c2:
            daily = df.set_index("sale_date")["selling_price"].resample("D").sum().reset_index()
            daily.columns = ["date","revenue"]
            fig2 = px.area(daily, x="date", y="revenue", title="Daily Revenue")
            fig2.update_traces(fillcolor="rgba(46,111,216,0.12)", line_color="#2E6FD8", line_width=1.5)
            styled_fig(fig2); st.plotly_chart(fig2, width="stretch")
        c3, c4 = st.columns(2)
        with c3:
            dow_order = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
            dow = df.groupby("dow").agg(sales=("id","count"), revenue=("selling_price","sum")).reset_index()
            dow["dow"] = pd.Categorical(dow["dow"], categories=dow_order, ordered=True)
            dow = dow.sort_values("dow")
            fig3 = px.bar(dow, x="dow", y="sales", title="Sales by Day of Week", color="revenue", color_continuous_scale=[[0,"#070C18"],[1,"#2E6FD8"]])
            styled_fig(fig3); st.plotly_chart(fig3, width="stretch")
        with c4:
            monthly["MoM Growth %"] = monthly["revenue"].pct_change()*100
            fig4 = px.bar(monthly.dropna(), x="month", y="MoM Growth %", title="Month-over-Month Growth", color="MoM Growth %", color_continuous_scale=[[0,"#C05060"],[0.5,"#0F1A2E"],[1,"#7ADFA0"]])
            styled_fig(fig4); st.plotly_chart(fig4, width="stretch")

    with t2:
        c1, c2 = st.columns(2)
        with c1:
            top_c = df.groupby("customer_name")["selling_price"].sum().nlargest(10).reset_index()
            fig5 = px.bar(top_c, x="selling_price", y="customer_name", orientation="h", title="Top 10 Customers by Revenue", color="selling_price", color_continuous_scale=[[0,"#070C18"],[1,"#2E6FD8"]])
            styled_fig(fig5); fig5.update_layout(yaxis=dict(autorange="reversed")); st.plotly_chart(fig5, width="stretch")
        with c2:
            cp = df.groupby("customer_name")["pending_amount"].sum()
            cp = cp[cp > 0].nlargest(10).reset_index()
            if not cp.empty:
                fig6 = px.bar(cp, x="pending_amount", y="customer_name", orientation="h", title="Top Customers by Pending", color="pending_amount", color_continuous_scale=[[0,"#070C18"],[1,"#C05060"]])
                styled_fig(fig6); fig6.update_layout(yaxis=dict(autorange="reversed")); st.plotly_chart(fig6, width="stretch")
            else:
                st.success("No pending amounts.")
        cust_stats = df.groupby("customer_name").agg(visits=("id","count"), revenue=("selling_price","sum"), avg_order=("selling_price","mean")).reset_index()
        fig_scatter = px.scatter(cust_stats, x="visits", y="revenue", size="avg_order", hover_name="customer_name", title="Customer Value Matrix", color="revenue", color_continuous_scale=[[0,"#070C18"],[1,"#2E6FD8"]])
        styled_fig(fig_scatter, 330); st.plotly_chart(fig_scatter, width="stretch")
        seg = df.groupby("customer_name").agg(spend=("selling_price","sum")).reset_index()
        seg["tier"] = pd.cut(seg["spend"], bins=[0,5000,20000,50000,float("inf")], labels=["Bronze","Silver","Gold","Platinum"])
        sec("Customer Tier Distribution")
        sg = seg.groupby("tier", observed=True).agg(customers=("customer_name","count"), total=("spend","sum")).reset_index()
        sg.columns = ["Tier","Customers","Total Spend ₹"]
        st.dataframe(sg, width="stretch", hide_index=True)

    with t3:
        c1, c2 = st.columns(2)
        with c1:
            cd = df.groupby("product_category").size().reset_index(name="count")
            fig7 = px.pie(cd, values="count", names="product_category", title="Sales Volume by Category", hole=0.55, color_discrete_sequence=["#2E6FD8","#4D8AE8","#7ADFA0","#8BACD8","#E08090","#1A3D80","#3D9A6C","#4A9AC8","#9B9070","#A8C4F0"])
            styled_fig(fig7); st.plotly_chart(fig7, width="stretch")
        with c2:
            cp2 = df.groupby("product_category").agg(profit=("profit","sum"), revenue=("selling_price","sum")).reset_index()
            cp2["margin"] = (cp2["profit"]/cp2["revenue"]*100).round(1)
            fig8 = px.bar(cp2, x="product_category", y="profit", title="Profit by Category", color="margin", color_continuous_scale=[[0,"#070C18"],[1,"#7ADFA0"]])
            styled_fig(fig8); st.plotly_chart(fig8, width="stretch")
        cm = df.groupby(["month","product_category"])["selling_price"].sum().unstack(fill_value=0)
        if not cm.empty:
            fig9 = px.imshow(cm.T, title="Category × Month Heatmap", color_continuous_scale=[[0,"#070C18"],[0.4,"#1A3D80"],[1,"#2E6FD8"]], aspect="auto")
            styled_fig(fig9, 300); st.plotly_chart(fig9, width="stretch")

    with t4:
        c1, c2 = st.columns(2)
        with c1:
            pm = df.groupby("payment_method").size().reset_index(name="count")
            fig10 = px.pie(pm, values="count", names="payment_method", title="Payment Method Distribution", hole=0.58, color_discrete_sequence=["#2E6FD8","#4D8AE8","#7ADFA0","#8BACD8","#1A3D80","#E08090"])
            styled_fig(fig10); st.plotly_chart(fig10, width="stretch")
        with c2:
            ps = df.groupby("payment_received").agg(count=("id","count"), total=("pending_amount","sum")).reset_index()
            ps["label"] = ps["payment_received"].map({0:"Pending",1:"Received"})
            fig11 = px.bar(ps, x="label", y="count", title="Payment Status", color="label", color_discrete_map={"Pending":"#2E6FD8","Received":"#7ADFA0"})
            styled_fig(fig11); st.plotly_chart(fig11, width="stretch")
        aged = df[df["pending_amount"] > 0].copy()
        if not aged.empty:
            today_ts = pd.Timestamp(date.today())
            aged["days"] = (today_ts - aged["sale_date"]).dt.days
            aged["bucket"] = pd.cut(aged["days"], bins=[0,7,15,30,60,9999], labels=["0–7d","8–15d","16–30d","31–60d","60d+"])
            ag = aged.groupby("bucket", observed=True)["pending_amount"].sum().reset_index()
            fig12 = px.bar(ag, x="bucket", y="pending_amount", title="Pending — Aging Buckets", color="pending_amount", color_continuous_scale=[[0,"#2E6FD8"],[1,"#C05060"]])
            styled_fig(fig12); st.plotly_chart(fig12, width="stretch")
        else:
            st.success("No pending payments.")

    with t5:
        c1, c2 = st.columns(2)
        with c1:
            if "vendor" in df.columns:
                vd = (df[df["vendor"].astype(str).str.strip() != ""].groupby("vendor").agg(revenue=("selling_price","sum"), items=("id","count")).nlargest(10,"revenue").reset_index())
                if not vd.empty:
                    fig13 = px.bar(vd, x="revenue", y="vendor", orientation="h", title="Top Vendors by Revenue", color="revenue", color_continuous_scale=[[0,"#070C18"],[1,"#2E6FD8"]])
                    styled_fig(fig13); fig13.update_layout(yaxis=dict(autorange="reversed")); st.plotly_chart(fig13, width="stretch")
                else:
                    st.info("Add vendor names to see this chart.")
        with c2:
            if "product_description" in df.columns:
                pd2 = df[df["product_description"].astype(str).str.strip() != ""].copy()
                if not pd2.empty:
                    tm = (pd2.groupby("product_description").agg(margin=("margin","mean"), revenue=("selling_price","sum")).nlargest(10,"margin").reset_index())
                    tm["product_description"] = tm["product_description"].str[:30]
                    fig14 = px.bar(tm, x="margin", y="product_description", orientation="h", title="Top Products by Margin %", color="margin", color_continuous_scale=[[0,"#070C18"],[1,"#7ADFA0"]])
                    styled_fig(fig14); fig14.update_layout(yaxis=dict(autorange="reversed")); st.plotly_chart(fig14, width="stretch")
                else:
                    st.info("Add product descriptions to see this chart.")

    analytics_context = "\n\n".join([
        "Monthly summary:\n" + df_for_ai(df.groupby("month").agg(revenue=("selling_price","sum"), profit=("profit","sum"), sales=("id","count")).reset_index(), limit=30),
        "Category summary:\n" + df_for_ai(df.groupby("product_category").agg(revenue=("selling_price","sum"), profit=("profit","sum"), sales=("id","count")).reset_index(), limit=30),
        "Payment summary:\n" + df_for_ai(df.groupby("payment_method").agg(count=("id","count"), revenue=("selling_price","sum"), pending=("pending_amount","sum")).reset_index(), limit=30),
    ])
    render_ai_panel(
        "AI Analytics Insights",
        analytics_context,
        "analytics_ai",
        "Explain the strongest business insights from these analytics and give practical next actions.",
    )


def page_reminders():
    page_header("Reminders", "Payment Follow-ups")
    df = fetch_all()
    if df.empty:
        st.info("No data available.")
        return

    today_ts  = pd.Timestamp(date.today())
    df["days_old"] = (today_ts - df["sale_date"]).dt.days
    overdue_count = len(df[(df["pending_amount"] > 0) & (df["days_old"] > 30)])
    flagged_count = int((df["delay_status"] == 1).sum())

    if overdue_count or flagged_count:
        bc = st.columns(2)
        if overdue_count: bc[0].error(f"{overdue_count} payments overdue (30+ days)")
        if flagged_count: bc[1].warning(f"{flagged_count} transactions flagged")
    else:
        st.success("All clear — no overdue or flagged payments.")

    rule_sm()
    t1, t2, t3, t4 = st.tabs(["Overdue (30d+)","Flagged","High Value","Upcoming"])

    with t1:
        ov = df[(df["pending_amount"] > 0) & (df["days_old"] > 30)].sort_values("days_old", ascending=False)
        if ov.empty:
            st.success("No overdue payments.")
        else:
            st.warning(f"{len(ov)} overdue — ₹{ov['pending_amount'].sum():,.0f} total")
            for _, r in ov.iterrows():
                with st.expander(f"{r['customer_name']}  ·  ₹{r['pending_amount']:,.0f}  ·  {int(r['days_old'])} days"):
                    row_id = int(r["id"])
                    ca, cb, cc, cd, ce = st.columns([2,2,1,1,1])
                    ca.write(r["sale_date"].strftime("%d %b %Y"))
                    cb.write(r.get("product_category","—"))
                    with cc:
                        if st.button("Mark Paid", key=f"op_{row_id}", width="stretch"):
                            st.session_state.overdue_payment_editor_id = row_id
                            st.rerun()
                    with cd:
                        if st.button("Remind", key=f"or_{row_id}", width="stretch"):
                            st.toast(f"Reminder noted for {r['customer_name']}.")
                    with ce:
                        render_customer_bill_download(df, r["customer_name"], key=f"overdue_bill_{row_id}", label="Bill PDF")
                    if st.session_state.get("overdue_payment_editor_id") == row_id:
                        st.markdown("<div class='pay-form-note'>Enter full or partial payment details.</div>", unsafe_allow_html=True)
                        with st.form(f"overdue_payment_form_{row_id}"):
                            pc1, pc2, pc3, pc4 = st.columns([1, 1, 1, 1.2])
                            with pc1:
                                _, payment_amount, payment_ok = currency_input("Amount Received (₹)", f"overdue_payment_amount_{row_id}", money_value(r.get("pending_amount")))
                            with pc2:
                                payment_date = st.date_input("Paid Date", value=date.today(), key=f"overdue_payment_date_{row_id}")
                            with pc3:
                                payment_method = st.selectbox("Paid By", PAYMENT_COLLECTION_METHODS, key=f"overdue_payment_method_{row_id}")
                            with pc4:
                                received_by = st.text_input("Received By", value=default_receiver_name(), key=f"overdue_payment_received_by_{row_id}")
                            save_col, cancel_col = st.columns(2)
                            with save_col:
                                save_payment = st.form_submit_button("Save Payment", width="stretch")
                            with cancel_col:
                                cancel_payment = st.form_submit_button("Cancel", width="stretch")
                            if save_payment:
                                if not payment_ok:
                                    st.error("Payment amount must be a valid number.")
                                else:
                                    ok, message = record_payment(r, payment_amount, payment_date, payment_method, received_by)
                                    if ok:
                                        st.session_state.overdue_payment_editor_id = None
                                        st.success(message)
                                        st.rerun()
                                    else:
                                        st.error(message)
                            if cancel_payment:
                                st.session_state.overdue_payment_editor_id = None
                                st.rerun()

    with t2:
        dl = df[df["delay_status"] == 1].sort_values("pending_amount", ascending=False)
        if dl.empty:
            st.success("No flagged payments.")
        else:
            st.error(f"{len(dl)} flagged — ₹{dl['pending_amount'].sum():,.0f}")
            show = dl[["customer_name","sale_date","product_category","selling_price","pending_amount","days_old"]].copy()
            show["sale_date"] = show["sale_date"].dt.strftime("%d %b %Y")
            show.columns = ["Customer","Date","Category","Amount ₹","Pending ₹","Days Old"]
            st.dataframe(show, width="stretch", hide_index=True)
            sc = st.selectbox("Clear flag for:", dl["id"].tolist(), format_func=lambda x: f"#{x} — {dl[dl['id']==x]['customer_name'].values[0]}")
            if st.button("Clear Flag"):
                get_col().update_one({"id": sc}, {"$set": {"delay_status":0}})
                invalidate_cache(); st.success("Flag cleared."); st.rerun()

    with t3:
        hv = df[df["selling_price"] >= 10000].sort_values("selling_price", ascending=False).head(20).copy()
        if hv.empty:
            st.info("No high-value sales (₹10,000+) yet.")
        else:
            hv["sale_date"]        = hv["sale_date"].dt.strftime("%d %b %Y")
            hv["payment_received"] = hv["payment_received"].map({0:"Pending",1:"Paid"})
            show = hv[["customer_name","sale_date","product_category","selling_price","profit","payment_received"]].copy()
            show.columns = ["Customer","Date","Category","Amount ₹","Profit ₹","Status"]
            st.dataframe(show, width="stretch", hide_index=True)

    with t4:
        soon = df[(df["pending_amount"] > 0) & (df["days_old"] >= 7) & (df["days_old"] <= 30) & (df["delay_status"] == 0)].sort_values("days_old", ascending=False)
        if soon.empty:
            st.info("No follow-ups needed in the 7–30 day window.")
        else:
            st.info(f"{len(soon)} sales with pending payments between 7–30 days old.")
            show = soon[["customer_name","customer_phone","sale_date","product_category","pending_amount","days_old"]].copy()
            show["sale_date"] = show["sale_date"].dt.strftime("%d %b %Y")
            show.columns = ["Customer","Phone","Date","Category","Pending ₹","Days Old"]
            st.dataframe(show, width="stretch", hide_index=True)
            sec("Bill PDFs")
            for customer in sorted(soon["customer_name"].dropna().astype(str).unique(), key=str.casefold):
                c1, c2 = st.columns([3, 1])
                pending_total = float(df[df["customer_name"].astype(str).eq(customer)]["pending_amount"].map(money_value).sum())
                c1.write(f"**{customer}** · ₹{pending_total:,.2f} pending")
                with c2:
                    render_customer_bill_download(df, customer, key=f"upcoming_bill_{re.sub(r'[^0-9A-Za-z]+', '_', customer)}", label="Bill PDF")

    reminder_context = "\n\n".join([
        "Overdue rows:\n" + df_for_ai(df[(df["pending_amount"] > 0) & (df["days_old"] > 30)].sort_values("days_old", ascending=False), ["customer_name","customer_phone","sale_date","product_category","pending_amount","days_old"], 40),
        "Upcoming rows:\n" + df_for_ai(df[(df["pending_amount"] > 0) & (df["days_old"] >= 7) & (df["days_old"] <= 30)].sort_values("days_old", ascending=False), ["customer_name","customer_phone","sale_date","product_category","pending_amount","days_old"], 40),
    ])
    render_ai_panel(
        "AI Follow-up Planner",
        reminder_context,
        "reminders_ai",
        "Prioritize payment follow-ups and draft short WhatsApp reminder message templates.",
    )


def page_inventory():
    page_header("Inventory", "Stock Management")
    inv_col = get_db()["inventory"]
    t1, t2 = st.tabs(["Current Stock","Add / Update Stock"])

    with t1:
        items = list(inv_col.find({}, {"_id":0}))
        if not items:
            st.markdown("<div class='empty'><div class='empty-glyph'>◆</div><div>No inventory items yet.</div></div>", unsafe_allow_html=True)
        else:
            inv_df = pd.DataFrame(items)
            total_value  = (inv_df.get("quantity", pd.Series([0])) * inv_df.get("cost_price", pd.Series([0]))).sum()
            low_stock    = inv_df[inv_df.get("quantity", pd.Series([0])) <= inv_df.get("min_stock", pd.Series([5]))]
            out_of_stock = inv_df[inv_df.get("quantity", pd.Series([0])) == 0]
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total SKUs",      len(inv_df))
            m2.metric("Inventory Value", f"₹{total_value:,.0f}")
            m3.metric("Low Stock",       len(low_stock))
            m4.metric("Out of Stock",    len(out_of_stock))
            if not low_stock.empty: st.warning(f"{len(low_stock)} item(s) running low.")
            rule_sm()
            inventory_filter_options = ["All"] + CATEGORIES
            inventory_default = app_pref("default_inventory_category", "All")
            if inventory_default not in inventory_filter_options:
                inventory_default = "All"
            cat_f = st.selectbox(
                "Filter by Category",
                inventory_filter_options,
                index=inventory_filter_options.index(inventory_default),
                key="inventory_category_filter",
            )
            view  = inv_df.copy()
            if cat_f != "All" and "category" in view.columns: view = view[view["category"] == cat_f]
            if "quantity" in view.columns and "min_stock" in view.columns:
                view["Status"] = view.apply(lambda r: "Out of Stock" if r["quantity"]==0 else ("Low Stock" if r["quantity"]<=r["min_stock"] else "OK"), axis=1)
            st.dataframe(view, width="stretch", hide_index=True)
            if "category" in inv_df.columns and "quantity" in inv_df.columns:
                cat_stock = inv_df.groupby("category")["quantity"].sum().reset_index()
                fig = px.bar(cat_stock, x="category", y="quantity", title="Stock by Category", color="quantity", color_continuous_scale=[[0,"#C05060"],[0.4,"#2E6FD8"],[1,"#7ADFA0"]])
                styled_fig(fig, 260); st.plotly_chart(fig, width="stretch")

    with t2:
        sec("Add or Update Stock Item")
        with st.form("inv_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                item_name = st.text_input("Item Name *", placeholder="e.g. Banarasi Silk Saree")
                item_sku  = st.text_input("SKU / Code",  placeholder="e.g. SAR-001")
                item_cat  = st.selectbox("Category", CATEGORIES)
                item_vend = st.text_input("Vendor")
            with c2:
                item_qty  = st.number_input("Quantity *",       min_value=0, step=1)
                item_min  = st.number_input("Min Stock Alert",  min_value=0, step=1, value=5)
                item_cost = st.number_input("Cost Price (₹) *", min_value=0.0, step=50.0, format="%.2f")
                item_mrp  = st.number_input("Selling Price (₹)",min_value=0.0, step=50.0, format="%.2f")
            item_notes = st.text_area("Notes", height=55)
            if st.form_submit_button("Save Item", width="stretch"):
                if not item_name.strip():
                    st.error("Item name is required.")
                else:
                    inv_col.update_one(
                        {"sku": item_sku.strip() or item_name.strip()},
                        {"$set": {"name":item_name.strip(),"sku":item_sku.strip(),"category":item_cat,"vendor":item_vend.strip(),"quantity":item_qty,"min_stock":item_min,"cost_price":round(item_cost,2),"sell_price":round(item_mrp,2),"notes":item_notes.strip(),"updated_at":str(datetime.now())}},
                        upsert=True,
                    )
                    st.success(f"'{item_name.strip()}' saved to inventory.")
                    st.rerun()

    inv_docs = list(inv_col.find({}, {"_id": 0}))
    sales_df = fetch_all()
    inventory_context = "\n\n".join([
        "Inventory:\n" + df_for_ai(pd.DataFrame(inv_docs) if inv_docs else pd.DataFrame(), limit=120),
        "Recent sales demand:\n" + df_for_ai(
            sales_df.sort_values("sale_date", ascending=False) if not sales_df.empty else sales_df,
            ["sale_date", "vendor", "product_category", "product_description", "quantity", "selling_price", "profit", "pending_amount"],
            80,
        ),
    ])
    render_ai_panel(
        "AI Inventory Planner",
        inventory_context,
        "inventory_ai",
        "Suggest what to restock, what is slow-moving, which vendors/categories deserve attention, and any pricing or margin actions.",
    )

# =====================================================
# BACKUP & RESTORE PAGE
# =====================================================

def page_backup_restore():
    _, settings_col = st.columns([0.78, 0.22])
    with settings_col:
        render_top_settings()

    # ── Header ────────────────────────────────────────────────────────────
    st.markdown("""
    <div class='bk-header'>
        <span class='bk-header-icon'>🗄️</span>
        <div class='bk-header-title'>Backup &amp; Restore</div>
        <div class='bk-header-sub'>Database Checkpoint Management · Shree Krishna Boutique</div>
    </div>
    """, unsafe_allow_html=True)

    # ── Last backup timestamp from session state ──────────────────────────
    last_backup_ts = st.session_state.get("last_backup_ts", None)
    last_restore_ts = st.session_state.get("last_restore_ts", None)

    # ══════════════════════════════════════════════════════════
    # ROW 1 — Backup & Restore cards side by side
    # ══════════════════════════════════════════════════════════
    col_bk, col_re = st.columns(2, gap="large")

    # ── Backup card ───────────────────────────────────────────────────────
    with col_bk:
        st.markdown("""
        <div class='bk-card' style='animation-delay:0s'>
            <div class='bk-card-icon bk-icon-blue'>💾</div>
            <div class='bk-card-title'>Backup Database</div>
            <div class='bk-card-desc'>
                Export a complete checkpoint of all sales, customers, and inventory data.
                Download as Excel for safekeeping or migration.
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        df = fetch_all()

        # CSV checkpoint
        csv_data = df.assign(sale_date=df["sale_date"].astype(str)).to_csv(index=False) if not df.empty else "No data"
        st.download_button(
            label="⬇️  Download CSV Checkpoint",
            data=csv_data,
            file_name=f"boutique_checkpoint_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            width="stretch",
        )
        st.caption("Includes all sales, customer, and payment records")

        st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)

        # Excel checkpoint
        if not df.empty:
            excel_data = to_excel(df)
            ts_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            st.download_button(
                label="⬇️  Download Excel Checkpoint",
                data=excel_data,
                file_name=f"boutique_checkpoint_{ts_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )
            st.caption("Formatted spreadsheet with column headers and styling")
        else:
            st.info("No data to export yet.")

        if last_backup_ts:
            st.markdown(f"<div class='bk-ts'>Last export: {last_backup_ts}</div>", unsafe_allow_html=True)

        if st.button("📋  Record Manual Backup Note", width="stretch"):
            ts = datetime.now().strftime("%d %b %Y, %I:%M %p")
            st.session_state.last_backup_ts = ts
            st.success(f"✓ Manual backup noted at {ts}")
            st.rerun()

    # ── Restore card ──────────────────────────────────────────────────────
    with col_re:
        st.markdown("""
        <div class='bk-card' style='animation-delay:0.1s'>
            <div class='bk-card-icon bk-icon-green'>♻️</div>
            <div class='bk-card-title'>Restore from Checkpoint</div>
            <div class='bk-card-desc'>
                Upload a previously exported CSV checkpoint to restore records.
                Existing data will be preserved — only new records are added.
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        uploaded = st.file_uploader(
            "Upload Checkpoint File",
            type=["csv"],
            help="Upload a CSV file previously exported from this application",
            label_visibility="visible",
        )
        st.caption("200 MB max · CSV format only")

        if last_restore_ts:
            st.markdown(
                f"<span class='bk-status-badge bk-status-ok'>✓ Last restored: {last_restore_ts}</span>",
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)

        if uploaded is not None:
            try:
                restore_df = pd.read_csv(uploaded)
                row_count = len(restore_df)
                col_count = len(restore_df.columns)

                st.markdown(f"""
                <div class='bk-card' style='animation-delay:0.15s; border-color: rgba(37,99,235,0.3); margin-top:0.8rem'>
                    <div class='bk-card-title' style='font-size:0.85rem'>📊 File Preview</div>
                    <div class='bk-card-desc'>
                        <b>{row_count}</b> records · <b>{col_count}</b> columns detected<br>
                        Columns: {', '.join(restore_df.columns[:6].tolist())}{' …' if col_count > 6 else ''}
                    </div>
                </div>
                """, unsafe_allow_html=True)

                st.dataframe(restore_df.head(5), width="stretch", hide_index=True)

                st.warning(
                    "⚠️  This will **insert** records from the checkpoint into the live database. "
                    "Duplicate entries may result if records already exist.",
                )

                confirm = st.checkbox("I understand — proceed with restore")
                if confirm:
                    if st.button("🔄  Restore Database from Checkpoint", width="stretch"):
                        progress_placeholder = st.empty()
                        progress_placeholder.markdown(
                            "<div style='background:var(--bg-2);border-radius:999px;overflow:hidden;height:6px;margin:0.5rem 0'>"
                            "<div class='bk-progress-bar'></div></div>",
                            unsafe_allow_html=True,
                        )
                        status_msg = st.empty()
                        status_msg.info("Restoring records…")

                        # ── Actual restore logic ──────────────────────────
                        inserted = 0
                        skipped  = 0
                        errors   = []

                        # ── Column name normalisation ─────────────────────
                        # The app exports title-cased columns (e.g. "Customer Name")
                        # but internally uses snake_case (e.g. "customer_name").
                        # Build a mapping: title-cased export name → internal name.
                        EXPORT_TO_INTERNAL = {
                            "Id":                  "id",
                            "Customer Name":       "customer_name",
                            "Customer Phone":      "customer_phone",
                            "Sale Date":           "sale_date",
                            "Product Category":    "product_category",
                            "Product Description": "product_description",
                            "Vendor":              "vendor",
                            "Buying Price":        "buying_price",
                            "Selling Price":       "selling_price",
                            "Profit":              "profit",
                            "Profit Margin %":     "margin",
                            "Amount Paid":         "amount_paid",
                            "Pending Amount":      "pending_amount",
                            "Payment Status":      "_payment_status_str",  # converted below
                            "Status":              "_payment_status_str",
                            "Delayed":             "_delayed_str",          # converted below
                            "Payment Method":      "payment_method",
                            "Notes":               "notes",
                            "Created At":          "created_at",
                        }
                        # Rename columns that match the export format; leave unknown ones as-is
                        restore_df = restore_df.rename(
                            columns={k: v for k, v in EXPORT_TO_INTERNAL.items() if k in restore_df.columns}
                        )
                        # Also lowercase any remaining columns that weren't renamed
                        restore_df.columns = [
                            c.lower().replace(" ", "_") if c not in restore_df.columns else c
                            for c in restore_df.columns
                        ]

                        required_cols = {"customer_name", "selling_price"}
                        if not required_cols.issubset(set(restore_df.columns)):
                            progress_placeholder.empty()
                            status_msg.empty()
                            st.error(
                                f"Invalid checkpoint file. Required columns missing: "
                                f"{required_cols - set(restore_df.columns)}"
                            )
                        else:
                            for _, row in restore_df.iterrows():
                                try:
                                    doc = row.dropna().to_dict()

                                    # Convert "Payment Status" string → payment_received int
                                    if "_payment_status_str" in doc:
                                        ps = str(doc.pop("_payment_status_str")).strip().lower()
                                        doc["payment_received"] = 1 if ps in ("paid", "received", "1") else 0

                                    # Convert "Delayed" string → delay_status int
                                    if "_delayed_str" in doc:
                                        dl = str(doc.pop("_delayed_str")).strip().lower()
                                        doc["delay_status"] = 1 if dl in ("yes", "true", "1") else 0

                                    # Derive payment_received from pending_amount if not set
                                    if "payment_received" not in doc:
                                        pending = float(doc.get("pending_amount", 0) or 0)
                                        doc["payment_received"] = 0 if pending > 0 else 1

                                    # Default delay_status
                                    if "delay_status" not in doc:
                                        doc["delay_status"] = 0

                                    # Normalise numeric types
                                    for num_col in ["buying_price", "selling_price", "amount_paid", "pending_amount", "quantity", "profit", "margin"]:
                                        if num_col in doc:
                                            try:
                                                doc[num_col] = float(doc[num_col])
                                            except (ValueError, TypeError):
                                                doc.pop(num_col, None)
                                    for int_col in ["payment_received", "delay_status"]:
                                        if int_col in doc:
                                            doc[int_col] = int(doc[int_col])

                                    # Drop the old exported id — assign a fresh one
                                    doc.pop("id", None)
                                    doc["id"]          = get_next_id()
                                    doc["restored_at"] = str(datetime.now())
                                    get_col().insert_one(doc)
                                    inserted += 1
                                except Exception as e:
                                    skipped += 1
                                    errors.append(str(e))

                            invalidate_cache()
                            progress_placeholder.empty()
                            status_msg.empty()
                            ts = datetime.now().strftime("%d %b %Y, %I:%M %p")
                            st.session_state.last_restore_ts = ts

                            if inserted > 0:
                                st.success(
                                    f"✅  Restore complete — **{inserted}** records inserted"
                                    + (f", {skipped} skipped." if skipped else ".")
                                )
                            if errors:
                                with st.expander(f"⚠️  {len(errors)} row(s) had errors"):
                                    for err in errors[:10]:
                                        st.caption(err)
                            st.rerun()

            except Exception as e:
                st.error(f"Could not read file: {e}")

    # ══════════════════════════════════════════════════════════
    # ROW 2 — Database stats card
    # ══════════════════════════════════════════════════════════
    rule()
    st.markdown("""
    <div class='bk-card' style='animation-delay:0.2s'>
        <div class='bk-card-title'>📈  Current Database Status</div>
        <div class='bk-card-desc'>Live snapshot of records in the database</div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    df2 = fetch_all()
    m = metrics(df2)
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total Records",  m["sales"])
    c2.metric("Total Revenue",  f"₹{m['revenue']:,.0f}")
    c3.metric("Unique Customers", m["customers"])
    c4.metric("Pending Payments", f"₹{m['pending']:,.0f}")
    c5.metric("Data Health",    "✓ Live" if m["sales"] > 0 else "Empty")

    rule_sm()
    backup_context = "\n\n".join([
        f"Last backup noted: {last_backup_ts or 'none'}",
        f"Last restore noted: {last_restore_ts or 'none'}",
        f"Current metrics: sales={m['sales']}, revenue={m['revenue']:.2f}, profit={m['profit']:.2f}, pending={m['pending']:.2f}, customers={m['customers']}",
        "Recent records:\n" + df_for_ai(
            df2.sort_values("sale_date", ascending=False) if not df2.empty and "sale_date" in df2.columns else df2,
            ["id", "customer_name", "sale_date", "vendor", "product_category", "selling_price", "amount_paid", "pending_amount", "payment_received", "created_at"],
            80,
        ),
    ])
    render_ai_panel(
        "AI Backup & Data Health Audit",
        backup_context,
        "backup_ai",
        "Check backup freshness, data quality risks, duplicate/odd-looking records, and the safest next backup or restore action.",
    )

    rule_sm()
    st.caption(f"Database last queried: {datetime.now().strftime('%d %b %Y, %I:%M:%S %p')}  ·  Boutique Manager v2.0")

def page_work_notes():
    page_header("Work Notes", "Manual Last Edited Log")

    notes = get_work_notes()
    latest = notes[0] if notes else None
    c1, c2, c3 = st.columns(3)
    c1.metric("Last Noted Date", latest.get("work_date", "—") if latest else "—")
    c2.metric("Total Notes", len(notes))
    c3.metric("Last Saved By", latest.get("created_by", "—") if latest else "—")

    rule_sm()
    sec("Add Note")
    with st.form("work_note_form"):
        d1, d2 = st.columns([1, 2])
        with d1:
            work_date = st.date_input("Date You Worked", value=date.today(), key="work_note_date")
        with d2:
            note = st.text_input("Note", placeholder="Example: Updated accounts / checked passbook / entered pending payments", key="work_note_text")
        save_note = st.form_submit_button("Save Note", width="stretch")

    if save_note:
        save_work_note(work_date, note)
        st.success(f"Saved note for {work_date}.")
        st.rerun()

    rule_sm()
    sec("History")
    notes = get_work_notes()
    if not notes:
        st.info("No work notes saved yet.")
        return

    note_df = pd.DataFrame(notes)
    show = note_df[["id", "work_date", "note", "created_at", "created_by"]].copy()
    show["created_at"] = pd.to_datetime(show["created_at"], errors="coerce").dt.strftime("%d %b %Y, %I:%M %p").fillna(show["created_at"])
    show.columns = ["ID", "Date", "Note", "Saved On", "Saved By"]
    st.dataframe(show, width="stretch", hide_index=True, height=420)

    d1, d2 = st.columns([1, 3])
    with d1:
        delete_id = st.selectbox("Delete Note ID", show["ID"].tolist(), key="work_note_delete_id")
    with d2:
        st.caption("Delete only removes this manual note. It does not affect sales, bills, vendors, or passbook data.")
        if st.button("Delete Selected Note", key="work_note_delete", width="stretch"):
            delete_work_note(int(delete_id))
            st.success(f"Deleted note #{delete_id}.")
            st.rerun()

    render_ai_panel(
        "AI Work Notes Summary",
        "Work notes:\n" + show.to_csv(index=False),
        "work_notes_ai",
        "Summarize what work was done recently and suggest the next bookkeeping actions.",
    )

def build_ai_business_context() -> str:
    df = fetch_all()
    m = metrics(df)
    parts = [
        f"Today: {date.today()}",
        f"Metrics: sales={m['sales']}, revenue={m['revenue']:.2f}, profit={m['profit']:.2f}, pending={m['pending']:.2f}, customers={m['customers']}",
    ]
    if not df.empty:
        recent_cols = ["id", "customer_name", "customer_phone", "sale_date", "vendor", "product_category", "selling_price", "amount_paid", "pending_amount", "payment_method", "last_payment_date"]
        parts.append("Recent sales:\n" + df_for_ai(df.sort_values("sale_date", ascending=False), recent_cols, 30))
        pending = df[df["pending_amount"].map(money_value) > 0].sort_values("pending_amount", ascending=False)
        parts.append("Pending sales:\n" + df_for_ai(pending, recent_cols, 30))
        customer_summary = (df.groupby("customer_name").agg(
            transactions=("id", "count"),
            total_spent=("selling_price", "sum"),
            total_pending=("pending_amount", "sum"),
            last_sale=("sale_date", "max"),
        ).reset_index().sort_values("total_pending", ascending=False))
        parts.append("Customer summary:\n" + df_for_ai(customer_summary, limit=30))
    try:
        notes = get_work_notes(20)
        if notes:
            parts.append("Recent work notes:\n" + pd.DataFrame(notes).to_csv(index=False))
    except Exception:
        pass
    try:
        inv_docs = list(get_db()["inventory"].find({}, {"_id": 0}).limit(120))
        if inv_docs:
            parts.append("Inventory:\n" + df_for_ai(pd.DataFrame(inv_docs), limit=120))
    except Exception:
        pass
    try:
        bills = list(bill_history_collection().find({}, {"_id": 0}).sort("generated_at", -1).limit(50))
        if bills:
            parts.append("Recent bills:\n" + df_for_ai(pd.DataFrame(bills), limit=50))
    except Exception:
        pass
    try:
        keys = list(auth_keys_collection().find({}, {"_id": 0, "public_key_hash": 0}).sort("created_at", -1).limit(50))
        invites = list(auth_qr_collection().find({}, {"_id": 0, "secret_hash": 0, "pin_hash": 0}).sort("created_at", -1).limit(50))
        devices = list(auth_devices_collection().find({}, {"_id": 0, "user_agent": 0}).sort("last_login_at", -1).limit(50))
        if keys:
            parts.append("Encrypted key logins:\n" + df_for_ai(pd.DataFrame(keys), limit=50))
        if invites:
            parts.append("Temporary QR invites:\n" + df_for_ai(pd.DataFrame(invites), limit=50))
        if devices:
            parts.append("Login devices:\n" + df_for_ai(pd.DataFrame(devices), limit=50))
    except Exception:
        pass
    return "\n\n".join(parts)

def page_ai_assistant():
    page_header("AI Assistant", "Ask About Boutique Data")
    if not llm_is_configured():
        st.info(ai_setup_message())
        return

    context = build_ai_business_context()
    quick = st.selectbox("Quick Question", [
        "Summarize today's business status and what needs attention.",
        "List customers with the most pending amount and suggest follow-up priority.",
        "Find sales or vendors that look unusual.",
        "Draft polite payment reminder messages for pending customers.",
        "Summarize recent work notes and next actions.",
        "Review inventory and suggest restock or slow-moving item actions.",
        "Review login security and suggest keys/devices/QR invites to revoke.",
        "Audit backup readiness and data quality risks.",
        "Custom question",
    ])
    default_task = "" if quick == "Custom question" else quick
    question = st.text_area("Question", value=default_task, height=120, key="global_ai_question")
    if st.button("Ask AI", key="global_ai_run", width="stretch"):
        try:
            with st.spinner("Thinking..."):
                st.session_state.global_ai_answer = ask_llm(question, context)
        except Exception as exc:
            st.error(str(exc))
    if st.session_state.get("global_ai_answer"):
        rule_sm()
        st.markdown(st.session_state.global_ai_answer)


# =====================================================
# MAIN
# =====================================================

def logout_current_session():
    device_id = st.session_state.get("auth_device_id")
    if device_id:
        auth_devices_collection().update_one(
            {"_id": device_id},
            {"$set": {"active": False, "logged_out_at": datetime.now()}},
        )
    _clear_auth_state()

def main():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
    if "theme" not in st.session_state:
        st.session_state.theme = "light"
    if st.session_state.logged_in and "user_role" not in st.session_state:
        st.session_state.user_role = "admin"
    if st.session_state.logged_in:
        apply_persistent_app_settings()

    # Apply light mode CSS overrides if needed
    inject_theme()
    ensure_auth_indexes()

    if not st.session_state.logged_in:
        render_admin_login_strip()
        page_add_sale(public=True)
        return

    if _current_auth_device_revoked():
        st.warning("This login device has been revoked.")
        _clear_auth_state()
        st.rerun()

    page = sidebar()

    if   "Dashboard"   in page: page_dashboard()
    elif "Add Sale"    in page: page_add_sale(public=False)
    elif "Review"      in page: page_review()
    elif "Update"      in page: page_update()
    elif "Customer"    in page: page_customers()
    elif "Vendor List" in page: page_vendors()
    elif "Analytics"   in page: page_analytics()
    elif "Reminders"   in page: page_reminders()
    elif "Generate Bill" in page: page_generate_bill()
    elif "Passbook Reader" in page: page_passbook_reader()
    elif "Work Notes"  in page: page_work_notes()
    elif "AI Assistant" in page: page_ai_assistant()
    elif "Technical"   in page: page_technical()
    elif "Security"    in page: page_security_devices()
    elif "Backup"      in page: page_backup_restore()
    elif "Logout"      in page:
        logout_current_session()
        st.rerun()

if __name__ == "__main__":
    main()
